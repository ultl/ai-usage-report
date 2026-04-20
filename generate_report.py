#!/usr/bin/env python3
"""
AI Dev Journal — Best Practices & Analysis Report Generator

Reads the consolidated report .xlsx + profiles.yml, uses LLM to analyze:
1. Prompting best practices vs actual prompts (per person + team-wide)
2. C-Codex/AI tool objectives: bugs, constraints, workarounds
3. Team member subjectives: personal insights, prior experience

Outputs Markdown then converts to PDF.

Usage:
    python generate_report.py report.xlsx --profiles profiles.yml --model gpt-5.4-mini
    python generate_report.py report.xlsx --profiles profiles.yml -o report.md
"""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

try:
    import yaml
except ImportError:
    yaml = None

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")


def _call_openai(model: str, prompt: str, timeout: int = 600) -> str:
    base = (OPENAI_BASE_URL or "").rstrip("/")
    is_azure = "cognitiveservices.azure.com" in base or "openai.azure.com" in base

    if is_azure:
        host = re.sub(r"/openai(/v1)?$", "", base).rstrip("/")
        url = f"{host}/openai/deployments/{model}/chat/completions?api-version=2024-12-01-preview"
        headers = {"api-key": OPENAI_API_KEY or ""}
    else:
        url = f"{base}/chat/completions"
        headers = {}
        if OPENAI_API_KEY:
            headers["Authorization"] = f"Bearer {OPENAI_API_KEY}"

    body: dict[str, Any] = {
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.3,
        "max_completion_tokens": 12000,
    }
    if not is_azure:
        body["model"] = model

    r = requests.post(url, headers=headers, json=body, timeout=timeout)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]


def _cell_str(ws, r: int, c: int) -> str:
    v = ws.cell(r, c).value
    return str(v).strip() if v is not None else ""


def _cell_num(ws, r: int, c: int) -> float | None:
    v = ws.cell(r, c).value
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _find_section_row(ws, label: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value or "").strip().startswith(label):
            return r
    return None


def _load_profiles(path: Path | None) -> dict[str, dict[str, str]]:
    if path is None or not path.exists():
        return {}
    try:
        text = path.read_text(encoding="utf-8")
        if path.suffix.lower() in (".yml", ".yaml"):
            if yaml is None:
                print("  ⚠  PyYAML not installed", file=sys.stderr)
                return {}
            raw = yaml.safe_load(text)
        else:
            raw = json.loads(text)
        if isinstance(raw, dict):
            return {k: v for k, v in raw.items() if isinstance(v, dict)}
    except Exception as e:
        print(f"  ⚠  Failed to load profiles: {e}", file=sys.stderr)
    return {}


# --------------------------------------------------------------------------- #
# Data extraction
# --------------------------------------------------------------------------- #

def extract_data(xlsx_path: Path, profiles: dict) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)
    data: dict[str, Any] = {"profiles": profiles}

    # --- Dashboard KPIs ---
    if "📊 Dashboard" in wb.sheetnames:
        ws = wb["📊 Dashboard"]
        kpis = {}
        for r in range(5, 20):
            m = _cell_str(ws, r, 1)
            if not m:
                continue
            kpis[m] = {
                "assessed": _cell_str(ws, r, 2),
                "self_reported": _cell_str(ws, r, 3),
                "deviation": _cell_str(ws, r, 4),
            }
        data["kpis"] = kpis

        # Self-Report Accuracy
        acc_row = _find_section_row(ws, "SELF-REPORT ACCURACY")
        if acc_row:
            accuracy = []
            hr = acc_row + 1
            for r in range(hr + 1, hr + 20):
                staff = _cell_str(ws, r, 1)
                if not staff or staff.upper().startswith("TOTAL"):
                    break
                accuracy.append({
                    "staff": staff,
                    "assessed_saved": _cell_num(ws, r, 8),
                    "self_rep_saved": _cell_num(ws, r, 9),
                    "delta_saved": _cell_num(ws, r, 10),
                    "accuracy": _cell_str(ws, r, 14),
                    "verdict": _cell_str(ws, r, 15),
                })
            data["accuracy"] = accuracy

    # --- AI Lesson Compare (prompts + lessons) ---
    if "🤖 AI Lesson Compare" in wb.sheetnames:
        ws = wb["🤖 AI Lesson Compare"]
        lessons = []
        for r in range(5, ws.max_row + 1):
            staff = _cell_str(ws, r, 1)
            if not staff:
                break
            lessons.append({
                "staff": staff,
                "title": _cell_str(ws, r, 3),
                "tool": _cell_str(ws, r, 4),
                "task_desc": _cell_str(ws, r, 5)[:300],
                "user_prompt": _cell_str(ws, r, 6)[:500],
                "result": _cell_str(ws, r, 7)[:300],
                "user_lesson": _cell_str(ws, r, 8)[:300],
                "ai_lesson": _cell_str(ws, r, 9)[:300],
                "comparison": _cell_str(ws, r, 10),
                "suggested_prompt": _cell_str(ws, r, ws.max_column)[:500],
            })
        data["lessons"] = lessons

    # --- Error Charts ---
    if "🏷️ Error Charts" in wb.sheetnames:
        ws = wb["🏷️ Error Charts"]
        errors = []
        for r in range(6, 20):
            label = _cell_str(ws, r, 1)
            count = _cell_num(ws, r, 2)
            desc = _cell_str(ws, r, 3)
            if label and count:
                errors.append({"label": label, "count": int(count), "description": desc})
        data["errors"] = errors

    # --- Raw Log ---
    if "📝 Raw Log" in wb.sheetnames:
        ws = wb["📝 Raw Log"]
        raw_log = []
        for r in range(5, min(ws.max_row + 1, 50)):
            staff = _cell_str(ws, r, 1)
            if not staff:
                break
            raw_log.append({
                "staff": staff,
                "title": _cell_str(ws, r, 3),
                "tool": _cell_str(ws, r, 4),
                "category": _cell_str(ws, r, 5),
                "rating": _cell_num(ws, r, 7),
            })
        data["raw_log"] = raw_log

    # --- SDLC ---
    if "🧭 SDLC Summary" in wb.sheetnames:
        ws = wb["🧭 SDLC Summary"]
        sdlc = []
        for r in range(6, 18):
            stage = _cell_str(ws, r, 1)
            count = _cell_num(ws, r, 3)
            if stage and count and count > 0:
                sdlc.append({
                    "stage": stage,
                    "task_count": int(count),
                    "assessed_pct": _cell_num(ws, r, 5),
                    "self_rep_pct": _cell_num(ws, r, 6),
                })
        data["sdlc"] = sdlc

    return data


# --------------------------------------------------------------------------- #
# Report prompt
# --------------------------------------------------------------------------- #

REPORT_PROMPT = """<role>
You are a senior engineering consultant who specializes in AI adoption, prompt engineering
best practices, and developer productivity. You write detailed, actionable reports for
engineering leadership. Bear in mind that your audience is technical but not necessarily AI experts, so explain concepts clearly and avoid jargon. Your tone is professional, constructive, and focused on growth opportunities.
</role>

<context>
This report analyzes an AI Dev Journal pilot program. {n_staff} staff members used AI tools
across {n_sessions} sessions. The data includes:

- Each staff member's profile (role, experience, tech stack, workflow, known limitations)
- Their actual prompts and AI responses
- AI-inferred lessons about what went wrong in each prompt
- Error classifications (e.g., "Clear and Format", "Missing Context")
- Self-report accuracy: how each person's subjective time estimates compare to AI's blind assessment
- SDLC stage distribution

Two perspectives are compared throughout:
- **Assessed (AI)**: Blind objective estimate of hours (AI doesn't see user's numbers)
- **Self-Reported**: User's own subjective estimate

When a user under-reports, it means they think the task is easier than it objectively is,
or they don't recognize how much AI helped them.
</context>

<staff_profiles>
{profiles_json}
</staff_profiles>

<report_data>
{data_json}
</report_data>

<instructions>
Write a professional report in Markdown with the following structure.
For EACH section, provide per-person analysis FIRST, then team-wide summary.

# AI Dev Journal — Sprint 0 Best Practices & Analysis Report

## 1. Executive Summary
- 4-5 bullet points covering the key findings across all 3 dimensions
- Include the most impactful numbers

## 2. Prompting Best Practices Analysis

### 2.1 Team-Wide Prompt Quality Overview
- What are the most common prompt errors across the team? (use error data)
- What does this tell us about the team's prompt engineering maturity?

### 2.2 Per-Person Prompt Analysis
For EACH staff member:
- **[Name] — [Role]**
  - Their most frequent prompt errors (from error classification data)
  - Specific example: what their prompt looked like vs what it should look like
    (use the suggested_prompt from AI lesson data if available)
  - What best practices they are already following well
  - Top 2-3 actionable recommendations specific to their role and workflow

### 2.3 Prompting Standards Recommendation
- Based on the team's patterns, propose a rule-based prompting standard template
  that the team should adopt (with XML structure, role assignment, context, etc.)
- Show a before/after example using a real prompt from the data

## 3. AI Tool Objectives — Bugs, Constraints & Workarounds

### 3.1 Known Limitations from Team Experience
- Extract from profiles.yml notes: what bugs, constraints, and workarounds
  each team member has documented about their AI tools
- Organize by tool (Claude Code, Copilot, C-Codex, etc.)

### 3.2 Tool-Specific Findings from Session Data
- Which tools had lowest/highest ratings and why?
- Any patterns in error types per tool?

### 3.3 Recommended Workarounds & Best Practices per Tool
- Consolidate the team's workarounds into a shared knowledge base
- Add recommendations based on the error patterns

## 4. Team Member Subjective Analysis

### 4.1 Per-Person Insights
For EACH staff member:
- **[Name] — [Role]**
  - Self-report accuracy and what it reveals about their perception
    (e.g., "thinks tasks are easier than they are", "undervalues AI contribution",
    "accurate reporter who understands AI's value")
  - Their user_lesson entries: what personal insights did they share?
  - How their background/experience affects their AI usage effectiveness
  - Do they have prior experience that helps (or misleads) them?

### 4.2 Team-Wide Patterns
- Who adapts to AI tools fastest and why?
- Common subjective biases across the team
- Correlation between experience level and self-report accuracy

## 5. Recommendations

### 5.1 Immediate Actions (This Sprint)
- 3-4 specific actions with owner and expected impact

### 5.2 Medium-Term Improvements (Next 2-3 Sprints)
- 3-4 strategic improvements

### 5.3 Training Plan
- What training each person needs based on their specific gaps
- Prioritized by impact

## 6. Appendix: Prompting Standard Template
- The complete recommended prompting template
- With placeholders and instructions for each section

Rules:
- Use SPECIFIC numbers and quotes from the data
- Reference specific staff members by name with their role
- For each person's analysis, connect their profile (experience, stack) to their behavior
- Explain WHY patterns exist, not just WHAT they are
- When discussing accuracy, explain it as: the user provides X hours but AI expects Y hours,
  meaning the user thinks the task is [easier/harder] than it objectively is
- Be constructive, not judgmental — frame gaps as growth opportunities
- Total length: 2000-3000 words
- Use tables where appropriate for comparisons
</instructions>"""


# --------------------------------------------------------------------------- #
# Markdown to PDF conversion
# --------------------------------------------------------------------------- #

_PDF_CSS = """
@page { size: A4; margin: 2cm; }
body { font-family: 'Helvetica Neue', Arial, sans-serif; font-size: 11pt;
       line-height: 1.5; color: #1a1a1a; }
h1 { color: #1F4E78; font-size: 22pt; border-bottom: 2px solid #1F4E78;
     padding-bottom: 6px; margin-top: 0; }
h2 { color: #2E75B6; font-size: 16pt; margin-top: 24pt; border-bottom: 1px solid #ddd;
     padding-bottom: 4px; }
h3 { color: #595959; font-size: 13pt; margin-top: 16pt; }
h4 { color: #1F4E78; font-size: 11pt; margin-top: 12pt; }
table { border-collapse: collapse; width: 100%; margin: 12px 0; font-size: 9.5pt; }
th { background: #1F4E78; color: white; padding: 6px 8px; text-align: left; }
td { padding: 5px 8px; border-bottom: 1px solid #ddd; }
tr:nth-child(even) td { background: #f7fbff; }
code { background: #f0f0f0; padding: 1px 4px; border-radius: 3px; font-size: 10pt; }
pre { background: #f5f5f5; padding: 12px; border-radius: 4px; overflow-x: auto;
      font-size: 9pt; }
blockquote { border-left: 3px solid #2E75B6; margin: 8px 0; padding: 4px 12px;
             color: #555; background: #f9f9f9; }
img { max-width: 100%; height: auto; margin: 8px 0; }
strong { color: #1F4E78; }
"""


def _md_to_pdf(md_text: str, pdf_path: Path, charts_dir: Path) -> bool:
    """Convert markdown to PDF. Tries weasyprint → fpdf2 → pandoc."""

    # Method 1: weasyprint (best quality)
    try:
        import markdown
        from weasyprint import HTML

        # Convert MD → HTML
        extensions = ['tables', 'fenced_code', 'nl2br']
        html_body = markdown.markdown(md_text, extensions=extensions)

        # Resolve image paths to absolute
        def _abs_img(m: re.Match) -> str:
            src = m.group(1)
            p = Path(src)
            if not p.is_absolute():
                p = charts_dir / p.name
            return f'src="file://{p}"' if p.exists() else m.group(0)

        html_body = re.sub(r'src="([^"]+)"', _abs_img, html_body)

        full_html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>{_PDF_CSS}</style></head>
<body>{html_body}</body></html>"""

        HTML(string=full_html).write_pdf(str(pdf_path))
        return True
    except ImportError:
        pass
    except Exception as e:
        print(f"  ⚠  weasyprint failed: {e}", file=sys.stderr)

    # Method 2: fpdf2 (lightweight, no external deps)
    try:
        from fpdf import FPDF
        from fpdf.enums import XPos, YPos

        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=20)
        pdf.add_page()
        pdf.set_font("Helvetica", size=10)

        for line in md_text.split("\n"):
            # Handle headers
            if line.startswith("# "):
                pdf.set_font("Helvetica", "B", 18)
                pdf.cell(0, 12, line[2:].strip(), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                pdf.set_font("Helvetica", size=10)
            elif line.startswith("## "):
                pdf.ln(4)
                pdf.set_font("Helvetica", "B", 14)
                pdf.cell(0, 10, line[3:].strip(), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                pdf.set_font("Helvetica", size=10)
            elif line.startswith("### "):
                pdf.ln(2)
                pdf.set_font("Helvetica", "B", 12)
                pdf.cell(0, 8, line[4:].strip(), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                pdf.set_font("Helvetica", size=10)
            elif line.startswith("#### "):
                pdf.set_font("Helvetica", "BI", 10)
                pdf.cell(0, 7, line[5:].strip(), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
                pdf.set_font("Helvetica", size=10)
            elif line.startswith("!["):
                # Image: ![alt](path)
                img_match = re.match(r"!\[.*?\]\((.+?)\)", line)
                if img_match:
                    img_path = Path(img_match.group(1))
                    if not img_path.is_absolute():
                        img_path = charts_dir / img_path.name
                    if img_path.exists():
                        try:
                            pdf.image(str(img_path), w=pdf.epw)
                            pdf.ln(4)
                        except Exception:
                            pass
            elif line.startswith("|") and "---" not in line:
                # Table row - simplified rendering
                cells = [c.strip() for c in line.split("|")[1:-1]]
                if cells:
                    col_w = pdf.epw / len(cells)
                    for cell in cells:
                        is_header = cell.startswith("**") or line == line.upper()
                        if is_header:
                            pdf.set_font("Helvetica", "B", 9)
                        else:
                            pdf.set_font("Helvetica", size=9)
                        pdf.cell(col_w, 6, cell.replace("**", "")[:30], border=1)
                    pdf.ln()
                    pdf.set_font("Helvetica", size=10)
            elif line.startswith("> "):
                pdf.set_font("Helvetica", "I", 9)
                pdf.multi_cell(0, 5, "  " + line[2:].strip())
                pdf.set_font("Helvetica", size=10)
            elif line.startswith("- ") or line.startswith("* "):
                pdf.cell(5, 6, "•")
                pdf.multi_cell(0, 6, line[2:].strip())
            elif re.match(r"^\d+\. ", line):
                pdf.multi_cell(0, 6, line.strip())
            elif line.strip() == "---":
                pdf.ln(4)
            elif line.strip():
                pdf.multi_cell(0, 6, line.strip())
            else:
                pdf.ln(3)

        pdf.output(str(pdf_path))
        return True
    except ImportError:
        pass
    except Exception as e:
        print(f"  ⚠  fpdf2 failed: {e}", file=sys.stderr)

    # Method 3: pandoc (external tool)
    try:
        subprocess.run(
            ["pandoc", "-f", "markdown", "-o", str(pdf_path),
             "--pdf-engine=xelatex", "-V", "geometry:margin=1in",
             "-V", "mainfont:Helvetica Neue"],
            input=md_text.encode("utf-8"),
            check=True, capture_output=True, timeout=120)
        return True
    except (FileNotFoundError, subprocess.CalledProcessError):
        pass

    return False


def generate_report(xlsx_path: Path, profiles_path: Path | None,
                    charts_dir: Path, model: str, output: Path) -> int:
    profiles = _load_profiles(profiles_path)
    print(f"📊  Reading report: {xlsx_path}")
    if profiles:
        print(f"👤  Loaded {len(profiles)} profiles from {profiles_path}")

    data = extract_data(xlsx_path, profiles)
    if not data:
        print("No data extracted.", file=sys.stderr)
        return 1

    n_sessions = 0
    n_staff = 0
    kpis = data.get("kpis", {})
    if "Total Sessions" in kpis:
        n_sessions = kpis["Total Sessions"].get("assessed", "0")
    if "Staff Count" in kpis:
        n_staff = kpis["Staff Count"].get("assessed", "0")

    profiles_json = json.dumps(profiles, ensure_ascii=False, indent=2, default=str)
    data_json = json.dumps(data, ensure_ascii=False, indent=2, default=str)

    prompt = REPORT_PROMPT.format(
        n_staff=n_staff,
        n_sessions=n_sessions,
        profiles_json=profiles_json,
        data_json=data_json,
    )

    print(f"🤖  Generating report with {model} (this may take a minute)...")
    try:
        report_md = _call_openai(model, prompt)
    except Exception as e:
        print(f"Failed: {e}", file=sys.stderr)
        return 1

    # Clean markdown fences
    report_md = re.sub(r"^```(?:markdown)?\s*", "", report_md.strip())
    report_md = re.sub(r"\s*```$", "", report_md.strip())

    # Resolve chart figure paths
    def _resolve_fig(m: re.Match) -> str:
        alt, fname = m.group(1), m.group(2)
        fig = charts_dir / fname
        return f"![{alt}]({fig})" if fig.exists() else m.group(0)

    report_md = re.sub(r"!\[([^\]]*)\]\(([^)]+\.png)\)", _resolve_fig, report_md)

    output.write_text(report_md, encoding="utf-8")
    print(f"✔  Report saved to: {output}")
    print(f"   Word count: ~{len(report_md.split())}")

    # Convert Markdown to PDF
    pdf_path = output.with_suffix(".pdf")
    converted = _md_to_pdf(report_md, pdf_path, charts_dir)

    if converted:
        print(f"✔  PDF saved to: {pdf_path}")
    else:
        print(f"   PDF conversion failed. Install one of: weasyprint, fpdf2, or pandoc")
        print(f"   pip install weasyprint  OR  pip install fpdf2")

    return 0


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Generate best practices & analysis report from AI Dev Journal data.")
    ap.add_argument("report", type=Path, help="Input report .xlsx (output of ai_journal.py)")
    ap.add_argument("--profiles", type=Path, default=None,
                    help="YAML or JSON file with staff profiles")
    ap.add_argument("--charts-dir", type=Path, default=Path("charts_output"),
                    help="Directory containing chart PNGs (default: charts_output)")
    ap.add_argument("--model", default="gpt-5.4-mini",
                    help="Model for report generation (default: gpt-5.4-mini)")
    ap.add_argument("-o", "--output", type=Path, default=Path("best_practices_report.md"),
                    help="Output Markdown file (default: best_practices_report.md)")
    args = ap.parse_args()

    if not args.report.exists():
        print(f"Report not found: {args.report}", file=sys.stderr)
        return 1

    return generate_report(args.report, args.profiles, args.charts_dir,
                           args.model, args.output)


if __name__ == "__main__":
    sys.exit(main())
