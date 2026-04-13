#!/usr/bin/env python3
"""
AI Dev Journal - Consolidated Report Generator

Reads multiple AI Dev Journal .xlsx files (one per staff, identified by filename)
and produces a consolidated Excel report with per-staff, per-category, per-tool,
time-trend, and pivot views. Optionally uses an OpenAI-compatible API to infer
lessons from each task and compare against the user's own "Bài học rút ra".

Usage:
    python report.py file1.xlsx file2.xlsx ... -o report.xlsx
    python report.py *.xlsx -o report.xlsx --no-ai
    python report.py *.xlsx -o report.xlsx --model gpt-5.4-mini
    example of file_name = 'journal_khanh.xlsx'
"""

from __future__ import annotations
from dotenv import load_dotenv
import os
import argparse
import hashlib
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

load_dotenv()
# --------------------------------------------------------------------------- #
# Constants
# --------------------------------------------------------------------------- #

OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")


NHAT_KY_SHEET = "📝 Nhật Ký"
HEADER_ROW = 3
DATA_START_ROW = 4

# Columns in Nhật Ký (1-indexed) — matches template.xlsx
COL_STT = 1
COL_DATE = 2
COL_TITLE = 3
COL_TOOL = 4
COL_CATEGORY = 5
COL_TASK_DESC = 6
COL_PROMPT = 7
COL_RESULT = 8
COL_QUALITY_TEXT = 9
COL_RATING = 10
COL_EST_HOURS = 11       # EST (without AI)
COL_ACTUAL_HOURS = 12    # Actual (with AI)
COL_TIME_SAVED = 13      # Time saved
COL_USER_LESSON = 14
COL_TAGS = 15

CACHE_PATH = Path(".ai_journal_cache.json")

# Styling
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name="Arial", italic=True, size=10, color="595959")
CELL_FONT = Font(name="Arial", size=10)
NUMBER_FONT = Font(name="Arial", size=10, bold=True, color="1F4E78")
TOTAL_FILL = PatternFill("solid", start_color="D9E1F2")
TOTAL_FONT = Font(name="Arial", bold=True, size=10)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# --------------------------------------------------------------------------- #
# Data model
# --------------------------------------------------------------------------- #

@dataclass
class Session:
    staff: str
    source_file: str
    stt: Any
    date: Any
    title: str
    tool: str
    category: str
    task_desc: str
    prompt: str
    result: str
    quality_text: str
    rating: float | None
    est_hours: float | None      # EST (without AI)
    actual_hours: float | None   # Actual (with AI)
    time_saved: float | None     # Time saved (trusted from user)
    user_lesson: str
    tags: str
    ai_lesson: str = ""
    comparison: str = ""
    ai_rating: float | None = None
    ai_rating_reason: str = ""
    suggested_prompt: str = ""

    def row_hash(self) -> str:
        """Stable hash for caching AI inference across runs."""
        key = "|".join([
            self.staff,
            str(self.date),
            self.title,
            self.task_desc,
            self.prompt,
            self.result,
            self.user_lesson,
        ])
        return hashlib.sha1(key.encode("utf-8")).hexdigest()

    @property
    def efficiency(self) -> float | None:
        """Time saved / EST — what % of work AI eliminated."""
        if self.est_hours and self.time_saved is not None:
            return self.time_saved / self.est_hours
        return None


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #

def staff_from_filename(path: Path) -> str:
    """journal_khanh.xlsx -> khanh; ai-dev-khanh.xlsx -> khanh"""
    stem = path.stem
    parts = re.split(r"[_\-\s]+", stem)
    blacklist = {"journal", "ai", "dev", "nhatky", "nhật", "ký", "log"}
    meaningful = [p for p in parts if p.lower() not in blacklist and p]
    return (meaningful[-1] if meaningful else stem).strip().capitalize()


def _cell(ws, row: int, col: int) -> Any:
    v = ws.cell(row=row, column=col).value
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def _to_float(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def parse_file(path: Path) -> list[Session]:
    wb = load_workbook(path, data_only=True)
    if NHAT_KY_SHEET not in wb.sheetnames:
        print(f"  ⚠  {path.name}: no '{NHAT_KY_SHEET}' sheet, skipping", file=sys.stderr)
        return []
    ws = wb[NHAT_KY_SHEET]
    staff = staff_from_filename(path)
    sessions: list[Session] = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        title = _cell(ws, row, COL_TITLE)
        tool = _cell(ws, row, COL_TOOL)
        if title is None and tool is None:
            continue
        sessions.append(Session(
            staff=staff,
            source_file=path.name,
            stt=_cell(ws, row, COL_STT),
            date=_cell(ws, row, COL_DATE),
            title=str(title or ""),
            tool=str(tool or ""),
            category=str(_cell(ws, row, COL_CATEGORY) or ""),
            task_desc=str(_cell(ws, row, COL_TASK_DESC) or ""),
            prompt=str(_cell(ws, row, COL_PROMPT) or ""),
            result=str(_cell(ws, row, COL_RESULT) or ""),
            quality_text=str(_cell(ws, row, COL_QUALITY_TEXT) or ""),
            rating=_to_float(_cell(ws, row, COL_RATING)),
            est_hours=_to_float(_cell(ws, row, COL_EST_HOURS)),
            actual_hours=_to_float(_cell(ws, row, COL_ACTUAL_HOURS)),
            time_saved=_to_float(_cell(ws, row, COL_TIME_SAVED)),
            user_lesson=str(_cell(ws, row, COL_USER_LESSON) or ""),
            tags=str(_cell(ws, row, COL_TAGS) or ""),
        ))
    return sessions


# --------------------------------------------------------------------------- #
# Aggregation helpers
# --------------------------------------------------------------------------- #

def _agg(items: list[Session]) -> dict[str, Any]:
    """Compute standard metrics for a group of sessions."""
    n = len(items)
    est = sum(s.est_hours or 0 for s in items)
    actual = sum(s.actual_hours or 0 for s in items)
    saved = sum(s.time_saved or 0 for s in items)
    eff = (saved / est * 100) if est else 0
    rated = [s.rating for s in items if s.rating is not None]
    avg_rating = sum(rated) / len(rated) if rated else 0
    excellent = sum(1 for s in items if s.rating == 5)
    avg_saved = saved / n if n else 0
    return {
        "n": n, "est": round(est, 1), "actual": round(actual, 1),
        "saved": round(saved, 1), "eff": round(eff, 1),
        "avg_rating": round(avg_rating, 2), "excellent": excellent,
        "avg_saved": round(avg_saved, 1),
    }


# --------------------------------------------------------------------------- #
# OpenAI-compatible inference (batch)
# --------------------------------------------------------------------------- #

PROMPT_TEMPLATE = """<role>
Bạn là chuyên gia prompt engineering, thành thạo "Claude Prompting Best Practices" của Anthropic.
Nhiệm vụ của bạn: phân tích một phiên làm việc AI thực tế, chấm điểm, và viết lại prompt theo đúng best practices.
</role>

<best_practices_rubric>
Dùng CHÍNH XÁC các nguyên tắc sau (Anthropic Claude Prompting Best Practices) làm thước đo:

1. **Clear & Direct** — Prompt có nêu cụ thể định dạng đầu ra, ràng buộc, và các bước tuần tự không? "Golden rule": nếu một đồng nghiệp mới không hiểu prompt thì AI cũng không hiểu.
2. **Context & Motivation** — Prompt có giải thích *tại sao* (mục tiêu, đối tượng dùng, bối cảnh nghiệp vụ) để AI tổng quát hoá tốt hơn không?
3. **Examples (few-shot)** — Với task phức tạp/lặp lại, có kèm 2–5 ví dụ đa dạng, bọc trong <example> tags không?
4. **XML Structure** — Các phần khác nhau (instructions, context, input, examples) có được tách bằng XML tags nhất quán để tránh nhập nhằng không?
5. **Role Assignment** — Có gán vai trò/persona cụ thể cho AI để định hướng tone & chuyên môn không?
6. **Long-context Ordering** — Với input dài (20k+ tokens, tài liệu, data), dữ liệu dài có được đặt TRÊN câu hỏi/instructions không?
7. **Positive Instructions** — Nói AI *phải làm gì* thay vì *không được làm gì*?
8. **Ground in Quotes** — Với task phân tích tài liệu dài, có yêu cầu AI trích dẫn phần liên quan trước khi xử lý không?
9. **Self-check** — Có yêu cầu AI tự kiểm tra lại kết quả trước khi kết thúc không?
10. **Output Format Specification** — Schema đầu ra có được định nghĩa rõ (JSON schema, XML tags, markdown structure)?
</best_practices_rubric>

<session_to_analyze>
<title>{title}</title>
<tool>{tool}</tool>
<category>{category}</category>
<task_description>{task_desc}</task_description>
<user_prompt>{prompt}</user_prompt>
<ai_result>{result}</ai_result>
<user_self_lesson>{user_lesson}</user_self_lesson>
</session_to_analyze>

<instructions>
Thực hiện các bước sau theo thứ tự:

1. **Đọc <user_prompt>** và kiểm tra nó đáp ứng được nguyên tắc nào trong <best_practices_rubric>, vi phạm nguyên tắc nào. Xác định 1–2 nguyên tắc BỊ VI PHẠM NẶNG NHẤT.

2. **Viết `ai_lesson`** (tiếng Việt, 2–3 câu, CỰC KỲ CỤ THỂ):
   - Chỉ đích danh nguyên tắc bị vi phạm (tên + số thứ tự từ rubric).
   - Trích dẫn hoặc mô tả chính xác phần nào trong <user_prompt> thiếu/yếu (ví dụ: "prompt không có phần định dạng đầu ra", "prompt thiếu context về stack công nghệ", "không có ví dụ mẫu cho format").
   - Giải thích ngắn gọn hậu quả quan sát được trong <ai_result>.
   - KHÔNG viết chung chung như "nên cung cấp context đầy đủ" — phải chỉ rõ context NÀO bị thiếu.

3. **So sánh với <user_self_lesson>** → chọn MỘT nhãn cho `comparison`:
   - "Đồng thuận" — bài học của bạn và user trùng về nguyên tắc chính.
   - "Bổ sung" — bạn chỉ ra thêm nguyên tắc user chưa nhận ra.
   - "Khác biệt" — bạn và user chỉ ra nguyên tắc khác nhau.
   - "Người dùng để trống" — <user_self_lesson> rỗng hoặc "(trống)".

4. **Chấm `ai_rating` 1–5** mức độ <ai_result> đáp ứng <task_description>:
   - 1=không đạt, 2=kém, 3=trung bình, 4=tốt, 5=xuất sắc.
   - Viết `ai_rating_reason` (1 câu tiếng Việt) chỉ ra điểm cụ thể khớp/lệch giữa kết quả và nhiệm vụ.

5. **Viết `suggested_prompt`** — phiên bản cải tiến của <user_prompt>, ÁP DỤNG tất cả best practices liên quan. BẮT BUỘC:
   - Bắt đầu bằng <role>...</role> gán vai trò cụ thể cho AI.
   - Có <context>...</context> nêu bối cảnh, stack công nghệ, đối tượng, mục tiêu nghiệp vụ.
   - Có <task>...</task> liệt kê yêu cầu bằng các bước đánh số nếu thứ tự quan trọng.
   - Có <constraints>...</constraints> với ràng buộc rõ ràng (nói "phải làm X", tránh "không được Y").
   - Có <output_format>...</output_format> mô tả schema/định dạng mong muốn.
   - Nếu task cần ví dụ: kèm <examples><example>...</example></examples>.
   - Kết thúc bằng 1 câu yêu cầu AI self-check trước khi trả lời.
   - Prompt BẰNG TIẾNG VIỆT, sẵn sàng copy-paste dùng lại, dài 200–500 từ.
   - KHÔNG viết placeholder như "[điền vào đây]" — phải điền dữ liệu thật từ <task_description> của session.
</instructions>

<output_format>
Trả về DUY NHẤT một đối tượng JSON hợp lệ, không markdown, không giải thích thêm:
{{"ai_lesson": "...", "comparison": "...", "ai_rating": 4, "ai_rating_reason": "...", "suggested_prompt": "..."}}
</output_format>

Trước khi trả lời, hãy tự kiểm tra: (a) ai_lesson có chỉ đích danh nguyên tắc cụ thể không? (b) suggested_prompt có đủ 5 XML tag bắt buộc không? (c) JSON có hợp lệ không?"""


def _load_cache() -> dict[str, dict]:
    if CACHE_PATH.exists():
        try:
            return json.loads(CACHE_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def _save_cache(cache: dict[str, dict]) -> None:
    try:
        CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        print(f"  ⚠  Failed to write cache: {e}", file=sys.stderr)


def _truncate(s: str, n: int = 1200) -> str:
    return s if len(s) <= n else s[:n] + "…"


def _call_openai(model: str, prompt: str, timeout: int = 300) -> str:
    base = (OPENAI_BASE_URL or "").rstrip("/")
    is_azure = "cognitiveservices.azure.com" in base or "openai.azure.com" in base

    if is_azure:
        # Azure: /openai/deployments/{model}/chat/completions?api-version=...
        host = re.sub(r"/openai(/v1)?$", "", base).rstrip("/")
        url = f"{host}/openai/deployments/{model}/chat/completions?api-version=2024-12-01-preview"
        headers = {"api-key": OPENAI_API_KEY or ""}
    else:
        url = f"{base}/chat/completions"
        headers = {}
        if OPENAI_API_KEY:
            headers["Authorization"] = f"Bearer {OPENAI_API_KEY}"

    body: dict[str, Any] = {
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.2,
        "max_completion_tokens": 2048,
        "response_format": {"type": "json_object"},
    }
    if not is_azure:
        body["model"] = model

    r = requests.post(url, headers=headers, json=body, timeout=timeout)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]


def _parse_json_response(raw: str) -> tuple[str, str, float | None, str, str]:
    if not raw:
        return "", "", None, "", ""
    raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip(), flags=re.MULTILINE)

    def _extract(obj: dict) -> tuple[str, str, float | None, str, str]:
        ai_lesson = str(obj.get("ai_lesson", "")).strip()
        comparison = str(obj.get("comparison", "")).strip()
        rating_raw = obj.get("ai_rating")
        try:
            rating = float(rating_raw) if rating_raw is not None else None
            if rating is not None:
                rating = max(1.0, min(5.0, rating))
        except (ValueError, TypeError):
            rating = None
        reason = str(obj.get("ai_rating_reason", "")).strip()
        suggested = str(obj.get("suggested_prompt", "")).strip()
        return ai_lesson, comparison, rating, reason, suggested

    try:
        return _extract(json.loads(raw))
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", raw, flags=re.DOTALL)
        if m:
            try:
                return _extract(json.loads(m.group(0)))
            except json.JSONDecodeError:
                pass
        return raw.strip(), "", None, "", ""


def infer_lessons_batch(sessions: list[Session], model: str) -> None:
    """Run OpenAI-compatible API over all sessions. Cached by row hash."""
    if not sessions:
        return
    cache = _load_cache()
    total = len(sessions)
    print(f"\n🤖  Running AI ({model}) on {total} sessions...")
    hits = 0
    for i, s in enumerate(sessions, 1):
        h = s.row_hash()
        if h in cache:
            c = cache[h]
            s.ai_lesson = c.get("ai_lesson", "")
            s.comparison = c.get("comparison", "")
            s.ai_rating = c.get("ai_rating")
            s.ai_rating_reason = c.get("ai_rating_reason", "")
            s.suggested_prompt = c.get("suggested_prompt", "")
            hits += 1
            print(f"  [{i}/{total}] {s.staff} — {s.title[:50]} (cached)")
            continue
        prompt = PROMPT_TEMPLATE.format(
            title=_truncate(s.title, 200),
            tool=s.tool,
            category=s.category,
            task_desc=_truncate(s.task_desc),
            prompt=_truncate(s.prompt),
            result=_truncate(s.result),
            user_lesson=_truncate(s.user_lesson, 600) or "(trống)",
        )
        try:
            raw = _call_openai(model, prompt)
            ai_lesson, comparison, ai_rating, reason, suggested = _parse_json_response(raw)
            s.ai_lesson = ai_lesson
            s.comparison = comparison or ("Người dùng để trống" if not s.user_lesson else "Khác biệt")
            s.ai_rating = ai_rating
            s.ai_rating_reason = reason
            s.suggested_prompt = suggested
            cache[h] = {
                "ai_lesson": s.ai_lesson,
                "comparison": s.comparison,
                "ai_rating": s.ai_rating,
                "ai_rating_reason": s.ai_rating_reason,
                "suggested_prompt": s.suggested_prompt,
            }
            rating_str = f"{ai_rating:.0f}★" if ai_rating else "—"
            print(f"  [{i}/{total}] {s.staff} — {s.title[:50]}  →  {s.comparison} ({rating_str})")
        except requests.RequestException as e:
            s.ai_lesson = f"[Lỗi AI: {e}]"
            s.comparison = "—"
            print(f"  [{i}/{total}] ⚠  {e}", file=sys.stderr)
    _save_cache(cache)
    print(f"✔  Done. Cache hits: {hits}/{total}")


# --------------------------------------------------------------------------- #
# Report builder — helpers
# --------------------------------------------------------------------------- #

def _fmt_date(d: Any) -> str:
    if isinstance(d, datetime):
        return d.strftime("%d/%m/%Y")
    return str(d or "")


def _style_header(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def _style_data_range(ws, r1: int, r2: int, n_cols: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.font is None or cell.font.name != "Arial":
                cell.font = CELL_FONT
            cell.border = BORDER
            if cell.alignment is None or cell.alignment.horizontal is None:
                cell.alignment = LEFT


def _set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_block(ws, title: str, subtitle: str, n_cols: int) -> int:
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=n_cols)
    ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=n_cols)
    ws.row_dimensions[1].height = 22
    return 4  # header row


def _write_total_row(ws, row: int, n_cols: int, hr: int, sum_cols: list[int],
                     avg_cols: list[int] | None = None) -> None:
    """Write a styled TỔNG row with SUM/AVERAGE formulas."""
    ws.cell(row=row, column=1, value="TỔNG")
    for c in sum_cols:
        col_letter = get_column_letter(c)
        ws.cell(row=row, column=c, value=f"=SUM({col_letter}{hr+1}:{col_letter}{row-1})")
    for c in (avg_cols or []):
        col_letter = get_column_letter(c)
        ws.cell(row=row, column=c,
                value=f"=IFERROR(AVERAGE({col_letter}{hr+1}:{col_letter}{row-1}),0)")
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT


# --------------------------------------------------------------------------- #
# Report builder — sheets
# --------------------------------------------------------------------------- #

def build_summary_sheet(wb: Workbook, sessions: list[Session]) -> None:
    ws = wb.create_sheet("📊 Tổng Quan")
    a = _agg(sessions)
    n_staff = len({s.staff for s in sessions})
    n_cols = 7
    hr = _title_block(
        ws,
        "📊  TỔNG QUAN  —  AI Dev Journal Consolidated Report",
        f"Generated {datetime.now().strftime('%d/%m/%Y %H:%M')}  •  "
        f"{n_staff} staff  •  {a['n']} sessions",
        n_cols,
    )

    headers = ["Chỉ Số", "Giá Trị", "", "Top 5 Công Cụ AI", "Phiên", "Giờ Tiết Kiệm", "Hiệu Suất %"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, n_cols)

    kpis = [
        ("Tổng số phiên", a["n"]),
        ("Số staff", n_staff),
        ("Tổng EST (không AI)", f"{a['est']}h"),
        ("Tổng Actual (có AI)", f"{a['actual']}h"),
        ("Tổng giờ tiết kiệm", f"{a['saved']}h"),
        ("Hiệu suất AI", f"{a['eff']}%"),
        ("Giờ tiết kiệm TB/phiên", f"{a['avg_saved']}h"),
        ("Chất lượng trung bình", f"{a['avg_rating']} / 5"),
        ("Phiên xuất sắc (5★)", a["excellent"]),
    ]
    for i, (k, v) in enumerate(kpis):
        ws.cell(row=hr + 1 + i, column=1, value=k).font = CELL_FONT
        ws.cell(row=hr + 1 + i, column=2, value=v).font = NUMBER_FONT

    # Top tools by time saved
    tool_data: dict[str, dict] = {}
    for s in sessions:
        if s.tool:
            d = tool_data.setdefault(s.tool, {"n": 0, "saved": 0.0, "est": 0.0})
            d["n"] += 1
            d["saved"] += s.time_saved or 0
            d["est"] += s.est_hours or 0
    top_tools = sorted(tool_data.items(), key=lambda x: -x[1]["saved"])[:5]
    for i, (t, d) in enumerate(top_tools):
        eff = (d["saved"] / d["est"] * 100) if d["est"] else 0
        ws.cell(row=hr + 1 + i, column=4, value=t).font = CELL_FONT
        ws.cell(row=hr + 1 + i, column=5, value=d["n"]).font = NUMBER_FONT
        ws.cell(row=hr + 1 + i, column=6, value=round(d["saved"], 1)).font = NUMBER_FONT
        ws.cell(row=hr + 1 + i, column=7, value=f"{eff:.0f}%").font = NUMBER_FONT

    last_row = hr + max(len(kpis), len(top_tools))
    _style_data_range(ws, hr + 1, last_row, n_cols)
    _set_widths(ws, [28, 18, 3, 30, 10, 16, 14])


def build_per_staff_sheet(wb: Workbook, sessions: list[Session]) -> None:
    ws = wb.create_sheet("👤 Per Staff")
    headers = [
        "Staff", "Số Phiên", "EST (h)", "Actual (h)", "Tiết Kiệm (h)",
        "Hiệu Suất %", "TB Tiết Kiệm/Phiên", "Rating TB", "Phiên 5★", "Công Cụ Chính",
    ]
    n_cols = len(headers)
    hr = _title_block(ws, "👤  THỐNG KÊ THEO STAFF", "Ranking by total hours saved", n_cols)
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, n_cols)

    by_staff: dict[str, list[Session]] = {}
    for s in sessions:
        by_staff.setdefault(s.staff, []).append(s)

    rows = []
    for staff, items in by_staff.items():
        a = _agg(items)
        tools: dict[str, int] = {}
        for s in items:
            if s.tool:
                tools[s.tool] = tools.get(s.tool, 0) + 1
        main_tool = max(tools.items(), key=lambda x: x[1])[0] if tools else "—"
        rows.append((
            staff, a["n"], a["est"], a["actual"], a["saved"],
            a["eff"], a["avg_saved"], a["avg_rating"], a["excellent"], main_tool,
        ))

    rows.sort(key=lambda r: -r[4])  # by hours saved
    for i, row in enumerate(rows):
        for j, v in enumerate(row, 1):
            ws.cell(row=hr + 1 + i, column=j, value=v)

    total_row = hr + 1 + len(rows)
    _write_total_row(ws, total_row, n_cols, hr,
                     sum_cols=[2, 3, 4, 5, 9],
                     avg_cols=[6, 7, 8])
    _style_data_range(ws, hr + 1, total_row, n_cols)
    _set_widths(ws, [16, 10, 12, 12, 14, 12, 16, 10, 10, 24])


def build_per_tool_sheet(wb: Workbook, sessions: list[Session]) -> None:
    ws = wb.create_sheet("🔧 Per Tool")
    headers = [
        "Công Cụ AI", "Số Phiên", "EST (h)", "Actual (h)", "Tiết Kiệm (h)",
        "Hiệu Suất %", "TB Tiết Kiệm/Phiên", "Rating TB", "Phiên 5★",
    ]
    n_cols = len(headers)
    hr = _title_block(ws, "🔧  THỐNG KÊ THEO CÔNG CỤ AI", "Ranking by total hours saved", n_cols)
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, n_cols)

    by_tool: dict[str, list[Session]] = {}
    for s in sessions:
        key = s.tool or "(không rõ)"
        by_tool.setdefault(key, []).append(s)

    rows = []
    for tool, items in by_tool.items():
        a = _agg(items)
        rows.append((
            tool, a["n"], a["est"], a["actual"], a["saved"],
            a["eff"], a["avg_saved"], a["avg_rating"], a["excellent"],
        ))

    rows.sort(key=lambda r: -r[4])
    for i, row in enumerate(rows):
        for j, v in enumerate(row, 1):
            ws.cell(row=hr + 1 + i, column=j, value=v)

    total_row = hr + 1 + len(rows)
    _write_total_row(ws, total_row, n_cols, hr,
                     sum_cols=[2, 3, 4, 5, 9],
                     avg_cols=[6, 7, 8])
    _style_data_range(ws, hr + 1, total_row, n_cols)
    _set_widths(ws, [24, 10, 12, 12, 14, 12, 16, 10, 10])


def build_per_category_sheet(wb: Workbook, sessions: list[Session]) -> None:
    ws = wb.create_sheet("📂 Per Category")
    headers = [
        "Danh Mục", "Số Phiên", "EST (h)", "Actual (h)", "Tiết Kiệm (h)",
        "Hiệu Suất %", "TB Tiết Kiệm/Phiên", "Rating TB", "Phiên 5★",
    ]
    n_cols = len(headers)
    hr = _title_block(ws, "📂  THỐNG KÊ THEO DANH MỤC", "Sorted by total hours saved", n_cols)
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, n_cols)

    by_cat: dict[str, list[Session]] = {}
    for s in sessions:
        key = s.category or "(chưa phân loại)"
        by_cat.setdefault(key, []).append(s)

    rows = []
    for cat, items in by_cat.items():
        a = _agg(items)
        rows.append((
            cat, a["n"], a["est"], a["actual"], a["saved"],
            a["eff"], a["avg_saved"], a["avg_rating"], a["excellent"],
        ))

    rows.sort(key=lambda r: -r[4])
    for i, row in enumerate(rows):
        for j, v in enumerate(row, 1):
            ws.cell(row=hr + 1 + i, column=j, value=v)

    total_row = hr + 1 + len(rows)
    _write_total_row(ws, total_row, n_cols, hr,
                     sum_cols=[2, 3, 4, 5, 9],
                     avg_cols=[6, 7, 8])
    _style_data_range(ws, hr + 1, total_row, n_cols)
    _set_widths(ws, [24, 10, 12, 12, 14, 12, 16, 10, 10])


def build_time_trend_sheet(wb: Workbook, sessions: list[Session]) -> None:
    ws = wb.create_sheet("📅 Time Trend")
    headers = ["Ngày", "Số Phiên", "EST (h)", "Actual (h)", "Tiết Kiệm (h)", "Hiệu Suất %", "Rating TB"]
    n_cols = len(headers)
    hr = _title_block(ws, "📅  XU HƯỚNG THEO NGÀY", "Daily activity across all staff", n_cols)
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, n_cols)

    by_day: dict[str, list[Session]] = {}
    for s in sessions:
        key = _fmt_date(s.date) or "(no date)"
        by_day.setdefault(key, []).append(s)

    def _sort_key(k: str):
        try:
            return datetime.strptime(k, "%d/%m/%Y")
        except ValueError:
            return datetime.max

    for i, day in enumerate(sorted(by_day.keys(), key=_sort_key)):
        a = _agg(by_day[day])
        ws.cell(row=hr + 1 + i, column=1, value=day)
        ws.cell(row=hr + 1 + i, column=2, value=a["n"])
        ws.cell(row=hr + 1 + i, column=3, value=a["est"])
        ws.cell(row=hr + 1 + i, column=4, value=a["actual"])
        ws.cell(row=hr + 1 + i, column=5, value=a["saved"])
        ws.cell(row=hr + 1 + i, column=6, value=a["eff"])
        ws.cell(row=hr + 1 + i, column=7, value=a["avg_rating"])

    _style_data_range(ws, hr + 1, hr + len(by_day), n_cols)
    _set_widths(ws, [16, 10, 12, 12, 14, 12, 10])


def _build_pivot(
    wb: Workbook,
    sheet_name: str,
    title: str,
    sessions: list[Session],
    col_attr: str,
) -> None:
    ws = wb.create_sheet(sheet_name)
    staff_list = sorted({s.staff for s in sessions})
    col_values = sorted({getattr(s, col_attr) or "(none)" for s in sessions})

    n_cols = len(col_values) + 2  # Staff + values + TỔNG
    hr = _title_block(ws, title, "Cell = sessions | EST → Actual (saved)", n_cols)

    ws.cell(row=hr, column=1, value="Staff \\ " + col_attr)
    for i, v in enumerate(col_values, 2):
        ws.cell(row=hr, column=i, value=v)
    ws.cell(row=hr, column=n_cols, value="TỔNG")
    _style_header(ws, hr, n_cols)

    for ri, staff in enumerate(staff_list):
        ws.cell(row=hr + 1 + ri, column=1, value=staff)
        staff_total_n = 0
        staff_total_est = 0.0
        staff_total_actual = 0.0
        staff_total_saved = 0.0
        for ci, v in enumerate(col_values, 2):
            items = [
                s for s in sessions
                if s.staff == staff and (getattr(s, col_attr) or "(none)") == v
            ]
            if items:
                a = _agg(items)
                ws.cell(row=hr + 1 + ri, column=ci,
                        value=f"{a['n']} | {a['est']}→{a['actual']} ({a['saved']}h)")
                staff_total_n += a["n"]
                staff_total_est += a["est"]
                staff_total_actual += a["actual"]
                staff_total_saved += a["saved"]
            else:
                ws.cell(row=hr + 1 + ri, column=ci, value="—")
        ws.cell(
            row=hr + 1 + ri,
            column=n_cols,
            value=f"{staff_total_n} | {staff_total_est:.1f}→{staff_total_actual:.1f} ({staff_total_saved:.1f}h)",
        ).font = TOTAL_FONT

    _style_data_range(ws, hr + 1, hr + len(staff_list), n_cols)
    widths = [18] + [28] * len(col_values) + [28]
    _set_widths(ws, widths)


def build_raw_log_sheet(wb: Workbook, sessions: list[Session]) -> None:
    ws = wb.create_sheet("📝 Raw Log")
    headers = [
        "Staff", "Ngày", "Tên Phiên", "Công Cụ", "Danh Mục",
        "Mô Tả", "Rating", "EST (h)", "Actual (h)", "Tiết Kiệm (h)",
        "Hiệu Suất %", "Bài Học Người Dùng", "Tags", "Source File",
    ]
    n_cols = len(headers)
    hr = _title_block(
        ws,
        "📝  RAW LOG  —  Consolidated Sessions",
        f"{len(sessions)} rows",
        n_cols,
    )
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, n_cols)

    for i, s in enumerate(sessions):
        r = hr + 1 + i
        eff = f"{s.efficiency * 100:.0f}%" if s.efficiency is not None else "—"
        ws.cell(row=r, column=1, value=s.staff)
        ws.cell(row=r, column=2, value=_fmt_date(s.date))
        ws.cell(row=r, column=3, value=s.title)
        ws.cell(row=r, column=4, value=s.tool)
        ws.cell(row=r, column=5, value=s.category)
        ws.cell(row=r, column=6, value=s.task_desc)
        ws.cell(row=r, column=7, value=s.rating)
        ws.cell(row=r, column=8, value=s.est_hours)
        ws.cell(row=r, column=9, value=s.actual_hours)
        ws.cell(row=r, column=10, value=s.time_saved)
        ws.cell(row=r, column=11, value=eff)
        ws.cell(row=r, column=12, value=s.user_lesson)
        ws.cell(row=r, column=13, value=s.tags)
        ws.cell(row=r, column=14, value=s.source_file)

    _style_data_range(ws, hr + 1, hr + len(sessions), n_cols)
    _set_widths(ws, [14, 12, 32, 16, 16, 40, 8, 10, 10, 12, 10, 40, 24, 22])
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)


def build_ai_comparison_sheet(wb: Workbook, sessions: list[Session]) -> None:
    ws = wb.create_sheet("🤖 AI Lesson Compare")
    headers = [
        "Staff", "Ngày", "Tên Phiên", "Công Cụ",
        "Mô Tả Nhiệm Vụ", "Prompt Chính", "Kết Quả",
        "Bài Học Người Dùng", "Bài Học AI Suy Luận", "So Sánh",
        "User ★", "AI ★", "Δ (AI − User)", "Lý Do AI Chấm",
        "Prompt Đề Xuất (AI + User Lessons)",
    ]
    n_cols = len(headers)
    hr = _title_block(
        ws,
        "🤖  SO SÁNH BÀI HỌC & ĐỀ XUẤT PROMPT  —  AI vs Người Dùng",
        "AI infers lesson, rates output (1–5), and suggests an improved prompt applying both lessons",
        n_cols,
    )
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, n_cols)

    comparison_fills = {
        "Đồng thuận": PatternFill("solid", start_color="C6EFCE"),
        "Bổ sung": PatternFill("solid", start_color="FFEB9C"),
        "Khác biệt": PatternFill("solid", start_color="FFC7CE"),
        "Người dùng để trống": PatternFill("solid", start_color="D9D9D9"),
    }
    gap_green = PatternFill("solid", start_color="C6EFCE")
    gap_red = PatternFill("solid", start_color="FFC7CE")
    suggested_fill = PatternFill("solid", start_color="FFF2CC")

    for i, s in enumerate(sessions):
        r = hr + 1 + i
        ws.cell(row=r, column=1, value=s.staff)
        ws.cell(row=r, column=2, value=_fmt_date(s.date))
        ws.cell(row=r, column=3, value=s.title)
        ws.cell(row=r, column=4, value=s.tool)
        ws.cell(row=r, column=5, value=s.task_desc or "—")
        ws.cell(row=r, column=6, value=s.prompt or "—")
        ws.cell(row=r, column=7, value=s.result or "—")
        ws.cell(row=r, column=8, value=s.user_lesson or "(trống)")
        ws.cell(row=r, column=9, value=s.ai_lesson or "—")
        comp_cell = ws.cell(row=r, column=10, value=s.comparison or "—")
        if s.comparison in comparison_fills:
            comp_cell.fill = comparison_fills[s.comparison]
            comp_cell.alignment = CENTER

        user_cell = ws.cell(row=r, column=11, value=s.rating)
        ai_cell = ws.cell(row=r, column=12, value=s.ai_rating)
        user_cell.alignment = CENTER
        ai_cell.alignment = CENTER

        gap = None
        if s.rating is not None and s.ai_rating is not None:
            gap = round(s.ai_rating - s.rating, 1)
        gap_cell = ws.cell(row=r, column=13, value=gap)
        gap_cell.alignment = CENTER
        if gap is not None:
            if gap >= 1:
                gap_cell.fill = gap_green
            elif gap <= -1:
                gap_cell.fill = gap_red

        ws.cell(row=r, column=14, value=s.ai_rating_reason or "—")
        sug_cell = ws.cell(row=r, column=15, value=s.suggested_prompt or "—")
        if s.suggested_prompt:
            sug_cell.fill = suggested_fill

    _style_data_range(ws, hr + 1, hr + len(sessions), n_cols)
    _set_widths(ws, [14, 12, 26, 14, 36, 36, 36, 34, 38, 14, 8, 8, 12, 32, 65])
    ws.freeze_panes = ws.cell(row=hr + 1, column=5)

    for r in range(hr + 1, hr + 1 + len(sessions)):
        ws.row_dimensions[r].height = 150


# --------------------------------------------------------------------------- #
# Report orchestrator
# --------------------------------------------------------------------------- #

def build_report(sessions: list[Session], output: Path, with_ai: bool) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, sessions)
    build_per_staff_sheet(wb, sessions)
    build_per_tool_sheet(wb, sessions)
    build_per_category_sheet(wb, sessions)
    build_time_trend_sheet(wb, sessions)
    _build_pivot(wb, "🔀 Pivot Staff×Tool", "🔀  PIVOT  —  Staff × Công Cụ", sessions, "tool")
    _build_pivot(wb, "🔀 Pivot Staff×Cat", "🔀  PIVOT  —  Staff × Danh Mục", sessions, "category")
    build_raw_log_sheet(wb, sessions)
    if with_ai:
        build_ai_comparison_sheet(wb, sessions)

    wb.save(output)


# --------------------------------------------------------------------------- #
# Terminal summary
# --------------------------------------------------------------------------- #

def print_terminal_summary(sessions: list[Session]) -> None:
    a = _agg(sessions)
    n_staff = len({s.staff for s in sessions})

    print("\n" + "=" * 64)
    print("  📊  AI DEV JOURNAL — SUMMARY")
    print("=" * 64)
    print(f"  Staff: {n_staff}    Sessions: {a['n']}    Rating TB: {a['avg_rating']}/5    5★: {a['excellent']}")
    print(f"  EST (without AI):  {a['est']:>8}h")
    print(f"  Actual (with AI):  {a['actual']:>8}h")
    print(f"  Time saved:        {a['saved']:>8}h")
    print(f"  Efficiency:        {a['eff']:>7}%")
    print(f"  Avg saved/session: {a['avg_saved']:>8}h")

    # Per staff
    by_staff: dict[str, list[Session]] = {}
    for s in sessions:
        by_staff.setdefault(s.staff, []).append(s)

    print("\n  ── Per Staff " + "─" * 49)
    print(f"  {'Staff':<14} {'#':>4} {'EST':>7} {'Actual':>7} {'Saved':>7} {'Eff%':>6} {'Rating':>6}")
    for staff in sorted(by_staff, key=lambda k: -sum(s.time_saved or 0 for s in by_staff[k])):
        sa = _agg(by_staff[staff])
        print(f"  {staff:<14} {sa['n']:>4} {sa['est']:>6}h {sa['actual']:>6}h {sa['saved']:>6}h {sa['eff']:>5}% {sa['avg_rating']:>6}")

    # Per tool
    by_tool: dict[str, list[Session]] = {}
    for s in sessions:
        key = s.tool or "(không rõ)"
        by_tool.setdefault(key, []).append(s)

    print("\n  ── Per Tool " + "─" * 50)
    print(f"  {'Tool':<22} {'#':>4} {'EST':>7} {'Actual':>7} {'Saved':>7} {'Eff%':>6}")
    for tool in sorted(by_tool, key=lambda k: -sum(s.time_saved or 0 for s in by_tool[k])):
        ta = _agg(by_tool[tool])
        print(f"  {tool:<22} {ta['n']:>4} {ta['est']:>6}h {ta['actual']:>6}h {ta['saved']:>6}h {ta['eff']:>5}%")

    # Per category
    by_cat: dict[str, list[Session]] = {}
    for s in sessions:
        key = s.category or "(chưa phân loại)"
        by_cat.setdefault(key, []).append(s)

    print("\n  ── Per Category " + "─" * 46)
    print(f"  {'Category':<22} {'#':>4} {'EST':>7} {'Actual':>7} {'Saved':>7} {'Eff%':>6}")
    for cat in sorted(by_cat, key=lambda k: -sum(s.time_saved or 0 for s in by_cat[k])):
        ca = _agg(by_cat[cat])
        print(f"  {cat:<22} {ca['n']:>4} {ca['est']:>6}h {ca['actual']:>6}h {ca['saved']:>6}h {ca['eff']:>5}%")

    print("=" * 64)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Aggregate AI Dev Journal files into a consolidated report.",
    )
    ap.add_argument("files", nargs="+", type=Path, help="Input .xlsx files")
    ap.add_argument("-o", "--output", type=Path, default=Path("ai_journal_report.xlsx"))
    ap.add_argument("--model", default="qwen2.5:7b",
                    help="Model name for OpenAI-compatible API (default: qwen2.5:7b)")
    ap.add_argument("--no-ai", action="store_true", help="Skip AI lesson inference")
    args = ap.parse_args()

    all_sessions: list[Session] = []
    print(f"📂  Reading {len(args.files)} file(s)...")
    for f in args.files:
        if not f.exists():
            print(f"  ⚠  {f}: not found, skipping", file=sys.stderr)
            continue
        try:
            sess = parse_file(f)
            print(f"  ✔  {f.name} → staff='{staff_from_filename(f)}', {len(sess)} sessions")
            all_sessions.extend(sess)
        except Exception as e:
            print(f"  ✖  {f.name}: {e}", file=sys.stderr)

    if not all_sessions:
        print("No sessions parsed. Exiting.", file=sys.stderr)
        return 1

    if not args.no_ai:
        try:
            infer_lessons_batch(all_sessions, args.model)
        except Exception as e:
            print(f"⚠  AI inference failed: {e}", file=sys.stderr)
            print("   Continuing without AI lesson sheet.", file=sys.stderr)

    build_report(all_sessions, args.output, with_ai=not args.no_ai)
    print_terminal_summary(all_sessions)
    print(f"\n✔  Report saved to: {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
