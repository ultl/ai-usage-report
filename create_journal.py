#!/usr/bin/env python3
"""
AI Dev Journal — Create Journal from Prompt Log

Reads PROMPT_LOG.md, groups prompts into logical sessions/tasks, then uses LLM
to fill in all journal fields (title, category, description, result, quality,
rating, EST, actual, time saved, lesson, tags) based on the prompt content.

Outputs a new .xlsx journal file matching template.xlsx structure.

Usage:
    python create_journal.py PROMPT_LOG.md -o journal_ltluyen8.xlsx --model gpt-5.4-mini
    python create_journal.py PROMPT_LOG.md -o journal_ltluyen8.xlsx --model gpt-5.4-mini --staff Ltluyen8
    python create_journal.py PROMPT_LOG.md -o journal_ltluyen8.xlsx --profile profiles.yml --staff Ltluyen8
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment

try:
    import yaml
except ImportError:
    yaml = None

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"

NHAT_KY_SHEET = "📝 Nhật Ký"
HEADER_ROW = 3
DATA_START_ROW = 4


def _call_openai(model: str, prompt: str, timeout: int = 600) -> str:
    base = (OPENAI_BASE_URL or "").rstrip("/")
    is_azure = "cognitiveservices.azure.com" in base or "openai.azure.com" in base

    if is_azure:
        host = re.sub(r"/openai(/v1)?$", "", base).rstrip("/")
        url = f"{host}/openai/deployments/{model}/chat/completions?api-version=2024-12-01-preview"
        headers = {"api-key": OPENAI_API_KEY or ""}
    else:
        url = f"{base}/chat/completions"
        headers = {}
        if OPENAI_API_KEY:
            headers["Authorization"] = f"Bearer {OPENAI_API_KEY}"

    body: dict[str, Any] = {
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.2,
        "max_completion_tokens": 8192,
        "response_format": {"type": "json_object"},
    }
    if not is_azure:
        body["model"] = model

    r = requests.post(url, headers=headers, json=body, timeout=timeout)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]


def _load_profiles(path: Path | None) -> dict[str, dict[str, str]]:
    if path is None or not path.exists():
        return {}
    try:
        text = path.read_text(encoding="utf-8")
        if path.suffix.lower() in (".yml", ".yaml"):
            if yaml is None:
                return {}
            raw = yaml.safe_load(text)
        else:
            raw = json.loads(text)
        if isinstance(raw, dict):
            return {k: v for k, v in raw.items() if isinstance(v, dict)}
    except Exception:
        pass
    return {}


def parse_prompt_log(path: Path) -> list[dict[str, Any]]:
    """Parse PROMPT_LOG.md into a list of {number, date, session, text} dicts.

    Each prompt is tagged with its session identifier (e.g. "Session 1")
    so that downstream code can group all prompts from the same session.
    """
    content = path.read_text(encoding="utf-8")
    prompts: list[dict[str, Any]] = []
    current_date = ""
    current_session = ""

    for line in content.split("\n"):
        # Date/session header: ## 2026-04-20 — Session 1
        date_match = re.match(r"^## (\d{4}-\d{2}-\d{2})\s*[—–-]\s*(.*)", line)
        if date_match:
            current_date = date_match.group(1)
            current_session = date_match.group(2).strip() or current_date
            continue

        # Prompt header: ### Prompt 1
        prompt_match = re.match(r"^### Prompt (\d+)", line)
        if prompt_match:
            prompts.append({
                "number": prompt_match.group(1),
                "date": current_date,
                "session": current_session,
                "text": "",
            })
            continue

        # Content lines (blockquote)
        if prompts and line.startswith(">"):
            text = line.lstrip("> ").strip()
            if text:
                if prompts[-1]["text"]:
                    prompts[-1]["text"] += "\n" + text
                else:
                    prompts[-1]["text"] = text

    # Filter out empty prompts
    return [p for p in prompts if p["text"].strip()]


def group_prompts_by_session(prompts: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Group prompts into session bundles: [{session, date, prompts: [...]}]."""
    from collections import OrderedDict
    sessions: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for p in prompts:
        key = p["session"]
        if key not in sessions:
            sessions[key] = {
                "session": key,
                "date": p["date"],
                "prompts": [],
            }
        sessions[key]["prompts"].append(p)
    return list(sessions.values())


JOURNAL_PROMPT = """<role>
You are a senior software engineer filling out an AI Dev Journal for a team member.
You analyze their prompts/interactions with AI tools and create structured journal entries.
</role>

<staff_profile>
{profile_json}
</staff_profile>

<task>
Below are prompts already grouped by session. Each session contains all prompts from
a single working session. Create ONE journal entry per session.

For EACH session, fill in these fields:

- `date`: The date (from the prompt log), format YYYY-MM-DD
- `title`: Short descriptive title of the task (5-10 words)
- `tool`: AI tool used (e.g., "Claude Code", "Claude", "ChatGPT", "Copilot"). Infer from context.
- `category`: One of: Architecture, Backend, Frontend, Bug Fix, Documentation, Refactor, Report, Testing, DevOps, Research, Database Design, Business Logic, Integration
- `task_desc`: 1-2 sentence description of what the session was about
- `all_prompts`: ALL individual prompts from this session, joined with "\\n---\\n" separator.
  Include the FULL text of every prompt — do NOT summarize or truncate. This is the raw log.
- `prompt`: A summarized version of the main prompt (keep under 200 chars). The prompt needs to follow these principles (Anthropic Claude Prompting Best Practices) as your scoring rubric:
    1. **Clear & Direct** — Does the prompt specify output format, constraints, and sequential steps?
    2. **Context & Motivation** — Does the prompt explain *why* (goals, target audience, business context)?
    3. **Examples (few-shot)** — For complex/repeated tasks, are 2–5 diverse examples included?
    4. **XML Structure** — Are different sections separated with consistent XML tags?
    5. **Role Assignment** — Is a specific role/persona assigned to the AI?
    6. **Long-context Ordering** — For long inputs, is the data placed ABOVE the question/instructions?
    7. **Positive Instructions** — Does it tell AI *what to do* rather than *what not to do*?
    8. **Ground in Quotes** — For document analysis, does it ask AI to quote relevant parts first?
    9. **Self-check** — Does it ask AI to verify its own output before finishing?
    10. **Output Format Specification** — Is the output schema clearly defined?
- `result`: What was achieved/delivered (infer from the prompt intent)
- `quality`: Rating text like "5 ★ Excellent", "4 ★ Good", "3 ★ Average", "2 ★ Poor"
- `rating`: Numeric 1-5 (estimate based on how well-defined the prompt was and likely outcome)
- `est_hours`: Estimated hours WITHOUT AI (how long would this take manually?)
- `actual_hours`: Estimated hours WITH AI assistance
- `time_saved`: est_hours - actual_hours
- `lesson`: What lesson was learned or could be learned from this interaction
- `tags`: Relevant tags, comma-separated

Guidelines for estimating hours:
- Consider the staff member's role and experience level
- Simple prompts (quick questions): 0.5-1h manual, 0.1-0.25h with AI
- Medium tasks (code changes, analysis): 2-4h manual, 0.5-1.5h with AI
- Complex tasks (architecture, full features): 4-8h manual, 1-3h with AI
</task>

<sessions>
{sessions_json}
</sessions>

<output_format>
Return a JSON object with a "sessions" array. ONE entry per session:
{{"sessions": [
  {{
    "date": "2026-04-20",
    "title": "...",
    "tool": "...",
    "category": "...",
    "task_desc": "...",
    "all_prompts": "Prompt 1 full text\\n---\\nPrompt 2 full text\\n---\\n...",
    "prompt": "...",
    "result": "...",
    "quality": "4 ★ Good",
    "rating": 4,
    "est_hours": 4.0,
    "actual_hours": 1.5,
    "time_saved": 2.5,
    "lesson": "...",
    "tags": "..."
  }}
]}}

Return ALL sessions. Be realistic with hour estimates.
</output_format>"""


def generate_journal_entries(prompts: list[dict], model: str,
                             profile: dict[str, str]) -> list[dict]:
    """Group prompts by session, send each batch to LLM, get journal entries."""
    profile_info = dict(profile) if profile else {"note": "No profile provided"}

    session_groups = group_prompts_by_session(prompts)
    print(f"   Grouped into {len(session_groups)} session(s)")

    all_entries: list[dict] = []
    # Process sessions in batches (max 5 sessions per LLM call)
    batch_size = 5

    for start in range(0, len(session_groups), batch_size):
        batch = session_groups[start:start + batch_size]
        session_names = [s["session"] for s in batch]
        print(f"  • Processing session(s): {', '.join(session_names)}...")

        prompt = JOURNAL_PROMPT.format(
            profile_json=json.dumps(profile_info, ensure_ascii=False, indent=2),
            sessions_json=json.dumps(batch, ensure_ascii=False, indent=2),
        )

        try:
            raw = _call_openai(model, prompt)
            raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip(), flags=re.MULTILINE)
            parsed = json.loads(raw)
            sessions = parsed.get("sessions", []) if isinstance(parsed, dict) else parsed
            if isinstance(sessions, list):
                # For each entry, if LLM didn't return all_prompts, build it
                # from the source session data
                for entry in sessions:
                    if not entry.get("all_prompts"):
                        # Find matching session group and build all_prompts
                        entry_date = entry.get("date", "")
                        for sg in batch:
                            if sg["date"] == entry_date:
                                texts = [p["text"] for p in sg["prompts"]]
                                entry["all_prompts"] = "\n---\n".join(texts)
                                break
                all_entries.extend(sessions)
        except Exception as e:
            print(f"  ⚠  Batch failed: {e}", file=sys.stderr)

    return all_entries


def write_journal(entries: list[dict], output: Path, staff: str) -> None:
    """Write journal entries to xlsx. Only modifies the Nhật Ký sheet; all other sheets stay intact."""
    # If output already exists, open it to preserve other sheets
    # Otherwise copy from template
    if output.exists():
        wb = load_workbook(output)
        print(f"   Opening existing file: {output} (only updating {NHAT_KY_SHEET})")
    else:
        template = TEMPLATE_PATH
        if not template.exists():
            print(f"Template not found: {template}", file=sys.stderr)
            sys.exit(1)
        wb = load_workbook(template)

    ws = wb[NHAT_KY_SHEET]

    # Clear existing data rows in Nhật Ký only (keep header row 3)
    for r in range(DATA_START_ROW, ws.max_row + 1):
        for c in range(1, 16):
            ws.cell(r, c).value = None

    # Write entries
    for i, entry in enumerate(entries):
        r = DATA_START_ROW + i
        ws.cell(r, 1, value=i + 1)  # STT

        # Parse date
        date_str = entry.get("date", "")
        try:
            ws.cell(r, 2, value=datetime.strptime(date_str, "%Y-%m-%d"))
        except (ValueError, TypeError):
            ws.cell(r, 2, value=date_str)

        ws.cell(r, 3, value=entry.get("title", ""))
        ws.cell(r, 4, value=entry.get("tool", ""))
        ws.cell(r, 5, value=entry.get("category", ""))
        ws.cell(r, 6, value=entry.get("task_desc", ""))
        # Col 7: store ALL prompts from the session (full log);
        # fall back to summarized prompt if all_prompts is empty
        all_prompts = entry.get("all_prompts", "") or entry.get("prompt", "")
        cell7 = ws.cell(r, 7, value=all_prompts)
        cell7.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 8, value=entry.get("result", ""))
        ws.cell(r, 9, value=entry.get("quality", ""))
        ws.cell(r, 10, value=entry.get("rating"))
        cell11 = ws.cell(r, 11, value=entry.get("est_hours"))
        cell11.number_format = '0.0'
        cell12 = ws.cell(r, 12, value=entry.get("actual_hours"))
        cell12.number_format = '0.0'
        cell13 = ws.cell(r, 13, value=entry.get("time_saved"))
        cell13.number_format = '0.0'
        ws.cell(r, 14, value=entry.get("lesson", ""))
        ws.cell(r, 15, value=entry.get("tags", ""))
        
        print("Wrote entry:", entry.get("title", "Untitled"))

        # Set row height — taller for multi-prompt cells
        prompt_lines = all_prompts.count("\n") + 1
        ws.row_dimensions[r].height = max(60, min(prompt_lines * 15, 400))

    # Save
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Create AI Dev Journal .xlsx from PROMPT_LOG.md using LLM.")
    ap.add_argument("prompt_log", type=Path, help="Input PROMPT_LOG.md file")
    ap.add_argument("-o", "--output", type=Path, default=None,
                    help="Output .xlsx file (default: journal_<staff>.xlsx)")
    ap.add_argument("--model", default="gpt-5.4-mini",
                    help="Model for journal generation (default: gpt-5.4-mini)")
    ap.add_argument("--staff", default="Ltluyen8",
                    help="Staff name for the journal (default: Ltluyen8)")
    ap.add_argument("--profiles", type=Path, default=None,
                    help="YAML/JSON profiles file to get staff context")
    ap.add_argument("--template", type=Path, default=None,
                    help="Template .xlsx file (default: template.xlsx in same dir)")
    args = ap.parse_args()

    if not args.prompt_log.exists():
        print(f"Prompt log not found: {args.prompt_log}", file=sys.stderr)
        return 1

    global TEMPLATE_PATH
    if args.template:
        TEMPLATE_PATH = args.template

    output = args.output or Path(f"journal_{args.staff.lower()}.xlsx")

    # Load profile
    profiles = _load_profiles(args.profiles)
    profile: dict[str, str] = {}
    for k, v in profiles.items():
        if k.strip().casefold() == args.staff.strip().casefold():
            profile = v
            break
    if profile:
        profile["name"] = args.staff

    # Parse prompts
    print(f"📝  Reading prompt log: {args.prompt_log}")
    prompts = parse_prompt_log(args.prompt_log)
    print(f"   Found {len(prompts)} prompts")

    if not prompts:
        print("No prompts found.", file=sys.stderr)
        return 1

    # Generate journal entries
    print(f"🤖  Generating journal entries with {args.model}...")
    entries = generate_journal_entries(prompts, args.model, profile)
    print(f"   Generated {len(entries)} journal entries")

    if not entries:
        print("No entries generated.", file=sys.stderr)
        return 1

    # Write to xlsx
    write_journal(entries, output, args.staff)
    print(f"✔  Journal saved to: {output}")
    print(f"   {len(entries)} sessions for {args.staff}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
