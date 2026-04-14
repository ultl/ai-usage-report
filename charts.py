#!/usr/bin/env python3
"""
AI Dev Journal — Visual Charts Generator

Reads the consolidated report .xlsx produced by report.py and generates
CEO-ready charts as a multi-page PDF + individual PNGs.

Charts:
  1. SDLC Stage Tasks — stages on the Y-axis with task names listed per stage
  2. Staff AI Effectiveness — how effectively each staff uses AI for those tasks
  3. EST vs Actual — per tool and per category (grouped bars)
  4. User Satisfaction — rating distribution per tool (stacked bars)
  5. Prompt Error Classification — top errors bar + staff×error heatmap
     (errors classified via OpenAI-compatible API from ai_lesson text)

Usage:
    python charts.py report.xlsx
    python charts.py report.xlsx -o charts_output
    python charts.py report.xlsx --model gpt-5.4-mini
    python charts.py report.xlsx --no-ai   # skip AI classification
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import textwrap
import warnings
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.backends.backend_pdf import PdfPages
import numpy as np
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

from plot_charts import (
    CACHE_PATH as CHART_CLASSIFIER_CACHE_PATH,
    SDLC_TAXONOMY,
    classify_sessions as classify_chart_sessions,
    mark_unclassified as mark_sdlc_unclassified,
    read_sessions as read_chart_sessions,
)

load_dotenv()

OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# --------------------------------------------------------------------------- #
# Theme — professional dark navy matching Excel report (#1F4E78)
# --------------------------------------------------------------------------- #

NAVY = "#1F4E78"
BLUE = "#2E75B6"
LIGHT_BLUE = "#9DC3E6"
TEAL = "#00B0F0"
GREEN = "#70AD47"
ORANGE = "#ED7D31"
RED = "#FF4B4B"
GOLD = "#FFC000"
GRAY = "#595959"
LIGHT_GRAY = "#F2F2F2"
WHITE = "#FFFFFF"

PALETTE = [BLUE, ORANGE, GREEN, TEAL, GOLD, RED, "#7B68EE", "#20B2AA", "#FF69B4", "#8B4513"]

STAR_COLORS = {
    "1 star": "#FF4B4B",
    "2 star": "#FF8C42",
    "3 star": "#FFC000",
    "4 star": "#70AD47",
    "5 star": "#2E75B6",
}

GENERATED_PNG_NAMES = {
    # Current chart output names.
    "01_sdlc_tasks_by_stage.png",
    "02_staff_ai_effectiveness.png",
    "03_kpi_summary.png",
    "04_est_actual_tool.png",
    "05_est_actual_category.png",
    "06_rating_distribution.png",
    "07_top_errors.png",
    "08_error_heatmap.png",
    # Legacy chart output names from earlier ordering.
    "01_kpi_summary.png",
    "02_est_actual_staff.png",
    "03_est_actual_tool.png",
    "04_est_actual_category.png",
    "05_rating_distribution.png",
    "06_top_errors.png",
    "07_error_heatmap.png",
    "08_sdlc_tasks_by_stage.png",
}


def _first_present(mapping: dict[str, Any], *keys: str, default: Any = "—") -> Any:
    for key in keys:
        value = mapping.get(key)
        if value not in (None, ""):
            return value
    return default


def _clean_generated_pngs(out_dir: Path) -> None:
    for name in GENERATED_PNG_NAMES:
        path = out_dir / name
        if path.exists():
            path.unlink()


plt.rcParams.update({
    "figure.facecolor": WHITE,
    "axes.facecolor": WHITE,
    "axes.edgecolor": "#D9D9D9",
    "axes.labelcolor": GRAY,
    "axes.titleweight": "bold",
    "axes.titlesize": 14,
    "axes.labelsize": 11,
    "xtick.color": GRAY,
    "ytick.color": GRAY,
    "xtick.labelsize": 10,
    "ytick.labelsize": 10,
    "font.family": "sans-serif",
    "font.sans-serif": ["Helvetica Neue", "Arial", "Helvetica", "sans-serif"],
    "grid.color": "#E8E8E8",
    "grid.linewidth": 0.5,
    "legend.frameon": False,
    "legend.fontsize": 9,
})


# --------------------------------------------------------------------------- #
# Data reading from report xlsx
# --------------------------------------------------------------------------- #

def _read_table(ws, header_row: int, data_start: int, n_cols: int,
                skip_total: bool = True) -> list[dict[str, Any]]:
    """Read a sheet table into list of dicts using header_row as keys."""
    headers = [ws.cell(header_row, c).value for c in range(1, n_cols + 1)]
    rows = []
    for r in range(data_start, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, n_cols + 1)]
        if vals[0] is None:
            continue
        if skip_total and str(vals[0]).upper().startswith("TỔNG"):
            continue
        rows.append(dict(zip(headers, vals)))
    return rows


def read_report(path: Path) -> dict[str, Any]:
    """Extract all data needed for charts from the report xlsx."""
    wb = load_workbook(path, data_only=True)
    data: dict[str, Any] = {}

    # Per Staff (header=row4, data=row5+)
    if "👤 Per Staff" in wb.sheetnames:
        data["per_staff"] = _read_table(wb["👤 Per Staff"], 4, 5, 10)

    # Per Tool
    if "🔧 Per Tool" in wb.sheetnames:
        data["per_tool"] = _read_table(wb["🔧 Per Tool"], 4, 5, 9)

    # Per Category
    if "📂 Per Category" in wb.sheetnames:
        data["per_category"] = _read_table(wb["📂 Per Category"], 4, 5, 9)

    # Raw Log — need per-row data for rating distribution
    if "📝 Raw Log" in wb.sheetnames:
        data["raw_log"] = _read_table(wb["📝 Raw Log"], 4, 5, 14, skip_total=False)

    # AI Lesson Compare — for error classification
    if "🤖 AI Lesson Compare" in wb.sheetnames:
        data["ai_compare"] = _read_table(wb["🤖 AI Lesson Compare"], 4, 5, 15, skip_total=False)

    # Summary KPIs
    if "📊 Tổng Quan" in wb.sheetnames:
        ws = wb["📊 Tổng Quan"]
        kpis = {}
        for r in range(5, 14):
            k = ws.cell(r, 1).value
            v = ws.cell(r, 2).value
            if k:
                kpis[k] = v
        data["kpis"] = kpis

    return data


# --------------------------------------------------------------------------- #
# Error classification via OpenAI-compatible API
# --------------------------------------------------------------------------- #

ERROR_LABELS = [
    "Thiếu Output Format",
    "Thiếu Context / Motivation",
    "Prompt Không Rõ Ràng",
    "Thiếu Ví Dụ (Few-shot)",
    "Không Gán Role",
    "Thiếu XML Structure",
    "Thiếu Self-check",
    "Scope Mơ Hồ",
    "Thiếu Constraints",
    "Sai Thứ Tự Long-context",
]

CLASSIFY_PROMPT = """<role>
Bạn là chuyên gia phân loại lỗi prompt engineering theo Claude Prompting Best Practices.
</role>

<task>
Đọc danh sách các bài học (từ AI suy luận) về lỗi prompt của từng phiên làm việc.
Phân loại MỖI phiên vào ĐÚNG 1–3 nhãn lỗi từ danh sách bên dưới.
</task>

<error_labels>
1. Thiếu Output Format — không định nghĩa schema/format đầu ra
2. Thiếu Context / Motivation — không giải thích tại sao, bối cảnh nghiệp vụ
3. Prompt Không Rõ Ràng — yêu cầu chung chung, không cụ thể
4. Thiếu Ví Dụ (Few-shot) — task phức tạp mà không kèm ví dụ
5. Không Gán Role — không gán vai trò/persona cho AI
6. Thiếu XML Structure — không dùng XML tags tách phần
7. Thiếu Self-check — không yêu cầu AI tự kiểm tra
8. Scope Mơ Hồ — phạm vi task không xác định rõ
9. Thiếu Constraints — không nêu ràng buộc/giới hạn
10. Sai Thứ Tự Long-context — data dài đặt sai vị trí
</error_labels>

<sessions>
{sessions_json}
</sessions>

<output_format>
Trả về DUY NHẤT một JSON object có key "results" chứa array. Mỗi phần tử là object:
{{"results": [{{"id": 0, "labels": ["Thiếu Output Format", "Thiếu Context / Motivation"]}}, ...]}}

Chỉ dùng tên nhãn CHÍNH XÁC từ danh sách trên. Mỗi phiên 1–3 nhãn.
Trả JSON thuần, không markdown.
</output_format>"""


def _call_openai(model: str, prompt: str, timeout: int = 300) -> str:
    base = (OPENAI_BASE_URL or "").rstrip("/")
    is_azure = "cognitiveservices.azure.com" in base or "openai.azure.com" in base

    if is_azure:
        host = re.sub(r"/openai(/v1)?$", "", base).rstrip("/")
        url = f"{host}/openai/deployments/{model}/chat/completions?api-version=2024-12-01-preview"
        headers = {"api-key": OPENAI_API_KEY or ""}
    else:
        url = f"{base}/chat/completions"
        headers = {}
        if OPENAI_API_KEY:
            headers["Authorization"] = f"Bearer {OPENAI_API_KEY}"

    body: dict[str, Any] = {
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_completion_tokens": 4096,
        "response_format": {"type": "json_object"},
    }
    if not is_azure:
        body["model"] = model

    r = requests.post(url, headers=headers, json=body, timeout=timeout)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]


def classify_errors(ai_compare: list[dict], model: str) -> list[dict]:
    """Classify prompt errors from AI lessons using the LLM."""
    sessions_for_api = []
    for i, row in enumerate(ai_compare):
        ai_lesson = str(row.get("Bài Học AI Suy Luận") or "")
        user_lesson = str(row.get("Bài Học Người Dùng") or "")
        if not ai_lesson or ai_lesson == "—":
            continue
        sessions_for_api.append({
            "id": i,
            "staff": row.get("Staff", ""),
            "title": str(row.get("Tên Phiên") or "")[:80],
            "ai_lesson": ai_lesson[:300],
            "user_lesson": user_lesson[:200] if user_lesson != "(trống)" else "",
        })

    if not sessions_for_api:
        return []

    prompt = CLASSIFY_PROMPT.format(sessions_json=json.dumps(sessions_for_api, ensure_ascii=False))
    print(f"  🏷️  Classifying {len(sessions_for_api)} sessions with {model}...")

    try:
        raw = _call_openai(model, prompt)
        raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip(), flags=re.MULTILINE)
        parsed = json.loads(raw)

        # Unwrap: could be [...], {"results":[...]}, or a single {...}
        items: list[dict] = []
        if isinstance(parsed, list):
            items = parsed
        elif isinstance(parsed, dict):
            # Try common wrapper keys
            for key in ("results", "data", "sessions", "classifications"):
                if key in parsed and isinstance(parsed[key], list):
                    items = parsed[key]
                    break
            if not items and "id" in parsed:
                # Single object response
                items = [parsed]

        if not items:
            print(f"  ⚠  Unexpected response format: {type(parsed)}", file=sys.stderr)
            return []

        # Map back to original rows
        results = []
        for item in items:
            idx = item.get("id", -1)
            labels = item.get("labels", [])
            # Filter to valid labels only
            valid = [l for l in labels if l in ERROR_LABELS]
            if 0 <= idx < len(ai_compare):
                results.append({
                    "staff": ai_compare[idx].get("Staff", ""),
                    "title": ai_compare[idx].get("Tên Phiên", ""),
                    "labels": valid if valid else ["Prompt Không Rõ Ràng"],
                })
        print(f"  ✔  Classified {len(results)} sessions")
        return results
    except Exception as e:
        print(f"  ⚠  Classification failed: {e}", file=sys.stderr)
        return []


# --------------------------------------------------------------------------- #
# SDLC task data
# --------------------------------------------------------------------------- #

SDLC_DETAIL_SECTION = "All Task Details by SDLC Stage"


def _read_existing_sdlc_tasks(report_path: Path) -> list[dict[str, str]]:
    """Read task/stage rows when the input is already enriched by plot_charts.py."""
    wb = load_workbook(report_path, data_only=True)
    if "🧭 SDLC Summary" not in wb.sheetnames:
        return []

    ws = wb["🧭 SDLC Summary"]
    section_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == SDLC_DETAIL_SECTION:
            section_row = row
            break
    if section_row is None:
        return []

    header_row = section_row + 1
    headers = [
        str(ws.cell(row=header_row, column=col).value or "").strip()
        for col in range(1, ws.max_column + 1)
    ]
    header_map = {header: idx + 1 for idx, header in enumerate(headers) if header}
    required = {"SDLC Stage", "Task Name"}
    if not required.issubset(header_map):
        return []

    rows: list[dict[str, str]] = []
    for row in range(header_row + 1, ws.max_row + 1):
        stage = str(ws.cell(row=row, column=header_map["SDLC Stage"]).value or "").strip()
        task = str(ws.cell(row=row, column=header_map["Task Name"]).value or "").strip()
        if not stage and not task:
            break
        if not stage or not task:
            continue
        rows.append({
            "stage": stage if stage in SDLC_TAXONOMY else "Other",
            "task": task,
            "staff": str(ws.cell(row=row, column=header_map.get("Staff", 0)).value or "").strip()
            if header_map.get("Staff")
            else "",
            "date": str(ws.cell(row=row, column=header_map.get("Ngày", 0)).value or "").strip()
            if header_map.get("Ngày")
            else "",
        })
    return rows


def read_sdlc_task_rows(report_path: Path, model: str, no_ai: bool) -> list[dict[str, str]]:
    """Return task/stage rows for the SDLC PDF/PNG chart."""
    existing_rows = _read_existing_sdlc_tasks(report_path)
    if existing_rows:
        print(f"  🧭  Loaded {len(existing_rows)} SDLC task row(s) from existing SDLC Summary")
        return existing_rows

    wb = load_workbook(report_path)
    sessions = read_chart_sessions(wb)
    if no_ai:
        mark_sdlc_unclassified(sessions)
        print("  ⏭️  Skipping AI SDLC classification (--no-ai); SDLC tasks will be grouped as Other")
    else:
        classify_chart_sessions(
            sessions=sessions,
            model=model,
            cache_path=CHART_CLASSIFIER_CACHE_PATH,
            batch_size=20,
            timeout=300,
        )

    rows: list[dict[str, str]] = []
    for session in sessions:
        stage = session.sdlc_category if session.sdlc_category in SDLC_TAXONOMY else "Other"
        rows.append({
            "stage": stage,
            "task": session.title or session.task_desc or "(unnamed task)",
            "staff": session.staff,
            "date": session.date,
        })
    return rows


# --------------------------------------------------------------------------- #
# Chart builders
# --------------------------------------------------------------------------- #

def _add_watermark(fig: plt.Figure) -> None:
    fig.text(0.98, 0.02, "AI Dev Journal Report", fontsize=7, color="#BFBFBF",
             ha="right", va="bottom", style="italic")


def _save_chart(fig: plt.Figure, pdf: PdfPages, out_dir: Path, name: str) -> None:
    _add_watermark(fig)
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="This figure includes Axes that are not compatible with tight_layout.*",
            category=UserWarning,
        )
        fig.tight_layout(rect=[0, 0.03, 1, 0.97])
    pdf.savefig(fig, dpi=200)
    fig.savefig(out_dir / f"{name}.png", dpi=200, bbox_inches="tight",
                facecolor=WHITE, edgecolor="none")
    plt.close(fig)


def chart_kpi_summary(data: dict, pdf: PdfPages, out_dir: Path) -> None:
    """Executive KPI numbers overview."""
    kpis = data.get("kpis", {})
    per_staff = data.get("per_staff", [])

    fig, ax = plt.subplots(figsize=(14, 7))
    ax.axis("off")

    fig.text(0.5, 0.94, "AI DEV JOURNAL — EXECUTIVE SUMMARY",
             fontsize=22, fontweight="bold", color=NAVY, ha="center")

    # KPI boxes
    boxes = [
        ("Tổng Phiên", str(kpis.get("Tổng số phiên", "—")), BLUE),
        ("Số Staff", str(kpis.get("Số staff", "—")), TEAL),
        ("EST (Không AI)", str(kpis.get("Tổng EST (không AI)", "—")), ORANGE),
        ("Actual (Có AI)", str(kpis.get("Tổng Actual (có AI)", "—")), GREEN),
        ("Giờ Tiết Kiệm", str(kpis.get("Tổng giờ tiết kiệm", "—")), NAVY),
        ("Hiệu Suất AI", str(_first_present(kpis, "Hiệu suất AI", "Tỷ lệ thời gian tiết kiệm")), RED),
    ]

    for i, (label, value, color) in enumerate(boxes):
        col = i % 3
        row_idx = i // 3
        x = 0.10 + col * 0.30
        y = 0.62 - row_idx * 0.28

        rect = plt.Rectangle((x - 0.02, y - 0.05), 0.26, 0.22,
                              transform=fig.transFigure, facecolor=color,
                              alpha=0.10, edgecolor=color, linewidth=1.5,
                              clip_on=False, zorder=2)
        fig.patches.append(rect)

        fig.text(x + 0.11, y + 0.10, value, fontsize=26, fontweight="bold",
                 color=color, ha="center", va="center")
        fig.text(x + 0.11, y + 0.0, label, fontsize=10, color=GRAY,
                 ha="center", va="center")

    # Bottom: staff ranking mini-table
    if per_staff:
        table_y = 0.18
        fig.text(0.5, table_y + 0.06, "Ranking by Time Saved",
                 fontsize=11, fontweight="bold", color=NAVY, ha="center")
        header = f"{'Staff':<16} {'Phiên':>6} {'EST':>8} {'Actual':>8} {'Saved':>8} {'Eff%':>7}"
        fig.text(0.18, table_y, header, fontsize=9, fontfamily="monospace",
                 color=NAVY, fontweight="bold")
        for j, row in enumerate(per_staff[:6]):
            name = str(row.get("Staff", ""))[:14]
            eff = _first_present(row, "Hiệu Suất %", "Tỷ Lệ Tiết Kiệm %", default=0)
            line = (f"{name:<16} {row.get('Số Phiên', 0):>6} "
                    f"{row.get('EST (h)', 0):>7}h {row.get('Actual (h)', 0):>7}h "
                    f"{row.get('Tiết Kiệm (h)', 0):>7}h {eff:>6}%")
            fig.text(0.18, table_y - 0.035 * (j + 1), line, fontsize=8.5,
                     fontfamily="monospace", color=GRAY)

    _save_chart(fig, pdf, out_dir, "03_kpi_summary")


def _bar_est_actual(ax: plt.Axes, labels: list[str], est: list[float],
                    actual: list[float], saved: list[float]) -> None:
    """Draw grouped horizontal bars: EST vs Actual with saved annotation."""
    y = np.arange(len(labels))
    h = 0.35

    bars_est = ax.barh(y + h / 2, est, h, label="EST (Không AI)", color=ORANGE, alpha=0.85)
    bars_act = ax.barh(y - h / 2, actual, h, label="Actual (Có AI)", color=BLUE, alpha=0.85)

    max_val = max(est) if est else 1
    for i in range(len(labels)):
        if saved[i] > 0:
            ax.text(est[i] + max_val * 0.02, y[i] + h / 2,
                    f"-{saved[i]:.1f}h", va="center", fontsize=8,
                    color=GREEN, fontweight="bold")
        eff = (saved[i] / est[i] * 100) if est[i] else 0
        if eff > 0:
            ax.text(actual[i] + max_val * 0.02, y[i] - h / 2,
                    f"{eff:.0f}%", va="center", fontsize=8, color=BLUE, fontweight="bold")

    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.set_xlabel("Giờ (hours)")
    ax.legend(loc="lower right")
    ax.xaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    ax.grid(axis="x", alpha=0.3)
    ax.invert_yaxis()


def _chart_efficiency_view(
    rows: list[dict[str, Any]],
    label_col: str,
    title: str,
    filename: str,
    pdf: PdfPages,
    out_dir: Path,
) -> None:
    if not rows:
        return

    labels = [str(r.get(label_col, ""))[:25] for r in rows]
    est = [float(r.get("EST (h)", 0) or 0) for r in rows]
    actual = [float(r.get("Actual (h)", 0) or 0) for r in rows]
    saved = [float(r.get("Tiết Kiệm (h)", 0) or 0) for r in rows]

    fig_h = max(4.5, 1.2 * len(labels) + 2)
    fig, ax = plt.subplots(figsize=(12, fig_h))
    fig.suptitle(title, fontsize=16, fontweight="bold", color=NAVY, y=0.98)

    _bar_est_actual(ax, labels, est, actual, saved)

    total_saved = sum(saved)
    total_est = sum(est)
    eff = (total_saved / total_est * 100) if total_est else 0
    ax.set_title(f"Tổng tiết kiệm: {total_saved:.1f}h / {total_est:.1f}h ({eff:.0f}%)",
                 fontsize=11, color=GRAY, pad=10)

    _save_chart(fig, pdf, out_dir, filename)


def chart_staff_effectiveness(data: dict, pdf: PdfPages, out_dir: Path) -> None:
    """Page 2: how effectively each staff uses AI for these tasks."""
    _chart_efficiency_view(
        data.get("per_staff", []),
        "Staff",
        "Staff AI Effectiveness — These Tasks",
        "02_staff_ai_effectiveness",
        pdf,
        out_dir,
    )


def chart_efficiency(data: dict, pdf: PdfPages, out_dir: Path) -> None:
    """EST vs Actual grouped bars per tool/category."""
    views = [
        ("per_tool", "Công Cụ AI", "EST vs Actual — Theo Công Cụ AI", "04_est_actual_tool"),
        ("per_category", "Danh Mục", "EST vs Actual — Theo Danh Mục", "05_est_actual_category"),
    ]

    for key, label_col, title, filename in views:
        _chart_efficiency_view(data.get(key, []), label_col, title, filename, pdf, out_dir)


def chart_rating(data: dict, pdf: PdfPages, out_dir: Path) -> None:
    """Rating distribution stacked bar per tool."""
    raw = data.get("raw_log", [])
    if not raw:
        return

    # Count ratings per tool
    tool_ratings: dict[str, Counter] = defaultdict(Counter)
    for row in raw:
        tool = str(row.get("Công Cụ") or "(không rõ)")
        rating = row.get("Rating")
        if rating is not None:
            try:
                r = int(float(rating))
                if 1 <= r <= 5:
                    tool_ratings[tool][r] += 1
            except (ValueError, TypeError):
                pass

    if not tool_ratings:
        return

    tools = sorted(tool_ratings.keys(), key=lambda t: -sum(tool_ratings[t].values()))
    stars = [1, 2, 3, 4, 5]
    star_keys = [f"{s} star" for s in stars]
    star_display = ["1 Star", "2 Star", "3 Star", "4 Star", "5 Star"]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, max(4.5, len(tools) * 1.2 + 2)),
                                   gridspec_kw={"width_ratios": [2.5, 1]})
    fig.suptitle("User Satisfaction — Rating Distribution by AI Tool",
                 fontsize=16, fontweight="bold", color=NAVY, y=0.98)

    # Left: stacked horizontal bar
    y = np.arange(len(tools))
    left = np.zeros(len(tools))
    for s, key, display in zip(stars, star_keys, star_display):
        vals = [tool_ratings[t][s] for t in tools]
        ax1.barh(y, vals, left=left, label=display,
                 color=STAR_COLORS[key], alpha=0.85, height=0.6)
        for i, v in enumerate(vals):
            if v > 0:
                ax1.text(left[i] + v / 2, y[i], str(v), ha="center", va="center",
                         fontsize=8, fontweight="bold", color=WHITE)
        left += np.array(vals, dtype=float)

    ax1.set_yticks(y)
    ax1.set_yticklabels(tools)
    ax1.set_xlabel("Số phiên")
    ax1.legend(loc="lower right", ncol=5)
    ax1.grid(axis="x", alpha=0.3)
    ax1.invert_yaxis()

    # Right: average rating per tool
    avg_ratings = []
    for t in tools:
        counts = tool_ratings[t]
        total = sum(counts.values())
        avg = sum(s * c for s, c in counts.items()) / total if total else 0
        avg_ratings.append(avg)

    colors = [GREEN if a >= 4.5 else BLUE if a >= 3.5 else ORANGE if a >= 2.5 else RED
              for a in avg_ratings]
    bars = ax2.barh(y, avg_ratings, color=colors, alpha=0.85, height=0.5)
    for i, v in enumerate(avg_ratings):
        ax2.text(v + 0.05, y[i], f"{v:.1f}/5", va="center", fontsize=10,
                 fontweight="bold", color=colors[i])

    ax2.set_xlim(0, 5.5)
    ax2.set_xlabel("Rating trung bình")
    ax2.set_yticks(y)
    ax2.set_yticklabels([""] * len(tools))
    ax2.axvline(x=4, color=GREEN, linestyle="--", alpha=0.4, linewidth=1)
    ax2.grid(axis="x", alpha=0.3)
    ax2.invert_yaxis()
    ax2.set_title("Trung Bình", fontsize=11, color=GRAY)

    _save_chart(fig, pdf, out_dir, "06_rating_distribution")


def chart_errors(data: dict, error_data: list[dict], pdf: PdfPages, out_dir: Path) -> None:
    """Top error labels bar + staff×error heatmap."""
    if not error_data:
        return

    # Count total errors
    label_counter: Counter = Counter()
    staff_errors: dict[str, Counter] = defaultdict(Counter)
    for item in error_data:
        staff = item["staff"]
        for label in item["labels"]:
            label_counter[label] += 1
            staff_errors[staff][label] += 1

    if not label_counter:
        return

    # --- Top errors bar chart ---
    top_labels = label_counter.most_common(10)
    labels = [l for l, _ in top_labels]
    counts = [c for _, c in top_labels]

    fig, ax = plt.subplots(figsize=(12, max(5, len(labels) * 0.7 + 2)))
    fig.suptitle("Top Lỗi Prompt — Phân Loại Theo Best Practices",
                 fontsize=16, fontweight="bold", color=NAVY, y=0.98)

    y = np.arange(len(labels))
    max_c = max(counts)
    colors_bar = [plt.cm.Blues(0.4 + 0.5 * c / max_c) for c in counts]
    bars = ax.barh(y, counts, color=colors_bar, height=0.6, edgecolor=WHITE, linewidth=0.5)

    for i, (cnt, lbl) in enumerate(zip(counts, labels)):
        ax.text(cnt + max_c * 0.02, y[i], f"{cnt}", va="center",
                fontsize=10, fontweight="bold", color=NAVY)

    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=10)
    ax.set_xlabel("Số lần xuất hiện")
    ax.grid(axis="x", alpha=0.3)
    ax.invert_yaxis()
    ax.set_title(f"Tổng {sum(counts)} lỗi từ {len(error_data)} phiên",
                 fontsize=11, color=GRAY, pad=10)

    _save_chart(fig, pdf, out_dir, "07_top_errors")

    # --- Staff × Error heatmap ---
    staff_list = sorted(staff_errors.keys())
    if not staff_list or not labels:
        return

    matrix = np.zeros((len(staff_list), len(labels)))
    for i, staff in enumerate(staff_list):
        for j, label in enumerate(labels):
            matrix[i, j] = staff_errors[staff].get(label, 0)

    fig, ax = plt.subplots(figsize=(max(10, len(labels) * 1.2 + 2),
                                    max(4, len(staff_list) * 0.8 + 2.5)))
    fig.suptitle("Phân Bố Lỗi Prompt Theo Staff",
                 fontsize=16, fontweight="bold", color=NAVY, y=0.98)

    im = ax.imshow(matrix, cmap="Blues", aspect="auto", vmin=0)

    ax.set_xticks(np.arange(len(labels)))
    ax.set_xticklabels(labels, rotation=35, ha="right", fontsize=9)
    ax.set_yticks(np.arange(len(staff_list)))
    ax.set_yticklabels(staff_list, fontsize=11)

    # Annotate cells
    for i in range(len(staff_list)):
        for j in range(len(labels)):
            val = int(matrix[i, j])
            if val > 0:
                text_color = WHITE if val >= matrix.max() * 0.6 else NAVY
                ax.text(j, i, str(val), ha="center", va="center",
                        fontsize=11, fontweight="bold", color=text_color)

    cbar = fig.colorbar(im, ax=ax, shrink=0.7, pad=0.02)
    cbar.set_label("Số lần", fontsize=9)

    ax.set_title("Mỗi ô = số lần staff mắc lỗi tương ứng",
                 fontsize=10, color=GRAY, pad=12)

    _save_chart(fig, pdf, out_dir, "08_error_heatmap")


def _wrap_task_bullets(task_counts: Counter[str], width: int = 72) -> str:
    lines: list[str] = []
    for task_name, count in task_counts.items():
        suffix = f" (×{count})" if count > 1 else ""
        bullet = f"• {task_name}{suffix}"
        wrapped = textwrap.wrap(bullet, width=width, subsequent_indent="  ")
        lines.extend(wrapped or [bullet])
    return "\n".join(lines) if lines else "—"


def chart_sdlc_tasks(task_rows: list[dict[str, str]], pdf: PdfPages, out_dir: Path) -> None:
    """Page 1: SDLC stages on the Y-axis with tasks listed within each stage."""
    if not task_rows:
        return

    stage_tasks: dict[str, Counter[str]] = {stage: Counter() for stage in SDLC_TAXONOMY}
    for row in task_rows:
        stage = row.get("stage") or "Other"
        if stage not in SDLC_TAXONOMY:
            stage = "Other"
        task_name = row.get("task") or "(unnamed task)"
        stage_tasks[stage][task_name] += 1

    stages = [stage for stage in SDLC_TAXONOMY if stage_tasks[stage]]
    if not stages:
        return

    counts = [sum(stage_tasks[stage].values()) for stage in stages]
    max_count = max(counts) if counts else 1
    total_distinct_tasks = sum(len(stage_tasks[stage]) for stage in stages)
    fig_h = min(24, max(7, 2.5 + len(stages) * 0.8 + total_distinct_tasks * 0.22))

    fig, (ax, ax_text) = plt.subplots(
        1,
        2,
        figsize=(18, fig_h),
        gridspec_kw={"width_ratios": [1.2, 2.3], "wspace": 0.08},
    )
    fig.suptitle(
        "SDLC Stage to Tasks",
        fontsize=16,
        fontweight="bold",
        color=NAVY,
        y=0.98,
    )

    y = np.arange(len(stages))
    colors = [PALETTE[i % len(PALETTE)] for i in range(len(stages))]
    bars = ax.barh(y, counts, color=colors, alpha=0.85, height=0.6)
    for bar, count in zip(bars, counts):
        ax.text(
            count + max_count * 0.03,
            bar.get_y() + bar.get_height() / 2,
            f"{count}",
            va="center",
            fontsize=10,
            fontweight="bold",
            color=NAVY,
        )

    ax.set_yticks(y)
    ax.set_yticklabels(stages, fontsize=10)
    ax.set_xlabel("Task count")
    ax.set_xlim(0, max_count * 1.25)
    ax.grid(axis="x", alpha=0.3)
    ax.invert_yaxis()
    ax.set_title("SDLC stages", fontsize=11, color=GRAY, pad=10)

    ax_text.set_xlim(0, 1)
    ax_text.set_ylim(ax.get_ylim())
    ax_text.axis("off")
    ax_text.set_title("Tasks within each stage", fontsize=11, color=GRAY, pad=10)
    for idx, stage in enumerate(stages):
        ax_text.text(
            0.0,
            idx,
            _wrap_task_bullets(stage_tasks[stage]),
            va="center",
            ha="left",
            fontsize=8.2,
            color=GRAY,
            linespacing=1.25,
        )

    fig.text(
        0.5,
        0.04,
        "Left side uses SDLC Stage on the Y-axis; right side lists the task names grouped under each stage.",
        ha="center",
        fontsize=9,
        color=GRAY,
    )

    _save_chart(fig, pdf, out_dir, "01_sdlc_tasks_by_stage")


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main() -> int:
    ap = argparse.ArgumentParser(description="Generate CEO-ready charts from AI Dev Journal report.")
    ap.add_argument("report", type=Path, help="Input report .xlsx (output of report.py)")
    ap.add_argument("-o", "--output", type=Path, default=Path("charts_output"),
                    help="Output directory (default: charts_output)")
    ap.add_argument("--model", default="gpt-5.4-mini",
                    help="Model for error and SDLC classification (default: gpt-5.4-mini)")
    ap.add_argument("--no-ai", action="store_true",
                    help="Skip AI error and SDLC classification")
    args = ap.parse_args()

    if not args.report.exists():
        print(f"✖  File not found: {args.report}", file=sys.stderr)
        return 1

    out_dir = args.output
    out_dir.mkdir(parents=True, exist_ok=True)
    _clean_generated_pngs(out_dir)
    pdf_path = out_dir / "ai_journal_charts.pdf"

    print(f"📊  Reading report: {args.report}")
    data = read_report(args.report)

    # Error classification
    error_data: list[dict] = []
    if not args.no_ai and data.get("ai_compare"):
        error_data = classify_errors(data["ai_compare"], args.model)
    elif args.no_ai:
        print("  ⏭️  Skipping AI error classification (--no-ai)")

    sdlc_task_rows: list[dict[str, str]] = []
    try:
        sdlc_task_rows = read_sdlc_task_rows(args.report, args.model, args.no_ai)
    except Exception as exc:
        print(f"  ⚠  SDLC task chart data unavailable: {exc}", file=sys.stderr)

    print(f"📈  Generating charts → {out_dir}/")

    with PdfPages(str(pdf_path)) as pdf:
        chart_sdlc_tasks(sdlc_task_rows, pdf, out_dir)
        chart_staff_effectiveness(data, pdf, out_dir)
        chart_kpi_summary(data, pdf, out_dir)
        chart_efficiency(data, pdf, out_dir)
        chart_rating(data, pdf, out_dir)
        chart_errors(data, error_data, pdf, out_dir)

    pngs = list(out_dir.glob("*.png"))
    print(f"\n✔  Done!")
    print(f"   PDF: {pdf_path}")
    print(f"   PNGs: {len(pngs)} files in {out_dir}/")
    for p in sorted(pngs):
        print(f"     • {p.name}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
