#!/usr/bin/env python3
"""
AI Dev Journal — All-in-One Report & Charts Generator

Consolidation of report.py, plot_charts.py, charts.py, and report_with_charts.py
into a single script that:

  1. Parses input AI Dev Journal .xlsx files (one per staff)
  2. (Optional) Runs OpenAI-compatible AI lesson inference
  3. Builds a consolidated Excel report workbook with multiple analysis sheets
  4. Classifies prompt errors + SDLC stages via OpenAI API
  5. Adds Excel chart sheets and polishes the workbook
  6. Generates CEO-ready PDF + PNG visual charts via matplotlib
  7. Prints a terminal summary

Usage:
    python ai_journal.py data/*.xlsx -o report.xlsx
    python ai_journal.py data/*.xlsx -o report.xlsx --model gpt-5.4-mini
    python ai_journal.py data/*.xlsx -o report.xlsx --no-ai
    python ai_journal.py data/*.xlsx -o report.xlsx --skip-pdf
    python ai_journal.py data/*.xlsx -o report.xlsx --skip-charts
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import textwrap
import warnings
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.backends.backend_pdf import PdfPages
import numpy as np

try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass


# =========================================================================== #
#  Section 1 — Constants & Configuration
# =========================================================================== #

OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# Input sheet layout
NHAT_KY_SHEET = "📝 Nhật Ký"
HEADER_ROW = 3
DATA_START_ROW = 4

# Columns in Nhật Ký (1-indexed) — matches template.xlsx
COL_STT = 1
COL_DATE = 2
COL_TITLE = 3
COL_TOOL = 4
COL_CATEGORY = 5
COL_TASK_DESC = 6
COL_PROMPT = 7
COL_RESULT = 8
COL_QUALITY_TEXT = 9
COL_RATING = 10
COL_EST_HOURS = 11
COL_ACTUAL_HOURS = 12
COL_TIME_SAVED = 13
COL_USER_LESSON = 14
COL_TAGS = 15

# Sheet names
RAW_LOG_SHEET = "📝 Raw Log"
AI_COMPARE_SHEET = "🤖 AI Lesson Compare"
EFFICIENCY_SHEET = "📈 Efficiency Charts"
RATING_SHEET = "⭐ Rating Charts"
ERROR_DATA_SHEET = "🏷️ Prompt Error Data"
ERROR_CHART_SHEET = "🏷️ Error Charts"
SDLC_SHEET = "🧭 SDLC Summary"

GENERATED_SHEETS = [
    SDLC_SHEET, ERROR_DATA_SHEET, ERROR_CHART_SHEET,
    # legacy — removed but cleaned up if present from older runs
    "🧭 SDLC Data", EFFICIENCY_SHEET, RATING_SHEET,
]

# Cache
LESSON_CACHE_PATH = Path(".ai_journal_cache.json")
CHART_CACHE_PATH = Path(".ai_chart_cache.json")
TRANSLATE_CACHE_PATH = Path(".ai_translate_cache.json")
ESTIMATE_CACHE_PATH = Path(".ai_estimate_cache.json")
PROMPT_VERSION = "chart-classifier-v1"
MAX_ERROR_LABELS_FOR_CHART = 10
MAX_LABELS_PER_SESSION = 3

# --------------------------------------------------------------------------- #
# Excel styling
# --------------------------------------------------------------------------- #

HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name="Arial", italic=True, size=10, color="595959")
CELL_FONT = Font(name="Arial", size=10)
NUMBER_FONT = Font(name="Arial", size=10, bold=True, color="1F4E78")
TOTAL_FILL = PatternFill("solid", start_color="D9E1F2")
TOTAL_FONT = Font(name="Arial", bold=True, size=10)
SECTION_FILL = PatternFill("solid", start_color="DDEBF7")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Professional polish styling
PROFESSIONAL_NAVY = "17365D"
PROFESSIONAL_BLUE = "1F4E78"
PROFESSIONAL_TEAL = "00A6A6"
PROFESSIONAL_ORANGE = "ED7D31"
PROFESSIONAL_LIGHT_BLUE = "DDEBF7"
PROFESSIONAL_ALT_FILL = PatternFill("solid", start_color="F7FBFF")
PROFESSIONAL_HEADER_FILL = PatternFill("solid", start_color=PROFESSIONAL_NAVY)
PROFESSIONAL_SECTION_FILL = PatternFill("solid", start_color=PROFESSIONAL_LIGHT_BLUE)
PROFESSIONAL_HEADER_FONT = Font(name="Aptos", bold=True, color="FFFFFF", size=11)
PROFESSIONAL_TITLE_FONT = Font(name="Aptos Display", bold=True, size=16, color=PROFESSIONAL_NAVY)
PROFESSIONAL_SUBTITLE_FONT = Font(name="Aptos", italic=True, size=10, color="666666")
PROFESSIONAL_BODY_FONT = Font(name="Aptos", size=10, color="1F1F1F")

# --------------------------------------------------------------------------- #
# Matplotlib theme
# --------------------------------------------------------------------------- #

NAVY = "#1F4E78"
BLUE = "#2E75B6"
LIGHT_BLUE = "#9DC3E6"
TEAL = "#00B0F0"
GREEN = "#70AD47"
ORANGE = "#ED7D31"
RED = "#FF4B4B"
GOLD = "#FFC000"
GRAY = "#595959"
LIGHT_GRAY = "#F2F2F2"
WHITE = "#FFFFFF"

PALETTE = [BLUE, ORANGE, GREEN, TEAL, GOLD, RED, "#7B68EE", "#20B2AA", "#FF69B4", "#8B4513"]

STAR_COLORS = {
    "1 star": "#FF4B4B", "2 star": "#FF8C42", "3 star": "#FFC000",
    "4 star": "#70AD47", "5 star": "#2E75B6",
}

GENERATED_PNG_NAMES = {
    "01_sdlc_tasks_by_stage.png", "02_staff_ai_effectiveness.png",
    "03_kpi_summary.png", "04_est_actual_tool.png", "05_est_actual_category.png",
    "06_rating_distribution.png", "07_top_errors.png", "08_error_heatmap.png",
    "09_user_vs_ai_comparison.png",
    "01_kpi_summary.png", "02_est_actual_staff.png", "03_est_actual_tool.png",
    "04_est_actual_category.png", "05_rating_distribution.png",
    "06_top_errors.png", "07_error_heatmap.png", "08_sdlc_tasks_by_stage.png",
}

plt.rcParams.update({
    "figure.facecolor": WHITE, "axes.facecolor": WHITE,
    "axes.edgecolor": "#D9D9D9", "axes.labelcolor": GRAY,
    "axes.titleweight": "bold", "axes.titlesize": 14,
    "axes.labelsize": 11, "xtick.color": GRAY, "ytick.color": GRAY,
    "xtick.labelsize": 10, "ytick.labelsize": 10,
    "font.family": "sans-serif",
    "font.sans-serif": ["Helvetica Neue", "Arial", "Helvetica", "sans-serif"],
    "grid.color": "#E8E8E8", "grid.linewidth": 0.5,
    "legend.frameon": False, "legend.fontsize": 9,
})

# --------------------------------------------------------------------------- #
# Error & SDLC taxonomies
# --------------------------------------------------------------------------- #

ERROR_TAXONOMY: dict[str, str] = {
    "Clear and Format": "Prompt lacks clear requirements, output format/schema, or sequential steps/constraints.",
    "Missing Context": "Prompt lacks business context, tech stack, background data, goals, or target audience.",
    "Missing Examples": "Complex/repeated task but prompt lacks few-shot examples or examples are not diverse enough.",
    "Weak Structure": "Prompt mixes instruction, context, input, output without clear structure like XML tags.",
    "No Role": "Prompt does not assign an appropriate expert role/persona to the AI.",
    "Negative Instruction": "Prompt focuses on what NOT to do instead of positively guiding what AI should do.",
    "Missing Grounding": "Document/data analysis task but prompt does not ask AI to quote or ground in evidence.",
    "No Self-check": "Prompt does not ask AI to verify its own output before responding.",
    "Long Context Ordering": "Prompt has long input/documents but places questions or instructions in a position that causes noise.",
    "Ambiguous Scope": "Prompt lacks scope, completion criteria, constraints, or boundaries so AI may infer incorrectly.",
    "Tool Or Environment Missing": "Prompt lacks info about tool, runtime environment, framework, file/schema, or input data.",
    "Insufficient Lesson Data": "Both user lesson and AI lesson are missing or too vague to reliably infer prompt errors.",
}

SDLC_TAXONOMY: dict[str, str] = {
    "Planning / Requirements": "Clarifying goals, requirements, acceptance criteria, scope, or user needs.",
    "Design / Architecture": "System design, data model design, API design, architecture, UX flow design.",
    "Development / Implementation": "Writing new code or implementing a feature.",
    "Testing / QA": "Creating, running, or improving tests; manual QA; validation.",
    "Debugging / Bug Fix": "Investigating or fixing bugs, errors, regressions, or broken behavior.",
    "Refactoring / Code Quality": "Cleanup, refactor, simplification, style, maintainability, performance tuning.",
    "Deployment / Release": "Build, packaging, release, CI/CD deployment, migration rollout.",
    "Operations / Maintenance": "Monitoring, support, environment maintenance, dependency updates, ops tasks.",
    "Documentation": "Writing or improving docs, comments, guides, summaries, specs.",
    "Project Management / Collaboration": "Planning work, reporting status, coordinating tasks, reviews.",
    "Research / Learning": "Exploration, technical research, comparing options, learning a tool or concept.",
    "Other": "Use only when no listed SDLC category fits.",
}

# Error labels used by charts.py-style classification
ERROR_LABELS = [
    "Missing Output Format", "Missing Context / Motivation", "Unclear Prompt",
    "Missing Examples (Few-shot)", "No Role Assignment", "Missing XML Structure",
    "No Self-check", "Ambiguous Scope", "Missing Constraints", "Wrong Long-context Order",
]


# =========================================================================== #
#  Section 2 — Shared Utilities
# =========================================================================== #

def _to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    text = str(v).strip().replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except (ValueError, TypeError):
        return None


def _fmt_date(d: Any) -> str:
    if d is None:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%d/%m/%Y")
    return str(d).strip()


def _truncate(s: str, n: int = 1200) -> str:
    return s if len(s) <= n else s[:n] + "…"


def _cell(ws, row: int, col: int) -> Any:
    v = ws.cell(row=row, column=col).value
    if isinstance(v, str):
        v = v.strip()
        if v == "":
            return None
    return v


def _load_json_cache(path: Path) -> dict:
    if path.exists():
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(raw, dict):
                return raw
        except Exception:
            pass
    return {}


def _save_json_cache(path: Path, cache: dict) -> None:
    try:
        path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        print(f"  ⚠  Failed to write cache {path}: {e}", file=sys.stderr)


def _call_openai(model: str, prompt: str, timeout: int = 300) -> str:
    """Call OpenAI-compatible API. Supports standard OpenAI and Azure endpoints."""
    base = (OPENAI_BASE_URL or "").rstrip("/")
    if not base:
        raise RuntimeError("OPENAI_BASE_URL is required unless --no-ai is used")

    is_azure = "cognitiveservices.azure.com" in base or "openai.azure.com" in base

    endpoints: list[tuple[str, dict[str, str], bool]] = []
    if is_azure:
        host = re.sub(r"/openai(/v1)?$", "", base).rstrip("/")
        endpoints.append((
            f"{host}/openai/deployments/{model}/chat/completions?api-version=2024-12-01-preview",
            {"api-key": OPENAI_API_KEY or ""}, False,
        ))
        endpoints.append((
            f"{base}/chat/completions",
            {"api-key": OPENAI_API_KEY or ""}, True,
        ))
    else:
        headers = {}
        if OPENAI_API_KEY:
            headers["Authorization"] = f"Bearer {OPENAI_API_KEY}"
        endpoints.append((f"{base}/chat/completions", headers, True))

    body_template: dict[str, Any] = {
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.1,
        "max_completion_tokens": 4096,
        "response_format": {"type": "json_object"},
    }

    last_error: requests.RequestException | None = None
    for url, headers, include_model in endpoints:
        body = dict(body_template)
        if include_model:
            body["model"] = model

        variants: list[dict[str, Any]] = [body]
        no_rf = dict(body)
        no_rf.pop("response_format", None)
        variants.append(no_rf)
        mt = dict(no_rf)
        mt["max_tokens"] = mt.pop("max_completion_tokens")
        variants.append(mt)

        for candidate in variants:
            try:
                r = requests.post(url, headers=headers, json=candidate, timeout=timeout)
                r.raise_for_status()
                return r.json()["choices"][0]["message"]["content"]
            except requests.HTTPError as exc:
                last_error = exc
                status = exc.response.status_code if exc.response is not None else None
                if status not in {400, 404, 422}:
                    raise
            except requests.RequestException as exc:
                last_error = exc
                raise
    assert last_error is not None
    raise last_error


# =========================================================================== #
#  Section 3 — Data Model
# =========================================================================== #

@dataclass
class Session:
    staff: str
    source_file: str
    stt: Any = None
    date: Any = None
    title: str = ""
    tool: str = ""
    category: str = ""
    task_desc: str = ""
    prompt: str = ""
    result: str = ""
    quality_text: str = ""
    rating: float | None = None
    est_hours: float | None = None
    actual_hours: float | None = None
    time_saved: float | None = None
    user_lesson: str = ""
    tags: str = ""
    # AI hour estimation fields
    ai_est_hours: float | None = None
    ai_actual_hours: float | None = None
    ai_est_reason: str = ""
    # AI lesson inference fields
    ai_lesson: str = ""
    comparison: str = ""
    ai_rating: float | None = None
    ai_rating_reason: str = ""
    suggested_prompt: str = ""
    # Chart classification fields
    row_id: str = ""
    error_labels: list[str] = field(default_factory=list)
    error_evidence: str = ""
    sdlc_category: str = "Other"
    sdlc_confidence: float | None = None
    sdlc_reason: str = ""

    def row_hash(self) -> str:
        key = "|".join([self.staff, str(self.date), self.title,
                        self.task_desc, self.prompt, self.result, self.user_lesson])
        return hashlib.sha1(key.encode("utf-8")).hexdigest()

    @property
    def cache_hash(self) -> str:
        payload = {
            "version": PROMPT_VERSION, "staff": self.staff,
            "date": _fmt_date(self.date), "title": self.title,
            "tool": self.tool, "category": self.category,
            "task_desc": self.task_desc, "prompt": self.prompt,
            "result": self.result, "user_lesson": self.user_lesson,
            "ai_lesson": self.ai_lesson,
        }
        raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
        return hashlib.sha1(raw.encode("utf-8")).hexdigest()

    @property
    def efficiency(self) -> float | None:
        if self.est_hours and self.time_saved is not None:
            return self.time_saved / self.est_hours
        return None

    @property
    def saved_hours(self) -> float | None:
        if self.time_saved is not None:
            return self.time_saved
        if self.est_hours is not None and self.actual_hours is not None:
            return self.est_hours - self.actual_hours
        return None


# =========================================================================== #
#  Section 4 — Phase 1: Parse Input Files
# =========================================================================== #

def staff_from_filename(path: Path) -> str:
    stem = path.stem
    parts = re.split(r"[_\-\s]+", stem)
    blacklist = {"journal", "ai", "dev", "nhatky", "nhật", "ký", "log"}
    meaningful = [p for p in parts if p.lower() not in blacklist and p]
    return (meaningful[-1] if meaningful else stem).strip().capitalize()


def parse_file(path: Path) -> list[Session]:
    wb = load_workbook(path, data_only=True)
    if NHAT_KY_SHEET not in wb.sheetnames:
        print(f"  ⚠  {path.name}: no '{NHAT_KY_SHEET}' sheet, skipping", file=sys.stderr)
        return []
    ws = wb[NHAT_KY_SHEET]
    staff = staff_from_filename(path)
    sessions: list[Session] = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        title = _cell(ws, row, COL_TITLE)
        tool = _cell(ws, row, COL_TOOL)
        if title is None and tool is None:
            continue
        sessions.append(Session(
            staff=staff,
            source_file=path.name,
            stt=_cell(ws, row, COL_STT),
            date=_cell(ws, row, COL_DATE),
            title=str(title or ""),
            tool=str(tool or ""),
            category=str(_cell(ws, row, COL_CATEGORY) or ""),
            task_desc=str(_cell(ws, row, COL_TASK_DESC) or ""),
            prompt=str(_cell(ws, row, COL_PROMPT) or ""),
            result=str(_cell(ws, row, COL_RESULT) or ""),
            quality_text=str(_cell(ws, row, COL_QUALITY_TEXT) or ""),
            rating=_to_float(_cell(ws, row, COL_RATING)),
            est_hours=_to_float(_cell(ws, row, COL_EST_HOURS)),
            actual_hours=_to_float(_cell(ws, row, COL_ACTUAL_HOURS)),
            time_saved=_to_float(_cell(ws, row, COL_TIME_SAVED)),
            user_lesson=str(_cell(ws, row, COL_USER_LESSON) or ""),
            tags=str(_cell(ws, row, COL_TAGS) or ""),
        ))
    return sessions


# =========================================================================== #
#  Section 4b — Translation to English
# =========================================================================== #

TRANSLATE_FIELDS = ["title", "tool", "category", "task_desc", "prompt", "result",
                    "quality_text", "user_lesson", "tags"]

TRANSLATE_PROMPT = """<role>You are a professional translator specializing in software development.</role>

<task>
Translate the following Vietnamese text fields to natural, professional English.
You MUST return ALL fields for every item, even if a field is already in English — return it unchanged.
</task>

<rules>
- Translate all Vietnamese text to English.
- Keep technical terms, code snippets, file paths, and variable names unchanged.
- If a field is already in English, return it as-is (do NOT omit it).
- If a field is empty, return empty string.
- Use standard software development terminology in English.
- Do NOT add explanations — just translate.
- You MUST include every field (title, tool, category, task_desc, prompt, result, quality_text, user_lesson, tags) in every result object.
</rules>

<input>
{sessions_json}
</input>

<output_format>
Return ONLY a valid JSON object, no markdown:
{{"results": [{{"id": 0, "title": "...", "tool": "...", "category": "...", "task_desc": "...", "prompt": "...", "result": "...", "quality_text": "...", "user_lesson": "...", "tags": "..."}}, ...]}}

IMPORTANT: Every result object MUST contain ALL 9 fields. Do not skip any field.
</output_format>"""


def _translate_hash(s: Session) -> str:
    key = "|".join([getattr(s, f) or "" for f in TRANSLATE_FIELDS])
    return hashlib.sha1(key.encode("utf-8")).hexdigest()


def translate_sessions_batch(sessions: list[Session], model: str,
                             batch_size: int = 5) -> None:
    """Translate Vietnamese text fields to English using LLM."""
    cache = _load_json_cache(TRANSLATE_CACHE_PATH)

    # Compute hashes before any modifications
    hashes: dict[int, str] = {}
    misses: list[Session] = []
    hits = 0

    for s in sessions:
        h = _translate_hash(s)
        hashes[id(s)] = h
        if h in cache:
            c = cache[h]
            for f in TRANSLATE_FIELDS:
                if f in c:
                    setattr(s, f, c[f])
            hits += 1
        else:
            misses.append(s)

    total = len(sessions)
    print(f"\n🌐  Translating {total} sessions to English ({model})...")
    print(f"   Cache hits: {hits}/{total}, translating {len(misses)} session(s)...")

    for start in range(0, len(misses), batch_size):
        batch = misses[start:start + batch_size]
        print(f"  • Batch {start // batch_size + 1}: {len(batch)} session(s)")

        items = []
        for i, s in enumerate(batch):
            item: dict[str, Any] = {"id": i}
            for f in TRANSLATE_FIELDS:
                item[f] = _truncate(getattr(s, f) or "", 2000)
            items.append(item)

        prompt = TRANSLATE_PROMPT.format(
            sessions_json=json.dumps(items, ensure_ascii=False))

        try:
            raw = _call_openai(model, prompt, timeout=300)
            parsed = _parse_json_object(raw)

            results: list[dict] = []
            if isinstance(parsed, list):
                results = parsed
            elif isinstance(parsed, dict):
                for key in ("results", "data", "sessions"):
                    if key in parsed and isinstance(parsed[key], list):
                        results = parsed[key]
                        break
                if not results and "id" in parsed:
                    results = [parsed]

            # Mark which sessions got a response
            seen_idx: set[int] = set()
            for item in results:
                idx = item.get("id", -1)
                if not (0 <= idx < len(batch)):
                    continue
                seen_idx.add(idx)
                s = batch[idx]
                h = hashes[id(s)]
                translated: dict[str, str] = {}
                for f in TRANSLATE_FIELDS:
                    if f in item and item[f]:
                        val = str(item[f]).strip()
                        if val:
                            setattr(s, f, val)
                            translated[f] = val
                    else:
                        # Keep original value in cache so we don't re-query
                        translated[f] = getattr(s, f) or ""
                cache[h] = translated

            # Cache sessions the LLM didn't return (keep originals)
            for idx, s in enumerate(batch):
                if idx not in seen_idx:
                    h = hashes[id(s)]
                    cache[h] = {f: getattr(s, f) or "" for f in TRANSLATE_FIELDS}
        except Exception as e:
            print(f"  ⚠  Translation batch failed: {e}", file=sys.stderr)

    # Retry pass: catch any fields the LLM left in Vietnamese
    vn_re = re.compile(r'[àáảãạăắằẳẵặâấầẩẫậèéẻẽẹêếềểễệìíỉĩịòóỏõọôốồổỗộơớờởỡợùúủũụưứừửữựỳýỷỹỵđ]', re.I)
    retry: list[Session] = []
    for s in sessions:
        for f in TRANSLATE_FIELDS:
            if vn_re.search(getattr(s, f) or ""):
                retry.append(s)
                break

    if retry:
        print(f"  🔄  Retrying {len(retry)} session(s) with remaining Vietnamese...")
        for s in retry:
            item = {"id": 0}
            for f in TRANSLATE_FIELDS:
                item[f] = _truncate(getattr(s, f) or "", 2000)
            prompt = TRANSLATE_PROMPT.format(
                sessions_json=json.dumps([item], ensure_ascii=False))
            try:
                raw = _call_openai(model, prompt, timeout=300)
                parsed = _parse_json_object(raw)
                results_list = parsed.get("results", [parsed]) if isinstance(parsed, dict) else parsed
                if results_list:
                    r_item = results_list[0] if isinstance(results_list, list) else results_list
                    # Use the ORIGINAL hash stored before any modifications
                    orig_h = hashes[id(s)]
                    translated: dict[str, str] = {}
                    for f in TRANSLATE_FIELDS:
                        if f in r_item and r_item[f]:
                            val = str(r_item[f]).strip()
                            if val:
                                setattr(s, f, val)
                                translated[f] = val
                            else:
                                translated[f] = getattr(s, f) or ""
                        else:
                            translated[f] = getattr(s, f) or ""
                    cache[orig_h] = translated
            except Exception as e:
                print(f"  ⚠  Retry failed: {e}", file=sys.stderr)

    _save_json_cache(TRANSLATE_CACHE_PATH, cache)
    print(f"✔  Translation done. Cache: {TRANSLATE_CACHE_PATH}")


def _parse_json_object(raw: str) -> dict[str, Any]:
    text = re.sub(r"^```(?:json)?\s*|\s*```$", "", (raw or "").strip(), flags=re.MULTILINE)
    try:
        parsed = json.loads(text)
        if isinstance(parsed, (dict, list)):
            return parsed
    except json.JSONDecodeError:
        pass
    match = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if not match:
        raise ValueError("Model response did not contain a JSON object")
    parsed = json.loads(match.group(0))
    return parsed


# =========================================================================== #
#  Section 4c — AI Hour Estimation
# =========================================================================== #

ESTIMATE_PROMPT = """<role>
You are a senior engineering manager with 15+ years of experience estimating software development tasks.
You provide independent, objective hour estimates — your goal is to give a realistic benchmark so management can compare against the staff member's own self-reported estimates.
</role>

<staff_profile>
{profile_json}
</staff_profile>

<task>
For each task below, provide YOUR OWN independent estimate (do NOT look at the user's self-reported hours — they are intentionally hidden from you):

- `ai_est_hours`: How many hours would this task realistically take the staff member described above if they did it MANUALLY without any AI tools? Factor in their role, experience level, and tech stack familiarity.
- `ai_actual_hours`: How many hours would this task take WITH the AI tool listed? Factor in prompt iteration time, review time, and the tool's effectiveness for this type of task.
- `ai_est_reason`: One sentence explaining your reasoning — mention the key factors (task complexity, staff experience, AI tool fit).

Guidelines:
- Be realistic, not optimistic. Junior devs take 2-3x longer than seniors on the same task.
- Simple documentation tasks: 1-4h manual, 0.5-2h with AI.
- Complex architecture/design: 4-16h manual, 2-8h with AI.
- Bug fixes vary widely by complexity: 0.5-8h manual.
- AI tools help most with boilerplate, docs, and well-defined tasks; less with novel architecture or debugging.
- Round to nearest 0.5h.
</task>

<sessions>
{sessions_json}
</sessions>

<output_format>
Return ONLY a valid JSON object, no markdown:
{{"results": [{{"id": 0, "ai_est_hours": 8.0, "ai_actual_hours": 3.0, "ai_est_reason": "..."}}, ...]}}

IMPORTANT: Return ALL sessions. Use numeric values (float). Do not skip any.
</output_format>"""


def _load_profiles(path: Path | None) -> dict[str, dict[str, str]]:
    """Load staff profiles from YAML or JSON file. Keys are staff names."""
    if path is None or not path.exists():
        return {}
    try:
        text = path.read_text(encoding="utf-8")
        if path.suffix.lower() in (".yml", ".yaml"):
            try:
                import yaml
                raw = yaml.safe_load(text)
            except ImportError:
                print("  ⚠  PyYAML not installed. Install with: pip install pyyaml", file=sys.stderr)
                return {}
        else:
            raw = json.loads(text)
        if isinstance(raw, dict):
            return {k: v for k, v in raw.items() if isinstance(v, dict)}
    except Exception as e:
        print(f"  ⚠  Failed to load profiles from {path}: {e}", file=sys.stderr)
    return {}


def _match_profile(staff: str, profiles: dict[str, dict[str, str]]) -> dict[str, str]:
    """Case-insensitive profile lookup."""
    for key, profile in profiles.items():
        if key.strip().casefold() == staff.strip().casefold():
            return profile
    return {}


def _estimate_hash(s: Session) -> str:
    key = "|".join([s.staff, str(s.date), s.title, s.tool, s.category, s.task_desc])
    return hashlib.sha1(key.encode("utf-8")).hexdigest()


def estimate_hours_batch(sessions: list[Session], model: str,
                         profiles: dict[str, dict[str, str]],
                         batch_size: int = 5) -> None:
    """Use LLM to estimate EST and Actual hours from AI's perspective."""
    cache = _load_json_cache(ESTIMATE_CACHE_PATH)
    hashes: dict[int, str] = {}
    misses: list[Session] = []
    hits = 0

    for s in sessions:
        h = _estimate_hash(s)
        hashes[id(s)] = h
        if h in cache:
            c = cache[h]
            s.ai_est_hours = _to_float(c.get("ai_est_hours"))
            s.ai_actual_hours = _to_float(c.get("ai_actual_hours"))
            s.ai_est_reason = str(c.get("ai_est_reason") or "")
            hits += 1
        else:
            misses.append(s)

    total = len(sessions)
    print(f"\n⏱️  AI hour estimation ({model}) — {total} sessions...")
    print(f"   Cache hits: {hits}/{total}, estimating {len(misses)} session(s)...")

    for start in range(0, len(misses), batch_size):
        batch = misses[start:start + batch_size]
        print(f"  • Batch {start // batch_size + 1}: {len(batch)} session(s)")

        # Group by staff to include the right profile
        staff_name = batch[0].staff
        profile = _match_profile(staff_name, profiles)
        profile_info = {
            "name": staff_name,
            **profile,
        }
        if not profile:
            profile_info["note"] = "No profile provided — estimate based on task complexity alone."

        items = []
        for i, s in enumerate(batch):
            items.append({
                "id": i,
                "title": _truncate(s.title, 200),
                "tool": s.tool,
                "category": s.category,
                "task_desc": _truncate(s.task_desc, 500),
            })

        prompt = ESTIMATE_PROMPT.format(
            profile_json=json.dumps(profile_info, ensure_ascii=False),
            sessions_json=json.dumps(items, ensure_ascii=False))

        try:
            raw = _call_openai(model, prompt, timeout=300)
            parsed = _parse_json_object(raw)

            results: list[dict] = []
            if isinstance(parsed, list):
                results = parsed
            elif isinstance(parsed, dict):
                for key in ("results", "data", "sessions"):
                    if key in parsed and isinstance(parsed[key], list):
                        results = parsed[key]
                        break
                if not results and "id" in parsed:
                    results = [parsed]

            for item in results:
                idx = item.get("id", -1)
                if not (0 <= idx < len(batch)):
                    continue
                s = batch[idx]
                h = hashes[id(s)]
                s.ai_est_hours = _to_float(item.get("ai_est_hours"))
                s.ai_actual_hours = _to_float(item.get("ai_actual_hours"))
                s.ai_est_reason = str(item.get("ai_est_reason") or "").strip()
                cache[h] = {
                    "ai_est_hours": s.ai_est_hours,
                    "ai_actual_hours": s.ai_actual_hours,
                    "ai_est_reason": s.ai_est_reason,
                }
        except Exception as e:
            print(f"  ⚠  Estimation batch failed: {e}", file=sys.stderr)

    _save_json_cache(ESTIMATE_CACHE_PATH, cache)
    print(f"✔  Estimation done. Cache: {ESTIMATE_CACHE_PATH}")


# =========================================================================== #
#  Section 5 — AI Lesson Inference (Phase 1 AI)
# =========================================================================== #

LESSON_PROMPT_TEMPLATE = """<role>
You are an expert in prompt engineering, well-versed in Anthropic's "Claude Prompting Best Practices".
Your task: analyze a real AI work session, score the output, and rewrite the prompt following best practices.
</role>

<best_practices_rubric>
Use EXACTLY these principles (Anthropic Claude Prompting Best Practices) as your scoring rubric:

1. **Clear & Direct** — Does the prompt specify output format, constraints, and sequential steps?
2. **Context & Motivation** — Does the prompt explain *why* (goals, target audience, business context)?
3. **Examples (few-shot)** — For complex/repeated tasks, are 2–5 diverse examples included?
4. **XML Structure** — Are different sections separated with consistent XML tags?
5. **Role Assignment** — Is a specific role/persona assigned to the AI?
6. **Long-context Ordering** — For long inputs, is the data placed ABOVE the question/instructions?
7. **Positive Instructions** — Does it tell AI *what to do* rather than *what not to do*?
8. **Ground in Quotes** — For document analysis, does it ask AI to quote relevant parts first?
9. **Self-check** — Does it ask AI to verify its own output before finishing?
10. **Output Format Specification** — Is the output schema clearly defined?
</best_practices_rubric>

<session_to_analyze>
<title>{title}</title>
<tool>{tool}</tool>
<category>{category}</category>
<task_description>{task_desc}</task_description>
<user_prompt>{prompt}</user_prompt>
<ai_result>{result}</ai_result>
<user_self_lesson>{user_lesson}</user_self_lesson>
</session_to_analyze>

<instructions>
Perform the following steps in order:

1. **Read <user_prompt>** and check which principles from <best_practices_rubric> it meets or violates. Identify the 1–2 MOST SEVERELY VIOLATED principles.

2. **Write `ai_lesson`** (English, 2–3 sentences, EXTREMELY SPECIFIC):
   - Name the violated principle (name + number from rubric).
   - Quote or describe exactly which part of <user_prompt> is missing/weak.
   - Briefly explain the observed consequence in <ai_result>.
   - Do NOT write generically — specify exactly WHAT context is missing.

3. **Compare with <user_self_lesson>** → choose ONE label for `comparison`:
   - "Agree" — your lesson and user's align on the main principle.
   - "Supplement" — you identify additional principles the user missed.
   - "Disagree" — you and user point to different principles.
   - "User left blank" — <user_self_lesson> is empty or "(empty)".

4. **Score `ai_rating` 1–5** how well <ai_result> meets <task_description>:
   - 1=fail, 2=poor, 3=average, 4=good, 5=excellent.
   - Write `ai_rating_reason` (1 sentence in English).

5. **Write `suggested_prompt`** — improved version of <user_prompt>, APPLYING all relevant best practices. MUST include:
   - Start with <role>...</role> assigning a specific role.
   - Include <context>...</context> with background and tech stack.
   - Include <task>...</task> listing requirements with numbered steps.
   - Include <constraints>...</constraints> with clear constraints.
   - Include <output_format>...</output_format> describing expected schema.
   - If task needs examples: include <examples><example>...</example></examples>.
   - End with 1 sentence asking AI to self-check before responding.
   - Prompt in ENGLISH, ready to copy-paste, 200–500 words.
   - Do NOT use placeholders — fill in real data from <task_description>.
</instructions>

<output_format>
Return ONLY a single valid JSON object, no markdown, no extra explanation:
{{"ai_lesson": "...", "comparison": "...", "ai_rating": 4, "ai_rating_reason": "...", "suggested_prompt": "..."}}
</output_format>

Before responding, self-check: (a) does ai_lesson name a specific principle? (b) does suggested_prompt have all 5 required XML tags? (c) is the JSON valid?"""


def _parse_lesson_response(raw: str) -> tuple[str, str, float | None, str, str]:
    if not raw:
        return "", "", None, "", ""
    raw = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw.strip(), flags=re.MULTILINE)

    def _extract(obj: dict) -> tuple[str, str, float | None, str, str]:
        ai_lesson = str(obj.get("ai_lesson", "")).strip()
        comparison = str(obj.get("comparison", "")).strip()
        rating_raw = obj.get("ai_rating")
        try:
            rating = float(rating_raw) if rating_raw is not None else None
            if rating is not None:
                rating = max(1.0, min(5.0, rating))
        except (ValueError, TypeError):
            rating = None
        reason = str(obj.get("ai_rating_reason", "")).strip()
        suggested = str(obj.get("suggested_prompt", "")).strip()
        return ai_lesson, comparison, rating, reason, suggested

    try:
        return _extract(json.loads(raw))
    except json.JSONDecodeError:
        m = re.search(r"\{.*\}", raw, flags=re.DOTALL)
        if m:
            try:
                return _extract(json.loads(m.group(0)))
            except json.JSONDecodeError:
                pass
        return raw.strip(), "", None, "", ""


def infer_lessons_batch(sessions: list[Session], model: str) -> None:
    if not sessions:
        return
    cache = _load_json_cache(LESSON_CACHE_PATH)
    total = len(sessions)
    print(f"\n🤖  Running AI ({model}) on {total} sessions...")
    hits = 0
    for i, s in enumerate(sessions, 1):
        h = s.row_hash()
        if h in cache:
            c = cache[h]
            s.ai_lesson = c.get("ai_lesson", "")
            s.comparison = c.get("comparison", "")
            s.ai_rating = c.get("ai_rating")
            s.ai_rating_reason = c.get("ai_rating_reason", "")
            s.suggested_prompt = c.get("suggested_prompt", "")
            hits += 1
            print(f"  [{i}/{total}] {s.staff} — {s.title[:50]} (cached)")
            continue
        prompt = LESSON_PROMPT_TEMPLATE.format(
            title=_truncate(s.title, 200), tool=s.tool, category=s.category,
            task_desc=_truncate(s.task_desc), prompt=_truncate(s.prompt),
            result=_truncate(s.result),
            user_lesson=_truncate(s.user_lesson, 600) or "(empty)",
        )
        try:
            raw = _call_openai(model, prompt)
            ai_lesson, comparison, ai_rating, reason, suggested = _parse_lesson_response(raw)
            s.ai_lesson = ai_lesson
            s.comparison = comparison or ("User left blank" if not s.user_lesson else "Disagree")
            s.ai_rating = ai_rating
            s.ai_rating_reason = reason
            s.suggested_prompt = suggested
            cache[h] = {
                "ai_lesson": s.ai_lesson, "comparison": s.comparison,
                "ai_rating": s.ai_rating, "ai_rating_reason": s.ai_rating_reason,
                "suggested_prompt": s.suggested_prompt,
            }
            rating_str = f"{ai_rating:.0f}" if ai_rating else "—"
            print(f"  [{i}/{total}] {s.staff} — {s.title[:50]}  →  {s.comparison} ({rating_str})")
        except requests.RequestException as e:
            s.ai_lesson = f"[AI Error: {e}]"
            s.comparison = "—"
            print(f"  [{i}/{total}] ⚠  {e}", file=sys.stderr)
    _save_json_cache(LESSON_CACHE_PATH, cache)
    print(f"✔  Done. Cache hits: {hits}/{total}")


# =========================================================================== #
#  Section 6 — Aggregation Helpers
# =========================================================================== #

def _agg(items: list[Session]) -> dict[str, Any]:
    n = len(items)
    est = sum(s.est_hours or 0 for s in items)
    actual = sum(s.actual_hours or 0 for s in items)
    saved = sum(s.time_saved or 0 for s in items)
    eff = (saved / est * 100) if est else 0
    rated = [s.rating for s in items if s.rating is not None]
    avg_rating = sum(rated) / len(rated) if rated else 0
    excellent = sum(1 for s in items if s.rating == 5)
    avg_saved = saved / n if n else 0
    # AI estimates
    ai_est = sum(s.ai_est_hours or 0 for s in items)
    ai_actual = sum(s.ai_actual_hours or 0 for s in items)
    ai_saved = round(ai_est - ai_actual, 1) if ai_est else 0
    ai_eff = (ai_saved / ai_est * 100) if ai_est else 0
    has_ai = any(s.ai_est_hours is not None for s in items)
    return {
        "n": n, "est": round(est, 1), "actual": round(actual, 1),
        "saved": round(saved, 1), "eff": round(eff, 1),
        "avg_rating": round(avg_rating, 2), "excellent": excellent,
        "avg_saved": round(avg_saved, 1),
        "ai_est": round(ai_est, 1), "ai_actual": round(ai_actual, 1),
        "ai_saved": round(ai_saved, 1), "ai_eff": round(ai_eff, 1),
        "has_ai": has_ai,
    }


def _sum_float(values: list[float | None]) -> float:
    return round(sum(v or 0 for v in values), 2)


def aggregate_sessions(items: list[Session]) -> dict[str, float | int]:
    est = _sum_float([s.est_hours for s in items])
    actual = _sum_float([s.actual_hours for s in items])
    saved = _sum_float([s.saved_hours for s in items])
    efficiency = round(saved / est * 100, 2) if est else 0
    return {"sessions": len(items), "est": est, "actual": actual,
            "saved": saved, "efficiency": efficiency}


def group_aggregate(sessions: list[Session], attr: str) -> list[tuple[str, dict[str, float | int]]]:
    groups: dict[str, list[Session]] = defaultdict(list)
    for s in sessions:
        key = getattr(s, attr) or "(none)"
        groups[str(key)].append(s)
    rows = [(k, aggregate_sessions(v)) for k, v in groups.items()]
    rows.sort(key=lambda r: (-float(r[1]["saved"]), r[0]))
    return rows


def date_aggregate(sessions: list[Session]) -> list[tuple[str, dict[str, float | int]]]:
    rows = group_aggregate(sessions, "date")

    def sort_key(r):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(r[0], fmt), r[0]
            except ValueError:
                pass
        return datetime.max, r[0]

    return sorted(rows, key=sort_key)


# =========================================================================== #
#  Section 7 — Report Sheet Builders (Phase 1 output)
# =========================================================================== #

def _style_header(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[row].height = 28


def _style_data_range(ws, r1: int, r2: int, n_cols: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.font is None or cell.font.name != "Arial":
                cell.font = CELL_FONT
            cell.border = BORDER
            if cell.alignment is None or cell.alignment.horizontal is None:
                cell.alignment = LEFT


def _set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _title_block(ws, title: str, subtitle: str, n_cols: int) -> int:
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=n_cols)
    ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=n_cols)
    ws.row_dimensions[1].height = 22
    return 4


def _write_total_row(ws, row: int, n_cols: int, hr: int, sum_cols: list[int],
                     avg_cols: list[int] | None = None) -> None:
    ws.cell(row=row, column=1, value="TOTAL")
    for c in sum_cols:
        col_letter = get_column_letter(c)
        ws.cell(row=row, column=c, value=f"=SUM({col_letter}{hr+1}:{col_letter}{row-1})")
    for c in (avg_cols or []):
        col_letter = get_column_letter(c)
        ws.cell(row=row, column=c,
                value=f"=IFERROR(AVERAGE({col_letter}{hr+1}:{col_letter}{row-1}),0)")
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT


def build_dashboard_sheet(wb: Workbook, sessions: list[Session]) -> None:
    """Combined Overview + Breakdown (Staff/Tool/Category) + Rating + Daily Trend on one sheet.
    When AI estimates are available, the report leads with Assessed (AI) as the objective
    baseline and shows Self-Reported (User) alongside it with deviation columns."""
    ws = wb.create_sheet("📊 Dashboard")
    a = _agg(sessions)
    has_ai = a["has_ai"]
    n_cols = 16 if has_ai else 10

    def _section(ws, r, text):
        ws.cell(row=r, column=1, value=text).font = Font(
            name="Arial", bold=True, size=12, color="1F4E78")
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=n_cols)

    gap_over = PatternFill("solid", start_color="FFC7CE")
    gap_under = PatternFill("solid", start_color="C6EFCE")

    # ── KPI Overview ──
    n_staff = len({s.staff for s in sessions})
    ws.cell(row=1, column=1,
            value="📊  AI DEV JOURNAL  —  Objective Assessment Report").font = TITLE_FONT
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=n_cols)
    subtitle = (f"Generated {datetime.now().strftime('%d/%m/%Y %H:%M')}  •  "
                f"{n_staff} staff  •  {a['n']} sessions")
    if has_ai:
        subtitle += "  •  Assessed = AI objective estimate  •  Self-Reported = user input"
    ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=n_cols)
    ws.row_dimensions[1].height = 22

    calc_desc = (
        "EST = Estimated hours without AI  •  Actual = Hours spent with AI  •  "
        "Saved = EST − Actual  •  Savings % = Saved / EST × 100  •  "
        "Δ = Self-Reported − Assessed  •  "
        "Accuracy = Self-Rep Saved / Assessed Saved × 100"
    ) if has_ai else (
        "EST = Estimated hours without AI  •  Actual = Hours spent with AI  •  "
        "Saved = EST − Actual  •  Savings % = Saved / EST × 100"
    )
    ws.cell(row=3, column=1, value=calc_desc).font = Font(
        name="Arial", italic=True, size=9, color="808080")
    ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=n_cols)

    kpi_hr = 5
    if has_ai:
        kpi_headers = ["Metric", "Assessed (AI)", "Self-Reported", "Deviation",
                       "", "Top 5 AI Tools", "Sessions", "Hours Saved", "Savings %"]
        kpi_n = len(kpi_headers)
        for i, h in enumerate(kpi_headers, 1):
            ws.cell(row=kpi_hr, column=i, value=h)
        _style_header(ws, kpi_hr, kpi_n)

        d_est = round(a["est"] - a["ai_est"], 1)
        d_act = round(a["actual"] - a["ai_actual"], 1)
        d_sav = round(a["saved"] - a["ai_saved"], 1)
        d_eff = round(a["eff"] - a["ai_eff"], 1)
        kpis: list[tuple[str, Any, Any, Any]] = [
            ("Total Sessions",          a["n"],             a["n"],             ""),
            ("Staff Count",             n_staff,            n_staff,            ""),
            ("EST (No AI)",             f"{a['ai_est']}h",  f"{a['est']}h",    f"{d_est:+.1f}h"),
            ("Actual (With AI)",        f"{a['ai_actual']}h", f"{a['actual']}h", f"{d_act:+.1f}h"),
            ("Hours Saved",             f"{a['ai_saved']}h", f"{a['saved']}h",  f"{d_sav:+.1f}h"),
            ("Savings %",              f"{a['ai_eff']}%",   f"{a['eff']}%",    f"{d_eff:+.1f}%"),
            ("Avg Saved/Session",       f"{a['avg_saved']}h", "",              ""),
            ("Avg Quality Rating",      f"{a['avg_rating']} / 5", "",          ""),
            ("Excellent Sessions (5★)", a["excellent"],     "",                ""),
        ]
        for i, (k, assessed, self_rep, dev) in enumerate(kpis):
            ws.cell(row=kpi_hr + 1 + i, column=1, value=k).font = CELL_FONT
            ws.cell(row=kpi_hr + 1 + i, column=2, value=assessed).font = NUMBER_FONT
            ws.cell(row=kpi_hr + 1 + i, column=3, value=self_rep).font = CELL_FONT
            ws.cell(row=kpi_hr + 1 + i, column=4, value=dev).font = CELL_FONT
        tool_col_start = 6
    else:
        kpi_headers_simple = ["Metric", "Value", "", "Top 5 AI Tools", "Sessions", "Hours Saved", "Savings %"]
        kpi_n = len(kpi_headers_simple)
        for i, h in enumerate(kpi_headers_simple, 1):
            ws.cell(row=kpi_hr, column=i, value=h)
        _style_header(ws, kpi_hr, kpi_n)
        kpis_s = [
            ("Total Sessions", a["n"]), ("Staff Count", n_staff),
            ("Total EST (No AI)", f"{a['est']}h"), ("Total Actual (With AI)", f"{a['actual']}h"),
            ("Total Hours Saved", f"{a['saved']}h"), ("Time Savings %", f"{a['eff']}%"),
            ("Avg Saved/Session", f"{a['avg_saved']}h"),
            ("Avg Quality Rating", f"{a['avg_rating']} / 5"),
            ("Excellent Sessions (5★)", a["excellent"]),
        ]
        for i, (k, v) in enumerate(kpis_s):
            ws.cell(row=kpi_hr + 1 + i, column=1, value=k).font = CELL_FONT
            ws.cell(row=kpi_hr + 1 + i, column=2, value=v).font = NUMBER_FONT
        tool_col_start = 4

    tool_data: dict[str, dict] = {}
    for s in sessions:
        if s.tool:
            d = tool_data.setdefault(s.tool, {"n": 0, "saved": 0.0, "est": 0.0})
            d["n"] += 1
            d["saved"] += s.time_saved or 0
            d["est"] += s.est_hours or 0
    top_tools = sorted(tool_data.items(), key=lambda x: -x[1]["saved"])[:5]
    for i, (t, d) in enumerate(top_tools):
        eff = (d["saved"] / d["est"] * 100) if d["est"] else 0
        ws.cell(row=kpi_hr + 1 + i, column=tool_col_start, value=t).font = CELL_FONT
        ws.cell(row=kpi_hr + 1 + i, column=tool_col_start + 1, value=d["n"]).font = NUMBER_FONT
        ws.cell(row=kpi_hr + 1 + i, column=tool_col_start + 2, value=round(d["saved"], 1)).font = NUMBER_FONT
        ws.cell(row=kpi_hr + 1 + i, column=tool_col_start + 3, value=f"{eff:.0f}%").font = NUMBER_FONT

    last_kpi = kpi_hr + 9
    _style_data_range(ws, kpi_hr + 1, last_kpi, kpi_n)
    cursor = last_kpi + 2

    # ── Self-Report Accuracy per Staff ──
    if has_ai:
        _section(ws, cursor, "SELF-REPORT ACCURACY  —  Deviation from AI Objective Estimate (Δ = Self-Reported − Assessed)")
        cursor += 1
        cmp_headers = [
            "Staff",
            "Assessed EST", "Self-Reported EST", "Δ EST",
            "Assessed Actual", "Self-Reported Actual", "Δ Actual",
            "Assessed Saved", "Self-Reported Saved", "Δ Saved",
            "Assessed %", "Self-Reported %", "Δ %",
            "Accuracy",
        ]
        for i, h in enumerate(cmp_headers, 1):
            ws.cell(row=cursor, column=i, value=h)
        _style_header(ws, cursor, len(cmp_headers))

        by_staff_cmp: dict[str, list[Session]] = {}
        for s in sessions:
            by_staff_cmp.setdefault(s.staff, []).append(s)
        cmp_rows = []
        for staff in sorted(by_staff_cmp, key=lambda k: -sum(s.time_saved or 0 for s in by_staff_cmp[k])):
            sa = _agg(by_staff_cmp[staff])
            de = round(sa["est"] - sa["ai_est"], 1)
            da = round(sa["actual"] - sa["ai_actual"], 1)
            ds = round(sa["saved"] - sa["ai_saved"], 1)
            dp = round(sa["eff"] - sa["ai_eff"], 1)
            # Accuracy: how close self-reported saved is to assessed saved (%)
            acc = f"{sa['saved'] / sa['ai_saved'] * 100:.0f}%" if sa["ai_saved"] else "—"
            cmp_rows.append((
                staff,
                sa["ai_est"], sa["est"], de,
                sa["ai_actual"], sa["actual"], da,
                sa["ai_saved"], sa["saved"], ds,
                sa["ai_eff"], sa["eff"], dp,
                acc,
            ))
        for i, row in enumerate(cmp_rows):
            for j, v in enumerate(row, 1):
                cell = ws.cell(row=cursor + 1 + i, column=j, value=v)
                if j in (4, 7, 10, 13) and isinstance(v, (int, float)):
                    cell.fill = gap_over if v > 0.5 else gap_under if v < -0.5 else PatternFill()
        _style_data_range(ws, cursor + 1, cursor + len(cmp_rows), len(cmp_headers))
        cursor = cursor + len(cmp_rows) + 2

    # ── Breakdown helper ──
    # When AI available: Assessed first, Self-Reported second, then deltas
    if has_ai:
        bd_headers_base = ["Sessions",
                           "Assessed EST", "Assessed Actual", "Assessed Saved", "Assessed %",
                           "Self-Rep EST", "Self-Rep Actual", "Self-Rep Saved", "Self-Rep %",
                           "Δ Saved", "Avg Rating", "5★"]
    else:
        bd_headers_base = ["Sessions", "EST (h)", "Actual (h)", "Saved (h)",
                           "Savings %", "Avg Saved/Session", "Avg Rating", "5★ Sessions"]

    def _write_breakdown(label_col_name: str, grouped: dict[str, list[Session]],
                         extra_cols: list[str] | None = None) -> int:
        nonlocal cursor
        headers = [label_col_name] + bd_headers_base
        if extra_cols:
            headers += extra_cols
        ncols = len(headers)
        for i, h in enumerate(headers, 1):
            ws.cell(row=cursor, column=i, value=h)
        _style_header(ws, cursor, ncols)

        rows_data = []
        for key in sorted(grouped, key=lambda k: -sum(s.time_saved or 0 for s in grouped[k])):
            items = grouped[key]
            ga = _agg(items)
            if has_ai:
                ds = round(ga["saved"] - ga["ai_saved"], 1)
                row: list[Any] = [key, ga["n"],
                                  ga["ai_est"], ga["ai_actual"], ga["ai_saved"], ga["ai_eff"],
                                  ga["est"], ga["actual"], ga["saved"], ga["eff"],
                                  ds, ga["avg_rating"], ga["excellent"]]
            else:
                row = [key, ga["n"], ga["est"], ga["actual"], ga["saved"],
                       ga["eff"], ga["avg_saved"], ga["avg_rating"], ga["excellent"]]
            if extra_cols:
                tools: dict[str, int] = {}
                for s in items:
                    if s.tool:
                        tools[s.tool] = tools.get(s.tool, 0) + 1
                row.append(max(tools.items(), key=lambda x: x[1])[0] if tools else "—")
            rows_data.append(row)

        for i, row in enumerate(rows_data):
            for j, v in enumerate(row, 1):
                cell = ws.cell(row=cursor + 1 + i, column=j, value=v)
                # Color Δ Saved column
                if has_ai and j == 11 and isinstance(v, (int, float)):
                    cell.fill = gap_over if v > 0.5 else gap_under if v < -0.5 else PatternFill()

        total_r = cursor + 1 + len(rows_data)
        if has_ai:
            _write_total_row(ws, total_r, ncols, cursor,
                             sum_cols=[2, 3, 4, 5, 7, 8, 9, 11, 13],
                             avg_cols=[6, 10, 12])
        else:
            _write_total_row(ws, total_r, ncols, cursor,
                             sum_cols=[2, 3, 4, 5, 9], avg_cols=[6, 7, 8])
        _style_data_range(ws, cursor + 1, total_r, ncols)
        cursor = total_r + 2
        return cursor

    # ── By Staff ──
    _section(ws, cursor, "BY STAFF")
    cursor += 1
    by_staff: dict[str, list[Session]] = {}
    for s in sessions:
        by_staff.setdefault(s.staff, []).append(s)
    _write_breakdown("Staff", by_staff, extra_cols=["Primary Tool"])

    # ── By Tool ──
    _section(ws, cursor, "BY AI TOOL")
    cursor += 1
    by_tool: dict[str, list[Session]] = {}
    for s in sessions:
        by_tool.setdefault(s.tool or "(unknown)", []).append(s)
    _write_breakdown("AI Tool", by_tool)

    # ── By Category ──
    _section(ws, cursor, "BY CATEGORY")
    cursor += 1
    by_cat: dict[str, list[Session]] = {}
    for s in sessions:
        by_cat.setdefault(s.category or "(uncategorized)", []).append(s)
    _write_breakdown("Category", by_cat)

    # ── Rating Distribution ──
    _section(ws, cursor, "RATING DISTRIBUTION BY TOOL")
    cursor += 1
    rating_headers = ["Tool", "1★", "2★", "3★", "4★", "5★", "Total Rated", "Avg Rating"]
    for i, h in enumerate(rating_headers, 1):
        ws.cell(row=cursor, column=i, value=h)
    _style_header(ws, cursor, len(rating_headers))
    rating_rows = rating_counts_by_tool(sessions)
    for i, row in enumerate(rating_rows):
        for j, v in enumerate(row, 1):
            ws.cell(row=cursor + 1 + i, column=j, value=v)
    _style_data_range(ws, cursor + 1, cursor + len(rating_rows), len(rating_headers))
    cursor = cursor + len(rating_rows) + 2

    # ── Daily Trend ──
    _section(ws, cursor, "DAILY TREND")
    cursor += 1
    trend_headers = ["Date", "Sessions", "EST (h)", "Actual (h)", "Saved (h)", "Savings %", "Avg Rating"]
    n_trend = len(trend_headers)
    for i, h in enumerate(trend_headers, 1):
        ws.cell(row=cursor, column=i, value=h)
    _style_header(ws, cursor, n_trend)

    by_day: dict[str, list[Session]] = {}
    for s in sessions:
        key = _fmt_date(s.date) or "(no date)"
        by_day.setdefault(key, []).append(s)

    def _sort_key(k: str):
        try:
            return datetime.strptime(k, "%d/%m/%Y")
        except ValueError:
            return datetime.max

    sorted_days = sorted(by_day.keys(), key=_sort_key)
    for i, day in enumerate(sorted_days):
        da = _agg(by_day[day])
        ws.cell(row=cursor + 1 + i, column=1, value=day)
        ws.cell(row=cursor + 1 + i, column=2, value=da["n"])
        ws.cell(row=cursor + 1 + i, column=3, value=da["est"])
        ws.cell(row=cursor + 1 + i, column=4, value=da["actual"])
        ws.cell(row=cursor + 1 + i, column=5, value=da["saved"])
        ws.cell(row=cursor + 1 + i, column=6, value=da["eff"])
        ws.cell(row=cursor + 1 + i, column=7, value=da["avg_rating"])
    _style_data_range(ws, cursor + 1, cursor + len(sorted_days), n_trend)

    _set_widths(ws, [28, 12, 12, 12, 12, 12, 12, 12, 12, 10, 10, 10, 10, 10, 24, 24])


def build_raw_log_sheet(wb: Workbook, sessions: list[Session]) -> None:
    ws = wb.create_sheet("📝 Raw Log")
    has_ai_est = any(s.ai_est_hours is not None for s in sessions)
    headers = ["Staff", "Date", "Session Name", "Tool", "Category",
               "Description", "Rating"]
    widths = [14, 12, 32, 16, 16, 40, 8]
    if has_ai_est:
        headers += ["Assessed EST", "Assessed Actual", "Assessed Saved", "Assessed %",
                    "Self-Rep EST", "Self-Rep Actual", "Self-Rep Saved", "Self-Rep %",
                    "Δ EST", "Δ Actual", "Δ Saved", "AI Reason"]
        widths += [12, 12, 12, 10, 12, 12, 12, 10, 10, 10, 10, 40]
    else:
        headers += ["EST (h)", "Actual (h)", "Saved (h)", "Savings %"]
        widths += [12, 12, 12, 10]
    headers += ["User Lesson", "Tags", "Source File"]
    widths += [40, 24, 22]
    n_cols = len(headers)
    raw_subtitle = f"{len(sessions)} rows"
    if has_ai_est:
        raw_subtitle += ("  •  Assessed = AI blind estimate based on staff profile  •  "
                         "Self-Rep = user input  •  Δ = Self-Rep − Assessed  •  "
                         "Saved = EST − Actual  •  % = Saved / EST × 100")
    hr = _title_block(ws, "📝  RAW LOG  —  Consolidated Sessions",
                      raw_subtitle, n_cols)
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, n_cols)

    gap_over = PatternFill("solid", start_color="FFC7CE")   # red — user over-reports
    gap_under = PatternFill("solid", start_color="C6EFCE")  # green — user under-reports

    for i, s in enumerate(sessions):
        r = hr + 1 + i
        c = 1
        for v in [s.staff, _fmt_date(s.date), s.title, s.tool, s.category,
                  s.task_desc, s.rating]:
            ws.cell(row=r, column=c, value=v)
            c += 1
        if has_ai_est:
            ai_saved = None
            ai_eff = "—"
            if s.ai_est_hours is not None and s.ai_actual_hours is not None:
                ai_saved = round(s.ai_est_hours - s.ai_actual_hours, 1)
                ai_eff = f"{ai_saved / s.ai_est_hours * 100:.0f}%" if s.ai_est_hours else "—"
            user_eff = f"{s.efficiency * 100:.0f}%" if s.efficiency is not None else "—"
            # Assessed (AI) first
            for v in [s.ai_est_hours, s.ai_actual_hours, ai_saved, ai_eff]:
                ws.cell(row=r, column=c, value=v)
                c += 1
            # Self-Reported (User) second
            for v in [s.est_hours, s.actual_hours, s.time_saved, user_eff]:
                ws.cell(row=r, column=c, value=v)
                c += 1
            # Deltas (Self-Reported − Assessed)
            d_est = round(s.est_hours - s.ai_est_hours, 1) if s.est_hours is not None and s.ai_est_hours is not None else None
            d_act = round(s.actual_hours - s.ai_actual_hours, 1) if s.actual_hours is not None and s.ai_actual_hours is not None else None
            d_sav = round(s.time_saved - ai_saved, 1) if s.time_saved is not None and ai_saved is not None else None
            for delta in [d_est, d_act, d_sav]:
                cell = ws.cell(row=r, column=c, value=delta)
                if isinstance(delta, (int, float)):
                    cell.fill = gap_over if delta > 0.5 else gap_under if delta < -0.5 else PatternFill()
                c += 1
            ws.cell(row=r, column=c, value=s.ai_est_reason)
            c += 1
        else:
            user_eff = f"{s.efficiency * 100:.0f}%" if s.efficiency is not None else "—"
            for v in [s.est_hours, s.actual_hours, s.time_saved, user_eff]:
                ws.cell(row=r, column=c, value=v)
                c += 1
        for v in [s.user_lesson, s.tags, s.source_file]:
            ws.cell(row=r, column=c, value=v)
            c += 1

    _style_data_range(ws, hr + 1, hr + len(sessions), n_cols)
    _set_widths(ws, widths)
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)


def build_ai_comparison_sheet(wb: Workbook, sessions: list[Session]) -> None:
    ws = wb.create_sheet("🤖 AI Lesson Compare")
    has_ai_est = any(s.ai_est_hours is not None for s in sessions)
    headers = ["Staff", "Date", "Session Name", "Tool",
               "Task Description", "Main Prompt", "Result",
               "User Lesson", "AI Inferred Lesson", "Comparison",
               "User ★", "AI ★", "Δ Rating"]
    if has_ai_est:
        headers += ["Assessed EST", "Self-Rep EST", "Δ EST",
                    "Assessed Actual", "Self-Rep Actual", "Δ Actual",
                    "Assessed Saved", "Self-Rep Saved", "Δ Saved"]
    headers += ["AI Rating Reason", "Suggested Prompt"]
    n_cols = len(headers)
    cmp_subtitle = ("AI infers lesson, rates output (1–5), and suggests an improved prompt  •  "
                    "Δ Rating = AI ★ − User ★")
    if has_ai_est:
        cmp_subtitle += ("  •  Assessed = AI blind estimate  •  Self-Rep = user input  •  "
                         "Δ = Self-Rep − Assessed  •  Saved = EST − Actual")
    hr = _title_block(ws,
                      "🤖  LESSON COMPARISON & PROMPT SUGGESTIONS  —  AI vs User",
                      cmp_subtitle,
                      n_cols)
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, n_cols)

    comparison_fills = {
        "Agree": PatternFill("solid", start_color="C6EFCE"),
        "Supplement": PatternFill("solid", start_color="FFEB9C"),
        "Disagree": PatternFill("solid", start_color="FFC7CE"),
        "User left blank": PatternFill("solid", start_color="D9D9D9"),
    }
    gap_green = PatternFill("solid", start_color="C6EFCE")
    gap_red = PatternFill("solid", start_color="FFC7CE")
    suggested_fill = PatternFill("solid", start_color="FFF2CC")

    for i, s in enumerate(sessions):
        r = hr + 1 + i
        c = 1
        for v in [s.staff, _fmt_date(s.date), s.title, s.tool,
                  s.task_desc or "—", s.prompt or "—", s.result or "—",
                  s.user_lesson or "(empty)", s.ai_lesson or "—"]:
            ws.cell(row=r, column=c, value=v)
            c += 1

        comp_cell = ws.cell(row=r, column=c, value=s.comparison or "—")
        if s.comparison in comparison_fills:
            comp_cell.fill = comparison_fills[s.comparison]
            comp_cell.alignment = CENTER
        c += 1

        ws.cell(row=r, column=c, value=s.rating).alignment = CENTER
        c += 1
        ws.cell(row=r, column=c, value=s.ai_rating).alignment = CENTER
        c += 1

        gap = None
        if s.rating is not None and s.ai_rating is not None:
            gap = round(s.ai_rating - s.rating, 1)
        gap_cell = ws.cell(row=r, column=c, value=gap)
        gap_cell.alignment = CENTER
        if gap is not None:
            gap_cell.fill = gap_green if gap >= 1 else gap_red if gap <= -1 else PatternFill()
        c += 1

        if has_ai_est:
            ai_saved = None
            if s.ai_est_hours is not None and s.ai_actual_hours is not None:
                ai_saved = round(s.ai_est_hours - s.ai_actual_hours, 1)
            # Assessed first, Self-Reported second, delta = Self-Rep − Assessed
            for ai, u in [(s.ai_est_hours, s.est_hours),
                          (s.ai_actual_hours, s.actual_hours),
                          (ai_saved, s.time_saved)]:
                ws.cell(row=r, column=c, value=ai)
                c += 1
                ws.cell(row=r, column=c, value=u)
                c += 1
                delta = round(u - ai, 1) if u is not None and ai is not None else None
                d_cell = ws.cell(row=r, column=c, value=delta)
                if isinstance(delta, (int, float)):
                    d_cell.fill = gap_red if delta > 0.5 else gap_green if delta < -0.5 else PatternFill()
                c += 1

        ws.cell(row=r, column=c, value=s.ai_rating_reason or "—")
        c += 1
        sug_cell = ws.cell(row=r, column=c, value=s.suggested_prompt or "—")
        if s.suggested_prompt:
            sug_cell.fill = suggested_fill

    _style_data_range(ws, hr + 1, hr + len(sessions), n_cols)
    widths = [14, 12, 26, 14, 36, 36, 36, 34, 38, 14, 8, 8, 10]
    if has_ai_est:
        widths += [10, 10, 10, 10, 10, 10, 10, 10, 10]
    widths += [32, 65]
    _set_widths(ws, widths)
    ws.freeze_panes = ws.cell(row=hr + 1, column=5)
    for r in range(hr + 1, hr + 1 + len(sessions)):
        ws.row_dimensions[r].height = 150


def build_report(sessions: list[Session], with_ai: bool) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    build_dashboard_sheet(wb, sessions)
    build_raw_log_sheet(wb, sessions)
    if with_ai:
        build_ai_comparison_sheet(wb, sessions)
    return wb


# =========================================================================== #
#  Section 8 — Error + SDLC Classification (Phase 2 AI)
# =========================================================================== #

def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, indent=2)


def build_classifier_prompt(sessions: list[Session], error_taxonomy: dict[str, str]) -> str:
    session_payload = [
        {
            "id": s.row_id, "staff": s.staff, "date": _fmt_date(s.date),
            "title": _truncate(s.title, 220), "tool": s.tool,
            "category": s.category, "task_description": _truncate(s.task_desc),
            "user_prompt": _truncate(s.prompt), "ai_result": _truncate(s.result),
            "user_lesson": _truncate(s.user_lesson, 700) or "(empty)",
            "ai_inferred_lesson": _truncate(s.ai_lesson, 900) or "(empty)",
        }
        for s in sessions
    ]

    return f"""<role>
You are an expert in prompt engineering and software development process analysis.
You classify developer AI journal entries to create management charts: common prompt errors and SDLC work types.
</role>

<context>
Input consists of work sessions from an AI Dev Journal. Each session has a user-written lesson and an AI-inferred lesson.
The goal is to accurately classify the errors each user encountered, prioritizing the fixed taxonomy, and only creating new labels when no existing label fits.
</context>

<fixed_error_taxonomy>
{_json_dumps(error_taxonomy)}
</fixed_error_taxonomy>

<fixed_sdlc_taxonomy>
{_json_dumps(SDLC_TAXONOMY)}
</fixed_sdlc_taxonomy>

<journal_sessions>
{_json_dumps(session_payload)}
</journal_sessions>

<examples>
  <example>
    <input>
      {{"id":"S_EXAMPLE","task_description":"Asked AI to fix Excel table format but output had wrong columns","user_lesson":"Need to specify expected format clearly","ai_inferred_lesson":"Prompt lacked output schema and column constraints"}}
    </input>
    <output>
      {{"id":"S_EXAMPLE","error_labels":["Clear and Format"],"new_error_labels":[],"error_evidence":"Both user_lesson and ai_inferred_lesson indicate missing format/schema output.","sdlc_category":"Development / Implementation","sdlc_confidence":0.82,"sdlc_reason":"Task involves AI fixing format in an existing file/script."}}
    </output>
  </example>
</examples>

<instructions>
Follow these steps exactly:
1. For each session in <journal_sessions>, read task_description, user_prompt, ai_result, user_lesson, and ai_inferred_lesson.
2. Assign 1 to {MAX_LABELS_PER_SESSION} `error_labels` for prompt errors the user encountered:
   - Prioritize labels from <fixed_error_taxonomy>.
   - Only add new labels if no fixed label accurately describes the main error.
   - If lesson data is insufficient, use "Insufficient Lesson Data".
3. Assign exactly one `sdlc_category` from <fixed_sdlc_taxonomy> for the session's task.
4. Write `error_evidence` and `sdlc_reason` concisely in English, based on evidence in the input.
5. Return results for ALL sessions, preserving the `id`.
</instructions>

<output_format>
Return ONLY a single valid JSON object, no markdown:
{{
  "sessions": [
    {{
      "id": "S1",
      "error_labels": ["Clear and Format"],
      "new_error_labels": [],
      "error_evidence": "Brief reason in English",
      "sdlc_category": "Development / Implementation",
      "sdlc_confidence": 0.0,
      "sdlc_reason": "Brief reason in English"
    }}
  ],
  "new_taxonomy": [
    {{"label": "New Label", "definition": "Short definition"}}
  ]
}}
</output_format>

Before responding, self-check:
- Every input session has exactly one output object.
- All `sdlc_category` values are from fixed_sdlc_taxonomy.
- `error_labels` has at most {MAX_LABELS_PER_SESSION} labels.
- JSON is valid with no explanation outside JSON.
"""






def _coerce_labels(raw: Any, taxonomy: dict[str, str]) -> list[str]:
    if isinstance(raw, str):
        values = [part.strip() for part in re.split(r"[,;|]", raw)]
    elif isinstance(raw, list):
        values = [str(item).strip() for item in raw]
    else:
        values = []
    labels: list[str] = []
    known_casefold = {l.casefold(): l for l in taxonomy}
    for v in values:
        if not v:
            continue
        label = known_casefold.get(v.casefold(), v)
        if label not in labels:
            labels.append(label)
        if len(labels) >= MAX_LABELS_PER_SESSION:
            break
    return labels or ["Insufficient Lesson Data"]


def classify_sessions(sessions: list[Session], model: str,
                      cache_path: Path = CHART_CACHE_PATH,
                      batch_size: int = 20, timeout: int = 300) -> None:
    cache = _load_json_cache(cache_path)
    taxonomy = dict(ERROR_TAXONOMY)

    misses: list[Session] = []
    hits = 0
    for s in sessions:
        key = f"{model}:{s.cache_hash}"
        cached = cache.get(key)
        if cached:
            s.error_labels = list(cached.get("error_labels") or ["Insufficient Lesson Data"])
            s.error_evidence = str(cached.get("error_evidence") or "")
            s.sdlc_category = str(cached.get("sdlc_category") or "Other")
            s.sdlc_confidence = _to_float(cached.get("sdlc_confidence"))
            s.sdlc_reason = str(cached.get("sdlc_reason") or "")
            hits += 1
            continue
        misses.append(s)

    print(f"🤖  Classifying errors + SDLC with {model}. Cache hits: {hits}/{len(sessions)}")

    for start in range(0, len(misses), batch_size):
        batch = misses[start:start + batch_size]
        print(f"  • Batch {start // batch_size + 1}: {len(batch)} session(s)")
        prompt = build_classifier_prompt(batch, taxonomy)
        raw = _call_openai(model=model, prompt=prompt, timeout=timeout)
        parsed = _parse_json_object(raw)

        for item in parsed.get("new_taxonomy", []) if isinstance(parsed.get("new_taxonomy"), list) else []:
            if not isinstance(item, dict):
                continue
            label = str(item.get("label") or "").strip()
            definition = str(item.get("definition") or "").strip()
            if label and label not in taxonomy:
                taxonomy[label] = definition or "Model-added prompt error label."

        by_id: dict[str, dict[str, Any]] = {}
        result_sessions = parsed.get("sessions")
        if isinstance(result_sessions, list):
            by_id = {str(item.get("id") or ""): item for item in result_sessions
                     if isinstance(item, dict) and item.get("id")}

        for s in batch:
            item = by_id.get(s.row_id, {})
            for new_label in item.get("new_error_labels", []) if isinstance(item.get("new_error_labels"), list) else []:
                new_label_text = str(new_label).strip()
                if new_label_text and new_label_text not in taxonomy:
                    taxonomy[new_label_text] = "Model-added prompt error label."

            s.error_labels = _coerce_labels(item.get("error_labels"), taxonomy)
            s.error_evidence = str(item.get("error_evidence") or "").strip()
            sdlc_cat = str(item.get("sdlc_category") or "Other").strip()
            if sdlc_cat not in SDLC_TAXONOMY:
                sdlc_cat = "Other"
            s.sdlc_category = sdlc_cat
            s.sdlc_confidence = _to_float(item.get("sdlc_confidence"))
            s.sdlc_reason = str(item.get("sdlc_reason") or "").strip()

            cache[f"{model}:{s.cache_hash}"] = {
                "error_labels": s.error_labels, "error_evidence": s.error_evidence,
                "sdlc_category": s.sdlc_category, "sdlc_confidence": s.sdlc_confidence,
                "sdlc_reason": s.sdlc_reason,
            }
        _save_json_cache(cache_path, cache)

    print(f"✔  Classification complete. Cache: {cache_path}")


def mark_unclassified(sessions: list[Session]) -> None:
    for s in sessions:
        s.error_labels = ["Unclassified"]
        s.error_evidence = "AI classification skipped (--no-ai)."
        s.sdlc_category = "Other"
        s.sdlc_confidence = None
        s.sdlc_reason = "AI classification skipped (--no-ai)."


# =========================================================================== #
#  Section 9 — Excel Chart Sheet Builders (Phase 2 output)
# =========================================================================== #

def _section_label(ws, row: int, text: str, n_cols: int = 6) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = SECTION_FILL
    cell.font = Font(name="Arial", bold=True, size=12, color="1F4E78")
    cell.alignment = LEFT
    return row + 1


def _xl_title(ws, row: int, title: str, subtitle: str, n_cols: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    ws.cell(row=row, column=1).alignment = LEFT
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=n_cols)
    ws.cell(row=row + 1, column=1, value=subtitle).font = SUBTITLE_FONT
    ws.cell(row=row + 1, column=1).alignment = LEFT
    return row + 3


def _xl_style_range(ws, row1: int, row2: int, col1: int, col2: int, header_row: int | None = None) -> None:
    for row in ws.iter_rows(min_row=row1, max_row=row2, min_col=col1, max_col=col2):
        for cell in row:
            cell.border = BORDER
            cell.font = CELL_FONT
            cell.alignment = CENTER if isinstance(cell.value, (int, float)) else LEFT
    if header_row:
        for cell in ws.iter_rows(min_row=header_row, max_row=header_row, min_col=col1, max_col=col2).__next__():
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER


def _write_xl_table(ws, start_row: int, start_col: int, headers: list[str],
                    rows: list[list[Any]]) -> tuple[int, int]:
    for col_offset, header in enumerate(headers):
        ws.cell(row=start_row, column=start_col + col_offset, value=header)
    for row_offset, row in enumerate(rows, 1):
        for col_offset, value in enumerate(row):
            ws.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
    end_row = start_row + len(rows)
    end_col = start_col + len(headers) - 1
    _xl_style_range(ws, start_row, end_row, start_col, end_col, header_row=start_row)
    return end_row, end_col


def _bar_xl_chart(ws, title: str, min_row: int, max_row: int, min_col: int,
                  max_col: int, category_col: int, anchor: str,
                  y_title: str = "Value", stacked: bool = False,
                  chart_type: str = "col", height: float = 9, width: float = 18) -> None:
    if max_row <= min_row:
        return
    chart = BarChart()
    chart.type = chart_type
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = ws.cell(row=min_row, column=category_col).value
    if stacked:
        chart.grouping = "stacked"
        chart.overlap = 100
    data = Reference(ws, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=category_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = height
    chart.width = width
    ws.add_chart(chart, anchor)


# --- Rating helpers ---

def rating_counts_by_tool(sessions: list[Session]) -> list[list[Any]]:
    tools = sorted({s.tool or "(unknown)" for s in sessions})
    rows: list[list[Any]] = []
    for tool in tools:
        items = [s for s in sessions if (s.tool or "(unknown)") == tool]
        counts = Counter()
        ratings: list[float] = []
        for s in items:
            if s.rating is None:
                continue
            rounded = int(round(s.rating))
            if 1 <= rounded <= 5:
                counts[rounded] += 1
                ratings.append(s.rating)
        total = sum(counts.values())
        avg = round(sum(ratings) / len(ratings), 2) if ratings else 0
        rows.append([tool, counts[1], counts[2], counts[3], counts[4], counts[5], total, avg])
    rows.sort(key=lambda r: (-r[6], r[0]))
    return rows


# --- Error helpers ---

def collapsed_error_labels(sessions: list[Session]) -> tuple[list[str], Counter[str]]:
    raw_counts: Counter[str] = Counter()
    for s in sessions:
        for label in (s.error_labels or ["Unclassified"]):
            raw_counts[label] += 1
    top = [l for l, _ in raw_counts.most_common(MAX_ERROR_LABELS_FOR_CHART)]
    collapsed: Counter[str] = Counter()
    for label, count in raw_counts.items():
        collapsed[label if label in top else "Other"] += count
    labels = [l for l in top if l in collapsed]
    if collapsed.get("Other"):
        labels.append("Other")
    return labels, collapsed


def error_staff_matrix(sessions: list[Session], labels: list[str]) -> tuple[list[str], list[list[Any]]]:
    staff_list = sorted({s.staff for s in sessions})
    rows: list[list[Any]] = []
    top_set = set(labels) - {"Other"}
    for label in labels:
        row: list[Any] = [label]
        for staff in staff_list:
            count = 0
            for s in sessions:
                if s.staff != staff:
                    continue
                session_labels = s.error_labels or ["Unclassified"]
                if label == "Other":
                    count += sum(1 for item in session_labels if item not in top_set)
                elif label in session_labels:
                    count += 1
            row.append(count)
        rows.append(row)
    return staff_list, rows


# --- SDLC helpers ---

def sdlc_staff_matrix(sessions: list[Session]) -> tuple[list[str], list[list[Any]]]:
    staff_list = sorted({s.staff for s in sessions})
    rows: list[list[Any]] = []
    for cat in SDLC_TAXONOMY:
        row: list[Any] = [cat]
        for staff in staff_list:
            row.append(sum(1 for s in sessions if s.staff == staff and s.sdlc_category == cat))
        rows.append(row)
    return staff_list, rows



def sdlc_task_detail_rows(sessions: list[Session], has_ai: bool) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for cat in SDLC_TAXONOMY:
        stage_sessions = [s for s in sessions
                          if (s.sdlc_category if s.sdlc_category in SDLC_TAXONOMY else "Other") == cat]
        stage_sessions.sort(key=lambda s: (s.title.casefold(), s.staff.casefold(), str(s.date)))
        for s in stage_sessions:
            saved = s.saved_hours
            base = [cat, s.title or s.task_desc or "(unnamed task)",
                    s.staff, _fmt_date(s.date)]
            if has_ai:
                ai_saved = round(s.ai_est_hours - s.ai_actual_hours, 1) if s.ai_est_hours and s.ai_actual_hours else 0
                ai_eff = round(ai_saved / s.ai_est_hours * 100, 2) if s.ai_est_hours else 0
                u_eff = round(saved / s.est_hours * 100, 2) if saved is not None and s.est_hours else 0
                d_sav = round((saved or 0) - ai_saved, 1)
                base += [s.ai_est_hours or 0, s.ai_actual_hours or 0, ai_saved, ai_eff,
                         s.est_hours or 0, s.actual_hours or 0, saved or 0, u_eff, d_sav]
            else:
                efficiency = round(saved / s.est_hours * 100, 2) if saved is not None and s.est_hours else 0
                base += [s.est_hours or 0, s.actual_hours or 0, saved or 0, efficiency]
            rows.append(base)
    return rows


# --- Sheet builders ---


def build_error_data_sheet(wb, sessions: list[Session]) -> None:
    ws = wb.create_sheet(ERROR_DATA_SHEET)
    row = _xl_title(ws, 1, "🏷️ Prompt Error Classification Data",
                    "Multi-label error classification from user lesson + AI inferred lesson.", 12)

    headers = ["Staff", "Date", "Session Name", "Tool", "Category", "Description",
               "User Lesson", "AI Inferred Lesson", "Error Labels", "Evidence", "Source File"]
    rows = [[s.staff, _fmt_date(s.date), s.title, s.tool, s.category, s.task_desc,
             s.user_lesson, s.ai_lesson,
             ", ".join(s.error_labels or ["Unclassified"]), s.error_evidence, s.source_file]
            for s in sessions]
    _write_xl_table(ws, row, 1, headers, rows)
    _set_widths(ws, [14, 12, 26, 16, 18, 38, 38, 42, 30, 42, 22])
    ws.freeze_panes = "A4"
    for row_idx in range(row + 1, row + 1 + len(rows)):
        ws.row_dimensions[row_idx].height = 70


def build_error_chart_sheet(wb, sessions: list[Session]) -> None:
    ws = wb.create_sheet(ERROR_CHART_SHEET)
    row = _xl_title(ws, 1, "🏷️ Prompt Error Charts",
                    f"Top {MAX_ERROR_LABELS_FOR_CHART} error labels from AI classification  •  "
                    "Count = number of sessions where the error was detected  •  "
                    "Labels per session: 1–3 (multi-label)", 10)

    labels, counts = collapsed_error_labels(sessions)
    row = _section_label(ws, row, "Top Error Labels", n_cols=4)
    top_start = row
    top_rows = [[l, counts[l], ERROR_TAXONOMY.get(l, "")] for l in labels]
    top_end, _ = _write_xl_table(ws, row, 1, ["Error Label", "Count", "Description"], top_rows)
    _bar_xl_chart(ws, "Top Prompt Errors", top_start, top_end, 2, 2, 1, "F4", y_title="Occurrences")
    row = top_end + 3

    staff_list, matrix_rows = error_staff_matrix(sessions, labels)
    row = _section_label(ws, row, "Error Labels by Staff", n_cols=max(2, len(staff_list) + 1))
    matrix_start = row
    matrix_end, _ = _write_xl_table(ws, row, 1, ["Error Label", *staff_list], matrix_rows)
    if staff_list:
        _bar_xl_chart(ws, "Prompt Errors by Staff", matrix_start, matrix_end,
                      2, len(staff_list) + 1, 1, "E22", y_title="Occurrences")
    _set_widths(ws, [28, 12, 60, *([12] * len(staff_list))])
    ws.freeze_panes = "A4"


def build_sdlc_sheet(wb, sessions: list[Session]) -> None:
    ws = wb.create_sheet(SDLC_SHEET, 0)
    has_ai = any(s.ai_est_hours is not None for s in sessions)
    n_summary_cols = 16 if has_ai else 8
    sdlc_subtitle = (
        "Assessed = AI blind estimate  |  Self-Rep = user input  |  "
        "Saved = EST − Actual  |  % = Saved / EST × 100  |  Δ Saved = Self-Rep Saved − Assessed Saved"
    ) if has_ai else (
        "Saved = EST − Actual  |  Efficiency % = Saved / EST × 100"
    )
    row = _xl_title(ws, 1, "🧭 SDLC Task Summary", sdlc_subtitle, n_summary_cols)

    summary_counts = Counter(s.sdlc_category for s in sessions)
    total_sessions = len(sessions)
    sessions_by_stage = {
        cat: [s for s in sessions
              if (s.sdlc_category if s.sdlc_category in SDLC_TAXONOMY else "Other") == cat]
        for cat in SDLC_TAXONOMY
    }

    row = _section_label(ws, row, "Tasks and Efficiency by SDLC Stage", n_cols=n_summary_cols)
    task_names_by_stage: dict[str, Counter[str]] = {cat: Counter() for cat in SDLC_TAXONOMY}
    for s in sessions:
        cat = s.sdlc_category if s.sdlc_category in SDLC_TAXONOMY else "Other"
        task_name = s.title or s.task_desc or "(unnamed task)"
        task_names_by_stage.setdefault(cat, Counter())[task_name] += 1

    def _format_task_names(cat: str) -> str:
        tc = task_names_by_stage.get(cat, Counter())
        if not tc:
            return "—"
        bullets = []
        for tn, count in tc.items():
            suffix = f" (×{count})" if count > 1 else ""
            bullets.append(f"• {tn}{suffix}")
        return "\n".join(bullets)

    summary_rows = []
    for cat in SDLC_TAXONOMY:
        items = sessions_by_stage[cat]
        ag = _agg(items)
        base = [cat, _format_task_names(cat), summary_counts[cat],
                round(summary_counts[cat] / total_sessions * 100, 1) if total_sessions else 0]
        if has_ai:
            base += [ag["ai_eff"], ag["eff"]]
        else:
            base += [ag["eff"]]
        summary_rows.append(base)

    summary_start = row
    if has_ai:
        sum_headers = ["SDLC Stage", "Task Names", "Task Count", "Share %",
                       "Assessed %", "Self-Reported %"]
    else:
        sum_headers = ["SDLC Stage", "Task Names", "Task Count", "Share %",
                       "Efficiency %"]
    summary_end, _ = _write_xl_table(ws, row, 1, sum_headers, summary_rows)
    _bar_xl_chart(ws, "Tasks by SDLC Stage", summary_start, summary_end, 3, 3, 1,
                  "F4" if not has_ai else "G4", y_title="Tasks")
    if has_ai:
        _bar_xl_chart(ws, "Assessed vs Self-Reported Efficiency %", summary_start, summary_end,
                      5, 6, 1, "O4", y_title="Efficiency %", chart_type="bar", height=10, width=18)
    else:
        _bar_xl_chart(ws, "Efficiency % by SDLC Stage", summary_start, summary_end,
                      5, 5, 1, "F4", y_title="Efficiency %", chart_type="bar", height=10, width=18)

    staff_list, matrix_rows = sdlc_staff_matrix(sessions)
    matrix_row = summary_end + 3
    matrix_row = _section_label(ws, matrix_row, "Task Count by SDLC Stage and Staff",
                                n_cols=max(2, len(staff_list) + 1))
    matrix_start = matrix_row
    matrix_end, _ = _write_xl_table(ws, matrix_row, 1, ["SDLC Stage", *staff_list], matrix_rows)
    if staff_list:
        _bar_xl_chart(ws, "Tasks by SDLC Stage and Staff", matrix_start, matrix_end,
                      2, len(staff_list) + 1, 1, "F22", y_title="Tasks", stacked=True)

    detail_rows = sdlc_task_detail_rows(sessions, has_ai)
    detail_row = matrix_end + 3
    if has_ai:
        detail_headers = ["SDLC Stage", "Task Name", "Staff", "Date",
                          "Assessed EST", "Assessed Actual", "Assessed Saved", "Assessed %",
                          "Self-Rep EST", "Self-Rep Actual", "Self-Rep Saved", "Self-Rep %",
                          "Δ Saved"]
    else:
        detail_headers = ["SDLC Stage", "Task Name", "Staff", "Date",
                          "EST (h)", "Actual (h)", "Saved (h)", "Efficiency %"]
    detail_row = _section_label(ws, detail_row, "All Task Details by SDLC Stage",
                                n_cols=len(detail_headers))
    detail_start = detail_row
    detail_end, _ = _write_xl_table(ws, detail_row, 1, detail_headers, detail_rows)

    _set_widths(ws, [30, 46, 12, 10, 12, 12, 12, 10, 12, 12, 12, 10, 10])
    ws.freeze_panes = "A6"
    for row_idx in range(summary_start + 1, summary_end + 1):
        task_count = ws.cell(row=row_idx, column=3).value or 0
        ws.row_dimensions[row_idx].height = min(180, max(42, 18 * int(task_count)))
    for row_idx in range(detail_start + 1, detail_end + 1):
        ws.row_dimensions[row_idx].height = 36


# =========================================================================== #
#  Section 10 — Workbook Polish (Phase 2)
# =========================================================================== #

def _last_used_row(ws) -> int:
    for row in range(ws.max_row, 0, -1):
        if any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, ws.max_column + 1)):
            return row
    return 1


def _last_used_col(ws) -> int:
    for col in range(ws.max_column, 0, -1):
        if any(ws.cell(row=row, column=col).value not in (None, "") for row in range(1, ws.max_row + 1)):
            return col
    return 1


def _row_values(ws, row: int, max_col: int) -> list[str]:
    return [str(ws.cell(row=row, column=col).value or "").strip() for col in range(1, max_col + 1)]


def _looks_like_header(values: list[str]) -> bool:
    nonempty = [v for v in values if v]
    if len(nonempty) < 2:
        return False
    normalized = {v.casefold() for v in nonempty}
    header_markers = {
        "staff", "date", "session name", "tool", "ai tool",
        "category", "group", "sessions", "sdlc stage", "task name",
        "error label", "metric", "value", "rating", "description",
    }
    return bool(normalized & header_markers)


def _iter_header_rows(ws) -> list[tuple[int, int]]:
    max_col = _last_used_col(ws)
    rows: list[tuple[int, int]] = []
    for row in range(1, min(_last_used_row(ws), 120) + 1):
        values = _row_values(ws, row, max_col)
        if _looks_like_header(values):
            header_width = max((idx for idx, v in enumerate(values, 1) if v), default=max_col)
            rows.append((row, header_width))
    return rows


def _contiguous_table_end(ws, header_row: int, max_col: int) -> int:
    row = header_row + 1
    while row <= ws.max_row:
        if not any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, max_col + 1)):
            return row - 1
        row += 1
    return ws.max_row


def _format_numeric_columns(ws, header_row: int, max_col: int, end_row: int) -> None:
    if end_row <= header_row:
        return
    for col in range(1, max_col + 1):
        header = str(ws.cell(row=header_row, column=col).value or "")
        hcf = header.casefold()
        number_format: str | None = None
        if "%" in header or "savings" in hcf or "eff" in hcf:
            number_format = "0.0"
        elif "(h)" in hcf or "saved" in hcf or "actual" in hcf or "est" in hcf:
            number_format = "0.0"
        elif "rating" in hcf or "★" in header:
            number_format = "0.0"
        elif "sessions" in hcf or "count" in hcf:
            number_format = "0"
        if not number_format:
            continue
        for r in range(header_row + 1, end_row + 1):
            cell = ws.cell(row=r, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = number_format


def _add_color_scale(ws, header_row: int, max_col: int, end_row: int) -> None:
    if end_row <= header_row + 1:
        return
    for col in range(1, max_col + 1):
        header = str(ws.cell(row=header_row, column=col).value or "").casefold()
        if not any(m in header for m in ["%", "savings", "eff", "rating", "saved"]):
            continue
        if not any(isinstance(ws.cell(row=r, column=col).value, (int, float))
                   for r in range(header_row + 1, end_row + 1)):
            continue
        col_letter = get_column_letter(col)
        ws.conditional_formatting.add(
            f"{col_letter}{header_row + 1}:{col_letter}{end_row}",
            ColorScaleRule(start_type="min", start_color="F8696B",
                           mid_type="percentile", mid_value=50, mid_color="FFEB84",
                           end_type="max", end_color="63BE7B"))


def _shade_body_rows(ws, header_row: int, max_col: int, end_row: int) -> None:
    if end_row <= header_row:
        return
    if ws.title in {AI_COMPARE_SHEET, ERROR_DATA_SHEET}:
        return
    for r in range(header_row + 1, end_row + 1):
        if str(ws.cell(row=r, column=1).value or "").upper().startswith(("TOTAL", "TỔNG")):
            continue
        if (r - header_row) % 2 == 0:
            for col in range(1, max_col + 1):
                cell = ws.cell(row=r, column=col)
                if cell.fill.fill_type is None:
                    cell.fill = PROFESSIONAL_ALT_FILL


def _set_professional_filter(ws) -> None:
    filter_sheet_names = {
        "📊 Dashboard",
        "📝 Raw Log", AI_COMPARE_SHEET, ERROR_DATA_SHEET,
    }
    if ws.title == SDLC_SHEET:
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "All Task Details by SDLC Stage":
                header_row = row + 1
                end_row = _contiguous_table_end(ws, header_row, 8)
                if end_row > header_row:
                    ws.auto_filter.ref = f"A{header_row}:H{end_row}"
                return
    if ws.title not in filter_sheet_names:
        return
    header_rows = _iter_header_rows(ws)
    if not header_rows:
        return
    header_row, max_col = header_rows[0]
    end_row = _contiguous_table_end(ws, header_row, max_col)
    if end_row > header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(max_col)}{end_row}"


def _polish_sheet(ws) -> None:
    max_row = _last_used_row(ws)
    max_col = _last_used_col(ws)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.freeze_panes = ws.freeze_panes or ("A5" if max_row >= 5 else None)

    if ws.cell(row=1, column=1).value:
        ws.cell(row=1, column=1).font = PROFESSIONAL_TITLE_FONT
        ws.row_dimensions[1].height = 24
    if ws.cell(row=2, column=1).value:
        ws.cell(row=2, column=1).font = PROFESSIONAL_SUBTITLE_FONT
        ws.row_dimensions[2].height = 20

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value in (None, ""):
                continue
            if not cell.font or not cell.font.bold:
                cell.font = PROFESSIONAL_BODY_FONT
            cell.alignment = CENTER if isinstance(cell.value, (int, float)) else LEFT

    for header_row, header_width in _iter_header_rows(ws):
        for cell in ws.iter_rows(min_row=header_row, max_row=header_row,
                                 min_col=1, max_col=header_width).__next__():
            cell.fill = PROFESSIONAL_HEADER_FILL
            cell.font = PROFESSIONAL_HEADER_FONT
            cell.alignment = CENTER
        end_row = _contiguous_table_end(ws, header_row, header_width)
        _format_numeric_columns(ws, header_row, header_width, end_row)
        _add_color_scale(ws, header_row, header_width, end_row)
        _shade_body_rows(ws, header_row, header_width, end_row)

    for r in range(1, max_row + 1):
        cell = ws.cell(row=r, column=1)
        if cell.value and len([v for v in _row_values(ws, r, max_col) if v]) == 1 and r not in {1, 2}:
            cell.fill = PROFESSIONAL_SECTION_FILL
            cell.font = Font(name="Aptos", bold=True, size=12, color=PROFESSIONAL_BLUE)

    _set_professional_filter(ws)
    ws.page_setup.orientation = "landscape" if max_col > 8 else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def polish_workbook(wb) -> None:
    tab_colors = {
        SDLC_SHEET: PROFESSIONAL_TEAL,
        ERROR_CHART_SHEET: "C00000",
        ERROR_DATA_SHEET: "C00000",
        "📊 Dashboard": PROFESSIONAL_BLUE,
        "📝 Raw Log": "7F7F7F", AI_COMPARE_SHEET: "FFC000",
    }
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = tab_colors.get(ws.title, PROFESSIONAL_NAVY)
        _polish_sheet(ws)


def add_chart_sheets(wb, sessions: list[Session]) -> None:
    """Add chart/SDLC sheets to the workbook and polish."""
    for name in GENERATED_SHEETS:
        if name in wb.sheetnames:
            del wb[name]

    build_sdlc_sheet(wb, sessions)
    build_error_data_sheet(wb, sessions)
    build_error_chart_sheet(wb, sessions)

    generated_front_order = [SDLC_SHEET, ERROR_CHART_SHEET]
    for target_index, sheet_name in enumerate(generated_front_order):
        ws = wb[sheet_name]
        current_index = wb._sheets.index(ws)
        wb._sheets.insert(target_index, wb._sheets.pop(current_index))

    error_data_ws = wb[ERROR_DATA_SHEET]
    wb._sheets.append(wb._sheets.pop(wb._sheets.index(error_data_ws)))
    polish_workbook(wb)


# =========================================================================== #
#  Section 11 — Matplotlib Chart Builders (Phase 3)
# =========================================================================== #

def _add_watermark(fig: plt.Figure) -> None:
    fig.text(0.98, 0.02, "AI Dev Journal Report", fontsize=7, color="#BFBFBF",
             ha="right", va="bottom", style="italic")


def _save_chart(fig: plt.Figure, pdf: PdfPages, out_dir: Path, name: str) -> None:
    _add_watermark(fig)
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="This figure includes Axes.*", category=UserWarning)
        fig.tight_layout(rect=[0, 0.03, 1, 0.97])
    pdf.savefig(fig, dpi=200)
    fig.savefig(out_dir / f"{name}.png", dpi=200, bbox_inches="tight",
                facecolor=WHITE, edgecolor="none")
    plt.close(fig)


def _first_present(mapping: dict[str, Any], *keys: str, default: Any = "—") -> Any:
    for key in keys:
        value = mapping.get(key)
        if value not in (None, ""):
            return value
    return default


def _clean_generated_pngs(out_dir: Path) -> None:
    for name in GENERATED_PNG_NAMES:
        path = out_dir / name
        if path.exists():
            path.unlink()


def _build_chart_data(sessions: list[Session]) -> dict[str, Any]:
    """Build data dict for chart functions from in-memory sessions."""
    data: dict[str, Any] = {}

    # Per Staff
    by_staff: dict[str, list[Session]] = {}
    for s in sessions:
        by_staff.setdefault(s.staff, []).append(s)
    per_staff = []
    for staff, items in by_staff.items():
        a = _agg(items)
        per_staff.append({
            "Staff": staff, "Sessions": a["n"], "EST (h)": a["est"],
            "Actual (h)": a["actual"], "Saved (h)": a["saved"],
            "Savings %": a["eff"], "Avg Rating": a["avg_rating"],
        })
    per_staff.sort(key=lambda r: -r["Saved (h)"])
    data["per_staff"] = per_staff

    # Per Tool
    by_tool: dict[str, list[Session]] = {}
    for s in sessions:
        by_tool.setdefault(s.tool or "(unknown)", []).append(s)
    per_tool = []
    for tool, items in by_tool.items():
        a = _agg(items)
        per_tool.append({
            "AI Tool": tool, "Sessions": a["n"], "EST (h)": a["est"],
            "Actual (h)": a["actual"], "Saved (h)": a["saved"],
            "Savings %": a["eff"],
        })
    per_tool.sort(key=lambda r: -r["Saved (h)"])
    data["per_tool"] = per_tool

    # Per Category
    by_cat: dict[str, list[Session]] = {}
    for s in sessions:
        by_cat.setdefault(s.category or "(uncategorized)", []).append(s)
    per_cat = []
    for cat, items in by_cat.items():
        a = _agg(items)
        per_cat.append({
            "Category": cat, "Sessions": a["n"], "EST (h)": a["est"],
            "Actual (h)": a["actual"], "Saved (h)": a["saved"],
            "Savings %": a["eff"],
        })
    per_cat.sort(key=lambda r: -r["Saved (h)"])
    data["per_category"] = per_cat

    # KPIs
    a = _agg(sessions)
    n_staff = len({s.staff for s in sessions})
    data["kpis"] = {
        "Total Sessions": a["n"], "Staff Count": n_staff,
        "Total EST (No AI)": f"{a['est']}h", "Total Actual (With AI)": f"{a['actual']}h",
        "Total Hours Saved": f"{a['saved']}h",
        "Time Savings %": f"{a['eff']}%",
    }

    # Raw log for rating distribution
    data["raw_log"] = [{"Tool": s.tool, "Rating": s.rating} for s in sessions]

    # AI estimates for comparison chart
    data["has_ai_est"] = any(s.ai_est_hours is not None for s in sessions)
    if data["has_ai_est"]:
        data["kpis"]["AI EST (No AI)"] = f"{a['ai_est']}h"
        data["kpis"]["AI Actual (With AI)"] = f"{a['ai_actual']}h"
        data["kpis"]["AI Hours Saved"] = f"{a['ai_saved']}h"
        data["kpis"]["AI Savings %"] = f"{a['ai_eff']}%"

    return data


def _bar_est_actual(ax, labels: list[str], est: list[float],
                    actual: list[float], saved: list[float]) -> None:
    y = np.arange(len(labels))
    h = 0.35
    ax.barh(y + h / 2, est, h, label="EST (No AI)", color=ORANGE, alpha=0.85)
    ax.barh(y - h / 2, actual, h, label="Actual (With AI)", color=BLUE, alpha=0.85)

    max_val = max(est) if est else 1
    for i in range(len(labels)):
        if saved[i] > 0:
            ax.text(est[i] + max_val * 0.02, y[i] + h / 2,
                    f"-{saved[i]:.1f}h", va="center", fontsize=8, color=GREEN, fontweight="bold")
        eff = (saved[i] / est[i] * 100) if est[i] else 0
        if eff > 0:
            ax.text(actual[i] + max_val * 0.02, y[i] - h / 2,
                    f"{eff:.0f}%", va="center", fontsize=8, color=BLUE, fontweight="bold")

    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.set_xlabel("Hours")
    ax.legend(loc="lower right")
    ax.xaxis.set_major_locator(mticker.MaxNLocator(integer=True))
    ax.grid(axis="x", alpha=0.3)
    ax.invert_yaxis()


def _chart_efficiency_view(rows, label_col, title, filename, pdf, out_dir) -> None:
    if not rows:
        return
    labels = [str(r.get(label_col, ""))[:25] for r in rows]
    est = [float(r.get("EST (h)", 0) or 0) for r in rows]
    actual = [float(r.get("Actual (h)", 0) or 0) for r in rows]
    saved = [float(r.get("Saved (h)", 0) or 0) for r in rows]

    fig_h = max(4.5, 1.2 * len(labels) + 2)
    fig, ax = plt.subplots(figsize=(12, fig_h))
    fig.suptitle(title, fontsize=16, fontweight="bold", color=NAVY, y=0.98)
    _bar_est_actual(ax, labels, est, actual, saved)

    total_saved = sum(saved)
    total_est = sum(est)
    eff = (total_saved / total_est * 100) if total_est else 0
    ax.set_title(f"Total saved: {total_saved:.1f}h / {total_est:.1f}h ({eff:.0f}%)",
                 fontsize=11, color=GRAY, pad=10)
    _save_chart(fig, pdf, out_dir, filename)


def chart_kpi_summary(data: dict, pdf: PdfPages, out_dir: Path) -> None:
    kpis = data.get("kpis", {})
    per_staff = data.get("per_staff", [])

    fig, ax = plt.subplots(figsize=(14, 7))
    ax.axis("off")
    fig.text(0.5, 0.94, "AI DEV JOURNAL — EXECUTIVE SUMMARY",
             fontsize=22, fontweight="bold", color=NAVY, ha="center")

    boxes = [
        ("Total Sessions", str(kpis.get("Total Sessions", "—")), BLUE),
        ("Staff Count", str(kpis.get("Staff Count", "—")), TEAL),
        ("EST (No AI)", str(kpis.get("Total EST (No AI)", "—")), ORANGE),
        ("Actual (With AI)", str(kpis.get("Total Actual (With AI)", "—")), GREEN),
        ("Hours Saved", str(kpis.get("Total Hours Saved", "—")), NAVY),
        ("AI Efficiency", str(kpis.get("Time Savings %", "—")), RED),
    ]
    for i, (label, value, color) in enumerate(boxes):
        col = i % 3
        row_idx = i // 3
        x = 0.10 + col * 0.30
        y = 0.62 - row_idx * 0.28
        rect = plt.Rectangle((x - 0.02, y - 0.05), 0.26, 0.22,
                              transform=fig.transFigure, facecolor=color,
                              alpha=0.10, edgecolor=color, linewidth=1.5,
                              clip_on=False, zorder=2)
        fig.patches.append(rect)
        fig.text(x + 0.11, y + 0.10, value, fontsize=26, fontweight="bold",
                 color=color, ha="center", va="center")
        fig.text(x + 0.11, y + 0.0, label, fontsize=10, color=GRAY,
                 ha="center", va="center")

    if per_staff:
        table_y = 0.18
        fig.text(0.5, table_y + 0.06, "Ranking by Time Saved",
                 fontsize=11, fontweight="bold", color=NAVY, ha="center")
        header = f"{'Staff':<16} {'Sess':>6} {'EST':>8} {'Actual':>8} {'Saved':>8} {'Eff%':>7}"
        fig.text(0.18, table_y, header, fontsize=9, fontfamily="monospace",
                 color=NAVY, fontweight="bold")
        for j, row in enumerate(per_staff[:6]):
            name = str(row.get("Staff", ""))[:14]
            eff = _first_present(row, "Savings %", default=0)
            line = (f"{name:<16} {row.get('Sessions', 0):>6} "
                    f"{row.get('EST (h)', 0):>7}h {row.get('Actual (h)', 0):>7}h "
                    f"{row.get('Saved (h)', 0):>7}h {eff:>6}%")
            fig.text(0.18, table_y - 0.035 * (j + 1), line, fontsize=8.5,
                     fontfamily="monospace", color=GRAY)
    _save_chart(fig, pdf, out_dir, "03_kpi_summary")


def chart_staff_effectiveness(data: dict, pdf: PdfPages, out_dir: Path) -> None:
    _chart_efficiency_view(data.get("per_staff", []), "Staff",
                           "Staff AI Effectiveness — These Tasks",
                           "02_staff_ai_effectiveness", pdf, out_dir)


def chart_efficiency(data: dict, pdf: PdfPages, out_dir: Path) -> None:
    views = [
        ("per_tool", "AI Tool", "EST vs Actual — By AI Tool", "04_est_actual_tool"),
        ("per_category", "Category", "EST vs Actual — By Category", "05_est_actual_category"),
    ]
    for key, label_col, title, filename in views:
        _chart_efficiency_view(data.get(key, []), label_col, title, filename, pdf, out_dir)


def chart_rating(data: dict, pdf: PdfPages, out_dir: Path) -> None:
    raw = data.get("raw_log", [])
    if not raw:
        return

    tool_ratings: dict[str, Counter] = defaultdict(Counter)
    for row in raw:
        tool = str(row.get("Tool") or "(unknown)")
        rating = row.get("Rating")
        if rating is not None:
            try:
                r = int(float(rating))
                if 1 <= r <= 5:
                    tool_ratings[tool][r] += 1
            except (ValueError, TypeError):
                pass
    if not tool_ratings:
        return

    tools = sorted(tool_ratings.keys(), key=lambda t: -sum(tool_ratings[t].values()))
    stars = [1, 2, 3, 4, 5]
    star_keys = [f"{s} star" for s in stars]
    star_display = ["1 Star", "2 Star", "3 Star", "4 Star", "5 Star"]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, max(4.5, len(tools) * 1.2 + 2)),
                                   gridspec_kw={"width_ratios": [2.5, 1]})
    fig.suptitle("User Satisfaction — Rating Distribution by AI Tool",
                 fontsize=16, fontweight="bold", color=NAVY, y=0.98)

    y = np.arange(len(tools))
    left = np.zeros(len(tools))
    for s, key, display in zip(stars, star_keys, star_display):
        vals = [tool_ratings[t][s] for t in tools]
        ax1.barh(y, vals, left=left, label=display,
                 color=STAR_COLORS[key], alpha=0.85, height=0.6)
        for i, v in enumerate(vals):
            if v > 0:
                ax1.text(left[i] + v / 2, y[i], str(v), ha="center", va="center",
                         fontsize=8, fontweight="bold", color=WHITE)
        left += np.array(vals, dtype=float)

    ax1.set_yticks(y)
    ax1.set_yticklabels(tools)
    ax1.set_xlabel("Sessions")
    ax1.legend(loc="lower right", ncol=5)
    ax1.grid(axis="x", alpha=0.3)
    ax1.invert_yaxis()

    avg_ratings = []
    for t in tools:
        counts = tool_ratings[t]
        total = sum(counts.values())
        avg = sum(s * c for s, c in counts.items()) / total if total else 0
        avg_ratings.append(avg)

    colors = [GREEN if a >= 4.5 else BLUE if a >= 3.5 else ORANGE if a >= 2.5 else RED
              for a in avg_ratings]
    ax2.barh(y, avg_ratings, color=colors, alpha=0.85, height=0.5)
    for i, v in enumerate(avg_ratings):
        ax2.text(v + 0.05, y[i], f"{v:.1f}/5", va="center", fontsize=10,
                 fontweight="bold", color=colors[i])
    ax2.set_xlim(0, 5.5)
    ax2.set_xlabel("Average Rating")
    ax2.set_yticks(y)
    ax2.set_yticklabels([""] * len(tools))
    ax2.axvline(x=4, color=GREEN, linestyle="--", alpha=0.4, linewidth=1)
    ax2.grid(axis="x", alpha=0.3)
    ax2.invert_yaxis()
    ax2.set_title("Average", fontsize=11, color=GRAY)
    _save_chart(fig, pdf, out_dir, "06_rating_distribution")


def chart_errors_from_sessions(sessions: list[Session], pdf: PdfPages, out_dir: Path) -> None:
    """Top error labels bar + staff x error heatmap from classified sessions."""
    label_counter: Counter = Counter()
    staff_errors: dict[str, Counter] = defaultdict(Counter)
    for s in sessions:
        for label in (s.error_labels or []):
            label_counter[label] += 1
            staff_errors[s.staff][label] += 1

    if not label_counter:
        return

    # --- Top errors bar chart ---
    top_labels = label_counter.most_common(10)
    labels = [l for l, _ in top_labels]
    counts = [c for _, c in top_labels]

    fig, ax = plt.subplots(figsize=(12, max(5, len(labels) * 0.7 + 2)))
    fig.suptitle("Top Prompt Errors — Classified by Best Practices",
                 fontsize=16, fontweight="bold", color=NAVY, y=0.98)

    y = np.arange(len(labels))
    max_c = max(counts)
    colors_bar = [plt.cm.Blues(0.4 + 0.5 * c / max_c) for c in counts]
    ax.barh(y, counts, color=colors_bar, height=0.6, edgecolor=WHITE, linewidth=0.5)
    for i, (cnt, lbl) in enumerate(zip(counts, labels)):
        ax.text(cnt + max_c * 0.02, y[i], f"{cnt}", va="center",
                fontsize=10, fontweight="bold", color=NAVY)
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=10)
    ax.set_xlabel("Occurrences")
    ax.grid(axis="x", alpha=0.3)
    ax.invert_yaxis()
    n_sessions_with_errors = sum(1 for s in sessions if s.error_labels)
    ax.set_title(f"Total {sum(counts)} errors from {n_sessions_with_errors} sessions",
                 fontsize=11, color=GRAY, pad=10)
    _save_chart(fig, pdf, out_dir, "07_top_errors")

    # --- Staff × Error heatmap ---
    staff_list = sorted(staff_errors.keys())
    if not staff_list or not labels:
        return

    matrix = np.zeros((len(staff_list), len(labels)))
    for i, staff in enumerate(staff_list):
        for j, label in enumerate(labels):
            matrix[i, j] = staff_errors[staff].get(label, 0)

    fig, ax = plt.subplots(figsize=(max(10, len(labels) * 1.2 + 2),
                                    max(4, len(staff_list) * 0.8 + 2.5)))
    fig.suptitle("Prompt Error Distribution by Staff",
                 fontsize=16, fontweight="bold", color=NAVY, y=0.98)
    im = ax.imshow(matrix, cmap="Blues", aspect="auto", vmin=0)
    ax.set_xticks(np.arange(len(labels)))
    ax.set_xticklabels(labels, rotation=35, ha="right", fontsize=9)
    ax.set_yticks(np.arange(len(staff_list)))
    ax.set_yticklabels(staff_list, fontsize=11)

    for i in range(len(staff_list)):
        for j in range(len(labels)):
            val = int(matrix[i, j])
            if val > 0:
                text_color = WHITE if val >= matrix.max() * 0.6 else NAVY
                ax.text(j, i, str(val), ha="center", va="center",
                        fontsize=11, fontweight="bold", color=text_color)

    cbar = fig.colorbar(im, ax=ax, shrink=0.7, pad=0.02)
    cbar.set_label("Count", fontsize=9)
    ax.set_title("Each cell = number of times staff encountered the error",
                 fontsize=10, color=GRAY, pad=12)
    _save_chart(fig, pdf, out_dir, "08_error_heatmap")


def _wrap_task_bullets(task_counts: Counter[str], width: int = 72) -> str:
    lines: list[str] = []
    for task_name, count in task_counts.items():
        suffix = f" (×{count})" if count > 1 else ""
        bullet = f"• {task_name}{suffix}"
        wrapped = textwrap.wrap(bullet, width=width, subsequent_indent="  ")
        lines.extend(wrapped or [bullet])
    return "\n".join(lines) if lines else "—"


def chart_sdlc_tasks(sessions: list[Session], pdf: PdfPages, out_dir: Path) -> None:
    """SDLC stages: task count + Assessed vs Self-Reported efficiency comparison."""
    has_ai = any(s.ai_est_hours is not None for s in sessions)

    stage_tasks: dict[str, Counter[str]] = {stage: Counter() for stage in SDLC_TAXONOMY}
    stage_aggs: dict[str, dict] = {}
    for stage in SDLC_TAXONOMY:
        items = [s for s in sessions
                 if (s.sdlc_category if s.sdlc_category in SDLC_TAXONOMY else "Other") == stage]
        for s in items:
            stage_tasks[stage][s.title or s.task_desc or "(unnamed task)"] += 1
        stage_aggs[stage] = _agg(items)

    stages = [stage for stage in SDLC_TAXONOMY if stage_tasks[stage]]
    if not stages:
        return

    counts = [sum(stage_tasks[stage].values()) for stage in stages]
    assessed_eff = [stage_aggs[stage]["ai_eff"] for stage in stages]
    self_rep_eff = [stage_aggs[stage]["eff"] for stage in stages]
    max_count = max(counts) if counts else 1
    total_distinct = sum(len(stage_tasks[stage]) for stage in stages)
    fig_h = min(24, max(7, 2.5 + len(stages) * 0.8 + total_distinct * 0.22))

    if has_ai:
        fig, (ax_count, ax_eff, ax_text) = plt.subplots(
            1, 3, figsize=(22, fig_h),
            gridspec_kw={"width_ratios": [1.0, 1.0, 2.0], "wspace": 0.12})
    else:
        fig, (ax_count, ax_text) = plt.subplots(
            1, 2, figsize=(18, fig_h),
            gridspec_kw={"width_ratios": [1.2, 2.3], "wspace": 0.08})
        ax_eff = None

    fig.suptitle("SDLC Stage — Tasks + Efficiency" +
                 (" (Assessed vs Self-Reported)" if has_ai else ""),
                 fontsize=16, fontweight="bold", color=NAVY, y=0.98)

    y = np.arange(len(stages))
    colors = [PALETTE[i % len(PALETTE)] for i in range(len(stages))]

    # Left panel: task count
    bars = ax_count.barh(y, counts, color=colors, alpha=0.85, height=0.6)
    for bar, count in zip(bars, counts):
        ax_count.text(count + max_count * 0.03, bar.get_y() + bar.get_height() / 2,
                      str(count), va="center", fontsize=10, fontweight="bold", color=NAVY)
    ax_count.set_yticks(y)
    ax_count.set_yticklabels(stages, fontsize=10)
    ax_count.set_xlabel("Task count")
    ax_count.set_xlim(0, max_count * 1.35)
    ax_count.grid(axis="x", alpha=0.3)
    ax_count.invert_yaxis()
    ax_count.set_title("Tasks per stage", fontsize=11, color=GRAY, pad=10)

    # Middle panel: efficiency comparison (only when AI estimates exist)
    if ax_eff is not None:
        h = 0.25
        ax_eff.barh(y + h / 2, assessed_eff, h, label="Assessed (AI)",
                    color="#1A5276", alpha=0.85)
        ax_eff.barh(y - h / 2, self_rep_eff, h, label="Self-Reported",
                    color=ORANGE, alpha=0.75)
        max_eff = max(max(assessed_eff, default=1), max(self_rep_eff, default=1))
        for i in range(len(stages)):
            ax_eff.text(assessed_eff[i] + max_eff * 0.02, y[i] + h / 2,
                        f"{assessed_eff[i]:.0f}%", va="center", fontsize=8, color="#1A5276")
            ax_eff.text(self_rep_eff[i] + max_eff * 0.02, y[i] - h / 2,
                        f"{self_rep_eff[i]:.0f}%", va="center", fontsize=8, color=ORANGE)
        ax_eff.set_yticks(y)
        ax_eff.set_yticklabels([""] * len(stages))
        ax_eff.set_xlabel("Efficiency %")
        ax_eff.legend(loc="lower right", fontsize=8)
        ax_eff.grid(axis="x", alpha=0.3)
        ax_eff.invert_yaxis()
        ax_eff.set_title("Savings % = Saved / EST", fontsize=11, color=GRAY, pad=10)

    # Right panel: task list
    ax_text.set_xlim(0, 1)
    ax_text.set_ylim(ax_count.get_ylim())
    ax_text.axis("off")
    ax_text.set_title("Tasks within each stage", fontsize=11, color=GRAY, pad=10)
    for idx, stage in enumerate(stages):
        if has_ai:
            header = f"Assessed: {assessed_eff[idx]:.0f}%  |  Self-Rep: {self_rep_eff[idx]:.0f}%"
        else:
            header = f"Efficiency: {self_rep_eff[idx]:.1f}%"
        ax_text.text(0.0, idx,
                     f"{header}\n{_wrap_task_bullets(stage_tasks[stage])}",
                     va="center", ha="left", fontsize=8.2, color=GRAY, linespacing=1.25)

    footer = "Efficiency % = Saved / EST × 100"
    if has_ai:
        footer += "  |  Assessed = AI blind estimate  |  Self-Reported = user input"
    fig.text(0.5, 0.04, footer, ha="center", fontsize=9, color=GRAY)
    _save_chart(fig, pdf, out_dir, "01_sdlc_tasks_by_stage")


def chart_user_vs_ai(sessions: list[Session], pdf: PdfPages, out_dir: Path) -> None:
    """Grouped bar chart: User EST/Actual/Saved vs AI EST/Actual/Saved per staff."""
    if not any(s.ai_est_hours is not None for s in sessions):
        return

    by_staff: dict[str, list[Session]] = {}
    for s in sessions:
        by_staff.setdefault(s.staff, []).append(s)

    staff_names = sorted(by_staff, key=lambda k: -sum(s.time_saved or 0 for s in by_staff[k]))
    aggs = {staff: _agg(by_staff[staff]) for staff in staff_names}

    y = np.arange(len(staff_names))
    h = 0.18

    fig, axes = plt.subplots(1, 3, figsize=(18, max(4, len(staff_names) * 1.2 + 2)),
                              sharey=True)
    fig.suptitle("Assessed (AI) vs Self-Reported (User) — Per Staff",
                 fontsize=16, fontweight="bold", color=NAVY, y=0.98)

    metrics = [
        ("EST (No AI)", "ai_est", "est", "#D35400", ORANGE),
        ("Actual (With AI)", "ai_actual", "actual", "#1A5276", BLUE),
        ("Hours Saved", "ai_saved", "saved", "#196F3D", GREEN),
    ]

    for ax, (title, a_key, u_key, a_color, u_color) in zip(axes, metrics):
        a_vals = [aggs[s][a_key] for s in staff_names]
        u_vals = [aggs[s][u_key] for s in staff_names]

        ax.barh(y + h / 2, a_vals, h, label="Assessed (AI)", color=a_color, alpha=0.85)
        ax.barh(y - h / 2, u_vals, h, label="Self-Reported", color=u_color, alpha=0.65,
                edgecolor=a_color, linewidth=0.8)

        max_val = max(max(a_vals, default=1), max(u_vals, default=1))
        for i in range(len(staff_names)):
            if a_vals[i]:
                ax.text(a_vals[i] + max_val * 0.02, y[i] + h / 2,
                        f"{a_vals[i]:.1f}", va="center", fontsize=8, color=a_color)
            if u_vals[i]:
                ax.text(u_vals[i] + max_val * 0.02, y[i] - h / 2,
                        f"{u_vals[i]:.1f}", va="center", fontsize=8, color=u_color)

        ax.set_yticks(y)
        ax.set_yticklabels(staff_names)
        ax.set_xlabel("Hours")
        ax.set_title(title, fontsize=12, color=GRAY)
        ax.legend(loc="lower right", fontsize=8)
        ax.grid(axis="x", alpha=0.3)
        ax.invert_yaxis()

    _save_chart(fig, pdf, out_dir, "09_user_vs_ai_comparison")


def generate_pdf_charts(sessions: list[Session], out_dir: Path) -> None:
    """Generate all PDF + PNG charts from sessions data."""
    out_dir.mkdir(parents=True, exist_ok=True)
    _clean_generated_pngs(out_dir)
    pdf_path = out_dir / "ai_journal_charts.pdf"

    data = _build_chart_data(sessions)

    print(f"📈  Generating charts → {out_dir}/")
    with PdfPages(str(pdf_path)) as pdf:
        chart_sdlc_tasks(sessions, pdf, out_dir)
        chart_staff_effectiveness(data, pdf, out_dir)
        chart_kpi_summary(data, pdf, out_dir)
        chart_efficiency(data, pdf, out_dir)
        chart_rating(data, pdf, out_dir)
        chart_errors_from_sessions(sessions, pdf, out_dir)
        chart_user_vs_ai(sessions, pdf, out_dir)

    pngs = list(out_dir.glob("*.png"))
    print(f"   PDF: {pdf_path}")
    print(f"   PNGs: {len(pngs)} files in {out_dir}/")
    for p in sorted(pngs):
        print(f"     • {p.name}")


# =========================================================================== #
#  Section 12 — Terminal Summary
# =========================================================================== #

def print_terminal_summary(sessions: list[Session]) -> None:
    a = _agg(sessions)
    n_staff = len({s.staff for s in sessions})

    print("\n" + "=" * 64)
    print("  📊  AI DEV JOURNAL — SUMMARY")
    print("=" * 64)
    print(f"  Staff: {n_staff}    Sessions: {a['n']}    Rating TB: {a['avg_rating']}/5    5★: {a['excellent']}")
    print(f"  EST (without AI):  {a['est']:>8}h")
    print(f"  Actual (with AI):  {a['actual']:>8}h")
    print(f"  Time saved:        {a['saved']:>8}h")
    print(f"  Time Saved %:        {a['eff']:>7}%")
    print(f"  Avg saved/session: {a['avg_saved']:>8}h")

    by_staff: dict[str, list[Session]] = {}
    for s in sessions:
        by_staff.setdefault(s.staff, []).append(s)
    print("\n  ── Per Staff " + "─" * 49)
    print(f"  {'Staff':<14} {'#':>4} {'EST':>7} {'Actual':>7} {'Saved':>7} {'Eff%':>6} {'Rating':>6}")
    for staff in sorted(by_staff, key=lambda k: -sum(s.time_saved or 0 for s in by_staff[k])):
        sa = _agg(by_staff[staff])
        print(f"  {staff:<14} {sa['n']:>4} {sa['est']:>6}h {sa['actual']:>6}h {sa['saved']:>6}h {sa['eff']:>5}% {sa['avg_rating']:>6}")

    by_tool: dict[str, list[Session]] = {}
    for s in sessions:
        by_tool.setdefault(s.tool or "(unknown)", []).append(s)
    print("\n  ── Per Tool " + "─" * 50)
    print(f"  {'Tool':<22} {'#':>4} {'EST':>7} {'Actual':>7} {'Saved':>7} {'Eff%':>6}")
    for tool in sorted(by_tool, key=lambda k: -sum(s.time_saved or 0 for s in by_tool[k])):
        ta = _agg(by_tool[tool])
        print(f"  {tool:<22} {ta['n']:>4} {ta['est']:>6}h {ta['actual']:>6}h {ta['saved']:>6}h {ta['eff']:>5}%")

    by_cat: dict[str, list[Session]] = {}
    for s in sessions:
        by_cat.setdefault(s.category or "(uncategorized)", []).append(s)
    print("\n  ── Per Category " + "─" * 46)
    print(f"  {'Category':<22} {'#':>4} {'EST':>7} {'Actual':>7} {'Saved':>7} {'Eff%':>6}")
    for cat in sorted(by_cat, key=lambda k: -sum(s.time_saved or 0 for s in by_cat[k])):
        ca = _agg(by_cat[cat])
        print(f"  {cat:<22} {ca['n']:>4} {ca['est']:>6}h {ca['actual']:>6}h {ca['saved']:>6}h {ca['eff']:>5}%")

    print("=" * 64)


# =========================================================================== #
#  Section 13 — CLI & Main
# =========================================================================== #

def main() -> int:
    ap = argparse.ArgumentParser(
        description="AI Dev Journal — All-in-one report, charts, and PDF generator.",
    )
    ap.add_argument("files", nargs="+", type=Path, help="Input AI Dev Journal .xlsx files")
    ap.add_argument("-o", "--output", type=Path, default=Path("ai_journal_report.xlsx"),
                    help="Output .xlsx report (default: ai_journal_report.xlsx)")
    ap.add_argument("--charts-dir", type=Path, default=Path("charts_output"),
                    help="Output directory for PDF/PNG charts (default: charts_output)")
    ap.add_argument("--model", default="gpt-5.4-mini",
                    help="Model for OpenAI-compatible API (default: gpt-5.4-mini)")
    ap.add_argument("--no-ai", action="store_true",
                    help="Skip all AI inference (lessons + classification)")
    ap.add_argument("--no-lesson-ai", action="store_true",
                    help="Skip AI lesson inference only")
    ap.add_argument("--no-chart-ai", action="store_true",
                    help="Skip AI error/SDLC classification only")
    ap.add_argument("--skip-charts", action="store_true",
                    help="Skip adding Excel chart sheets")
    ap.add_argument("--skip-pdf", action="store_true",
                    help="Skip PDF/PNG chart generation")
    ap.add_argument("--profiles", type=Path, default=None,
                    help="YAML or JSON file with staff profiles for AI hour estimation")
    ap.add_argument("--no-estimate", action="store_true",
                    help="Skip AI hour estimation")
    ap.add_argument("--batch-size", type=int, default=20,
                    help="Sessions per AI classification batch (default: 20)")
    ap.add_argument("--timeout", type=int, default=300,
                    help="AI request timeout in seconds (default: 300)")
    args = ap.parse_args()

    # ── Phase 1: Parse input files ──
    all_sessions: list[Session] = []
    print(f"📂  Reading {len(args.files)} file(s)...")
    for f in args.files:
        if not f.exists():
            print(f"  ⚠  {f}: not found, skipping", file=sys.stderr)
            continue
        try:
            sess = parse_file(f)
            print(f"  ✔  {f.name} → staff='{staff_from_filename(f)}', {len(sess)} sessions")
            all_sessions.extend(sess)
        except Exception as e:
            print(f"  ✖  {f.name}: {e}", file=sys.stderr)

    if not all_sessions:
        print("No sessions parsed. Exiting.", file=sys.stderr)
        return 1

    # Assign row IDs for classification
    for i, s in enumerate(all_sessions):
        s.row_id = f"S{i + 1}"

    # ── Phase 0: Translate input data to English ──
    if not args.no_ai:
        try:
            translate_sessions_batch(all_sessions, args.model)
        except Exception as e:
            print(f"⚠  Translation failed: {e}", file=sys.stderr)
            print("   Continuing with original language.", file=sys.stderr)

    # ── Phase 0b: AI hour estimation ──
    if not args.no_ai and not args.no_estimate:
        profiles = _load_profiles(args.profiles)
        if not profiles and args.profiles:
            print(f"⚠  No profiles loaded from {args.profiles}", file=sys.stderr)
        try:
            estimate_hours_batch(all_sessions, args.model, profiles)
        except Exception as e:
            print(f"⚠  AI estimation failed: {e}", file=sys.stderr)
            print("   Continuing without AI estimates.", file=sys.stderr)

    # ── Phase 1 AI: Lesson inference ──
    with_lesson_ai = not args.no_ai and not args.no_lesson_ai
    if with_lesson_ai:
        try:
            infer_lessons_batch(all_sessions, args.model)
        except Exception as e:
            print(f"⚠  AI lesson inference failed: {e}", file=sys.stderr)
            print("   Continuing without AI lesson sheet.", file=sys.stderr)
            with_lesson_ai = False

    # ── Phase 1 output: Build report workbook ──
    wb = build_report(all_sessions, with_ai=with_lesson_ai)

    # ── Phase 2: Error + SDLC classification ──
    with_chart_ai = not args.no_ai and not args.no_chart_ai
    if not args.skip_charts:
        if with_chart_ai:
            try:
                classify_sessions(all_sessions, args.model,
                                  batch_size=args.batch_size, timeout=args.timeout)
            except Exception as e:
                print(f"⚠  AI classification failed: {e}", file=sys.stderr)
                print("   Adding chart sheets without classification.", file=sys.stderr)
                mark_unclassified(all_sessions)
        else:
            mark_unclassified(all_sessions)

        # ── Phase 2 output: Add chart sheets ──
        print("📊  Building chart sheets...")
        add_chart_sheets(wb, all_sessions)

    # ── Save workbook ──
    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"\n✔  Report saved to: {args.output}")

    # ── Phase 3: PDF/PNG charts ──
    if not args.skip_pdf:
        generate_pdf_charts(all_sessions, args.charts_dir)

    # ── Terminal summary ──
    print_terminal_summary(all_sessions)

    print(f"\n✔  Done!")
    return 0


if __name__ == "__main__":
    sys.exit(main())
